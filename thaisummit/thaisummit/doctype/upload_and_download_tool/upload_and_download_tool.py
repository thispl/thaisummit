# Copyright (c) 2023, TEAMPRO and contributors
# For license information, please see license.txt
import frappe
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook

from frappe.utils.csvutils import UnicodeWriter, read_csv_content

import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

class UploadandDownloadTool(Document):
	pass

@frappe.whitelist()
def download_pcs_part_master():
	filename = 'PCS Part Master'
	test = build_xlsx_response(filename)

# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	header = ["Item Code","Part No","Part Name","Vendor Code","Vendor Name","Model","Customer","Grade","Mat Type","Size Type","RM Length","RM Width","RM Thick","Strip Qty","Gross Weight","Net Weight","Process Cost","Admin Cost","Transport Cost",]
	ws.append(header)
	data = get_parts()
	for row in data:
		ws.append(row)
	for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=19):
		for cell in rows:
			cell.fill = PatternFill(fgColor='60add3', fill_type = "solid")
	bold_font = Font(bold=True,size=12)
	for cell in ws["1:1"]:
		cell.font = bold_font
	ws.sheet_view.zoomScale = 80
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file    

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def get_parts():
	data = []
	parts = frappe.get_all('PCS Part Master',['*'],order_by="item_code")
	for p in parts:
		data.append([p.item_code,p.parts_no,p.parts_name,p.vendor_code,p.vendor_name,p.model,p.customer,p.grade,p.mat_type,p.size_type,p.rm_length,p.rm_width,p.rm_thick,p.strip_qty,p.gross_weight,p.net_weight,p.process_cost,p.admin_cost,p.transport_cost])
	return data


# dowloading template 
@frappe.whitelist()
def download_template_pcs_part_master():
	filename = 'Template PCS Part Master'
	test = build_templatepart_xlsx_response(filename)

# return xlsx file object
def template_pcs_part_master_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	header = ["Item Code","Part No","Part Name","Vendor Code","Vendor Name","Model","Customer","Grade","Mat Type","Size Type","RM Length","RM Width","RM Thick","Strip Qty","Gross Weight","Net Weight","Process Cost","Admin Cost","Transport Cost",]
	ws.append(header)
	data = get_template_parts()
	for row in data:
		ws.append(row)
	for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=19):
		for cell in rows:
			cell.fill = PatternFill(fgColor='60add3', fill_type = "solid")
	bold_font = Font(bold=True,size=12)
	for cell in ws["1:1"]:
		cell.font = bold_font
	ws.sheet_view.zoomScale = 80
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file    

def build_templatepart_xlsx_response(filename):
	xlsx_file = template_pcs_part_master_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def get_template_parts():
	data = []
	parts = frappe.get_all('PCS Part Master',['*'],order_by="item_code")
	for p in parts:
		data.append(["","","","","","","","","","","","","","","","","","","",])
	return data


# download_pcs_scrap_master
@frappe.whitelist()
def download_pcs_scrap_master():
	filename = 'PCS Scrap Master'
	test = download_pcs_scrap_master_build_xlsx_response(filename)


# return xlsx file object
def download_pcs_scrap_master_make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.create_sheet(sheet_name, 0)
	header = ["Vendor Code","Vendor Name","IYM","RE",]
	ws.append(header)

	data = download_pcs_scrap_master_get_parts()

	for row in data:
		ws.append(row)

	for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=4):
		for cell in rows:
			cell.fill = PatternFill(fgColor='60add3', fill_type = "solid")

	bold_font = Font(bold=True,size=12)
	for cell in ws["1:1"]:
		cell.font = bold_font
	
	ws.sheet_view.zoomScale = 80

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file    


def download_pcs_scrap_master_build_xlsx_response(filename):
	xlsx_file = download_pcs_scrap_master_make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def download_pcs_scrap_master_get_parts():
	data = []
	parts = frappe.get_all('Scrap Master',['*'],order_by="vendor_code")
	for p in parts:
		data.append([p.vendor_code,p.vendor_name,p.iym,p.re])
	return data



@frappe.whitelist()
def download_template_pcs_scrap_master():
	filename = 'Template for PCS Scrap Master'
	test = download_template_pcs_scrap_master_build_xlsx_response(filename)


# return xlsx file object
def download_template_pcs_scrap_master_make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.create_sheet(sheet_name, 0)
	header = ["Vendor Code","Vendor Name","IYM","RE",]
	ws.append(header)

	data = download_template_pcs_scrap_master_get_parts()

	for row in data:
		ws.append(row)

	for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=4):
		for cell in rows:
			cell.fill = PatternFill(fgColor='60add3', fill_type = "solid")

	bold_font = Font(bold=True,size=12)
	for cell in ws["1:1"]:
		cell.font = bold_font
	
	ws.sheet_view.zoomScale = 80

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file    


def download_template_pcs_scrap_master_build_xlsx_response(filename):
	xlsx_file = download_template_pcs_scrap_master_make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def download_template_pcs_scrap_master_get_parts():
	data = []
	parts = frappe.get_all('Scrap Master',['*'],order_by="vendor_code")
	for p in parts:
		data.append(["","","","",])
	return data


@frappe.whitelist()
def enqueue_scrap_master(file):
	enqueue(attach_scrap_master, queue='default', timeout=6000, event='attach_scrap_master',file=file)
	return 'Succuss'

@frappe.whitelist()
def attach_scrap_master(file):
	
	# frappe.db.sql("""delete from `tabScrap Master`""")
	# frappe.log_error(title="Scrap Master Delete",message="Scrap Master Deleted By %s" % frappe.session.user)
	filepath = get_file(file)
	pps = read_csv_content(filepath[1])
	for p in pps[1:]:
		frappe.errprint(p)
		s_master = frappe.new_doc("Scrap Master")
		s_master.vendor_code = p[0]
		s_master.vendor_name = p[1]
		s_master.iym = p[2]
		s_master.re = p[3]
		s_master.save(ignore_permissions=True)
		frappe.db.commit()

@frappe.whitelist()
def enqueue_pcs_part_master(attach_file):
	enqueue(attach_pcs_part_master, queue='default', timeout=6000, event='attach_pcs_part_master',file = attach_file)
	return 'Succuss'

@frappe.whitelist()
def attach_pcs_part_master(attach_file):
	# frappe.db.sql("""delete from `tabScrap Master`""")
	# frappe.log_error(title="Scrap Master Delete",message="Scrap Master Deleted By %s" % frappe.session.user)
	filepath = get_file(attach_file)
	pps = read_csv_content(filepath[1])
	for p in pps[1:]:
		p_master = frappe.new_doc("PCS Part Master")
		p_master.item_code = p[0]
		p_master.parts_no = p[1]
		p_master.parts_name = p[2]
		p_master.vendor_code = p[3]
		p_master.vendor_name = p[4]
		p_master.model = p[5]
		p_master.customer = p[6]
		p_master.grade = p[7]
		p_master.mat_type = p[8]
		p_master.size_type = p[9]
		p_master.rm_length = p[10]
		p_master.rm_width = p[11]
		p_master.rm_thick = p[12]
		p_master.strip_qty = p[13]
		p_master.gross_weight = p[14]
		p_master.net_weight = p[15]
		p_master.process_cost = p[16]
		p_master.admin_cost = p[17]
		p_master.transport_cost = p[18]
		p_master.save(ignore_permissions=True)
		frappe.db.commit()