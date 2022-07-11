# Copyright (c) 2022, TEAMPRO and contributors
# For license information, please see license.txt

from email import message
import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (
    flt,
    cint,
    cstr,
    get_html_format,
    get_url_to_form,
    gzip_decompress,
    format_duration,
)
from frappe.model.document import Document


class InvoiceGRNDetail(Document):
    pass


@frappe.whitelist()
def download():
    filename = 'INVOICE GRN DETAIL'
    test = build_xlsx_response(filename)


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header = ["s.no", "Supplier Code", "Suppplier Name", "Invoice No", "Status",  "Invoice Date", "Mat No", "Part No",
              "Part Name", "QTY", "Rate", "Basic Amount", "Invoice Amount", "GRN Date", "GRN No", "GRN QTY", ]
    ws.append(header)
    if args.supplier:
        invoices = frappe.get_all("TSAI Invoice", {"invoice_date": (
            "between", (args.from_date, args.to_date)), 'supplier_code': args.supplier}, ['*'])
    else:
        invoices = frappe.get_all("TSAI Invoice", {"invoice_date": (
            "between", (args.from_date, args.to_date))}, ['*'])
    i = 1
    for inv in invoices:
        invoice_items = frappe.get_all(
            "Invoice Items", {"parent": inv.name}, ["*"])

        for item in invoice_items:
            if inv.status in ('OPEN', 'CLOSED'):
                status = '-'
            else:
                status = inv.status

            data = [i, inv.supplier_code, inv.supplier_name, inv.name, status, inv.invoice_date, item.mat_no, item.parts_no, item.parts_name,
                    item.key_qty, item.unit_price, item.basic_amount, item.invoice_amount, item.grn_date, item.grn_no, item.grn_qty]
            i += 1
            ws.append(data)

    bold_font = Font(bold=True, size=12)
    for cell in ws["1:1"]:
        cell.font = bold_font

    ws.sheet_view.zoomScale = 80

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file
