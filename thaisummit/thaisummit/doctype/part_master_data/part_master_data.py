# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

class PartMasterData(Document):
    @frappe.whitelist()
    def get_data(self):
        data = """<table class=table table-bordered >
        <tr><td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>S.No</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Mat No</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Part No</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Part Name</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Production Line</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Model</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Customer</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>MRP Sales Price</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>MRP Purchase Price</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Kanban Group</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Mat Type</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Operator</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Manual Welder</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Spater Cleaner</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Inspector</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Final Inspector</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Buffing and Rework</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Retap</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Boring</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Manpower Std</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Packing Std</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Cycle Time</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>UPH</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Min Day</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Max Day</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Parts Weight</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Scrap Weight</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>MRP Daily Order</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Transfer Rate</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Rack No</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Rack Resposible Person</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>After Spot Nut</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>Machine</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>FG MAT</center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>FG NAME</b></center></td>
        <td style="background-color:#ffedcc; padding:1px; border: 1px solid black; font-size:7.5px;"><center><b>NDOL</b></center></td>




        </tr>"""
        parts = frappe.get_all('TSAI Part Master',['*'],order_by="mat_no")
        i = 1
        for p in parts:
            data += """
            <tr>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:7.5px;"><center>%s</center></td>
            </tr>
            """%(i,p.mat_no,p.parts_no or '',p.parts_name or '',p.production_line or '',p.model or '',p.customer or '',p.mrp_sales_price,p.mrp_purchase_price,p.kanban_group or '',p.mat_type or '', p.operator,p.manual_welder,p.spater_cleaner,p.inspector,p.final_inspector,p.buffing_and_rework,p.retap,p.boring,
                        p.manpower_std,p.packing_std,p.cycle_time,p.uph,p.min_day,p.max_day,p.parts_weight,p.scrap_weight,p.mrp_daily_order,p.transfer_rate,p.rack_no or '',p.rack_resposible_person or '',p.after_spot_nut,p.machine,p.fg_mat,p.fg_name,p.ndol
                   )
            i += 1
        return data

@frappe.whitelist()
def download():
    filename = 'Part Master Data'
    test = build_xlsx_response(filename)


# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header = ["Mat No","Part No","Part Name","Production Line","Model","Customer","MRP Sales Price","MRP Purchase Price","Kanban Group","Mat Type","Operator","Manual Welder","Spater Cleaner","Inspector","Final Inspector","Buffing and Rework","Retap","Boring","Manpower Std","Packing Std","Cycle Time","UPH","Min Day","Max Day","Parts Weight","Scrap Weight","MRP Daily Order",
                "Transfer Rate","Rack No","Rack Responsible Person","After Spot Nut","Machine","FG MAT","FG NAME","NDOL"]
    ws.append(header)

    data = get_parts()

    for row in data:
        ws.append(row)

    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=34):
        for cell in rows:
            cell.fill = PatternFill(fgColor='60add3', fill_type = "solid")

    bold_font = Font(bold=True,size=12)
    for cell in ws["1:1"]:
        cell.font = bold_font
    
    ws.sheet_view.zoomScale = 80

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file    


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def get_parts():
    data = []
    parts = frappe.get_all('TSAI Part Master',['*'],order_by="mat_no")
    for p in parts:
        data.append([p.mat_no,p.parts_no,p.parts_name,p.production_line,p.model,p.customer,p.mrp_sales_price,p.mrp_purchase_price,p.kanban_group,p.mat_type,p.operator,p.manual_welder,p.spater_cleaner,p.inspector,p.final_inspector,p.buffing_and_rework,p.retap,p.boring,p.manpower_std,p.packing_std,p.cycle_time,p.uph,p.min_day,p.max_day,p.parts_weight,p.scrap_weight,p.mrp_daily_order,p.transfer_rate,p.rack_no or '-',
                    p.rack_resposible_person or '-',p.after_spot_nut or '-',p.machine or '-',p.fg_mat or '-',p.fg_name or '-',p.ndol or '-',
                    ])
    return data


@frappe.whitelist()
def enqueue_upload(file):
    enqueue(upload, queue='default', timeout=6000, event='upload',file=file)


@frappe.whitelist()
def upload(file):
    file = get_file(file)
    pps = read_xlsx_file_from_attached_file(fcontent=file[1])
    for pp in pps:
        pm = frappe.db.exists('TSAI Part Master',pp[0])
        if pm:
            doc = frappe.get_doc('TSAI Part Master',pp[0])
            doc.parts_no = pp[1]
            doc.parts_name = pp[2]
            doc.production_line = pp[3]
            doc.model = pp[4]
            doc.customer = pp[5]
            doc.mrp_sales_price = pp[6]
            doc.mrp_purchase_price = pp[7]
            doc.kanban_group = pp[8]
            doc.mat_type = pp[9]
            doc.operator = pp[10]
            doc.manual_welder = pp[11]
            doc.spater_cleaner = pp[12]
            doc.inspector = pp[13]   
            doc.final_inspector = pp[14]
            doc.buffing_and_rework = pp[15]
            doc.retap = pp[16] 
            doc.boring = pp[17]  
            doc.manpower_std = pp[18]
            doc.packing_std = pp[19]
            doc.cycle_time = pp[20]
            doc.uph = pp[21]
            doc.min_day = pp[22]
            doc.max_day = pp[23]
            doc.parts_weight = pp[24]
            doc.scrap_weight = pp[25]
            doc.mrp_daily_order = pp[26]
            doc.transfer_rate = pp[27]
            doc.rack_no = pp[28]
            doc.rack_resposible_person = pp[29]
            doc.after_spot_nut = pp[30]
            doc.machine = pp[31]   
            doc.fg_mat = pp[32]
            doc.fg_name = pp[33]
            doc.ndol = pp[34]
            doc.save(ignore_permissions=True)
            frappe.db.commit()
        else:
            doc = frappe.new_doc('TSAI Part Master')
            doc.mat_no = pp[0]
            doc.parts_no = pp[1]
            doc.parts_name = pp[2]
            doc.production_line = pp[3]
            doc.model = pp[4]
            doc.customer = pp[5]
            doc.mrp_sales_price = pp[6]
            doc.mrp_purchase_price = pp[7]
            doc.kanban_group = pp[8]
            doc.mat_type = pp[9]
            doc.operator = pp[10]
            doc.manual_welder = pp[11]
            doc.spater_cleaner = pp[12]
            doc.inspector = pp[13]   
            doc.final_inspector = pp[14]
            doc.buffing_and_rework = pp[15]
            doc.retap = pp[16] 
            doc.boring = pp[17]  
            doc.manpower_std = pp[18]
            doc.packing_std = pp[19]
            doc.cycle_time = pp[20]
            doc.uph = pp[21]
            doc.min_day = pp[22]
            doc.max_day = pp[23]
            doc.parts_weight = pp[24]
            doc.scrap_weight = pp[25]
            doc.mrp_daily_order = pp[26]
            doc.transfer_rate = pp[27]
            doc.rack_no = pp[28]
            doc.rack_resposible_person = pp[29]
            doc.after_spot_nut = pp[30]
            doc.machine = pp[31]   
            doc.fg_mat = pp[32]
            doc.fg_name = pp[33] 
            doc.ndol = pp[34]
            doc.save(ignore_permissions=True)
            frappe.db.commit()
    frappe.log_error('Part Master Uploaded successfully')