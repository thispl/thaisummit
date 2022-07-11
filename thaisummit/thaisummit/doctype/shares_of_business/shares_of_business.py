# Copyright (c) 2022, TEAMPRO and contributors
# For license information, please see license.txt


import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (
    flt,
    cint,
    cstr,
    get_html_format,
    get_url_to_form,
    gzip_decompress,
    format_duration,
)


class SharesofBusiness(Document):
    @frappe.whitelist()
    def get_data(self):
        data = """<table class='table table-bordered=1'>
		<tr>
		   <td style="background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;">
                 <center><b>S no.</b></center>
		  <td style="background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;">
                 <center><b>Supplier Code</b></center>
		  <td style="background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;">
                 <center><b>Supplier Name</b></center>
		  <td style="background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;">
                 <center><b>Mat No</b></center>
		  <td style="background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;">
                 <center><b>Part Name</b></center>
		  <td style="background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;">
                 <center><b>Share Of Business</b></center>
		</tr>"""
        sbe = frappe.get_all("Shares of Business Entry",'mat_no')
        mats = []
        for s in sbe:
            if s.mat_no not in mats:
                mats.append(s.mat_no)
        i = 1
        for mat in mats:
            d = 0
            shares = frappe.get_all('Shares of Business Entry',{'mat_no':mat},['*'])
            for s in shares:
                if d == 0:
                    data += """
                    <tr>
                    <td rowspan='2' style="padding:1px; border: 1px solid black; font-size:10px;"><center>%s</center></td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>%s</center></td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>%s</center></td>
                    </tr>
                    """ % (i,s.supplier_code, s.supplier_name, s.mat_no, s.part_name, s.share_of_business)
                else:
                    data += """
                    <tr>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>%s</center></td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>%s</center></td>
                    </tr>
                    """ % (s.supplier_code, s.supplier_name, s.mat_no, s.part_name, s.share_of_business)
                d += 1
            i += 1
        return data
@frappe.whitelist()
def download():
    filename = 'Share Of'
    test = build_xlsx_response(filename)


# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header = ["s.no","Supplier Code","Supplier Name","Mat No","Part Name","Share of Business"]
    ws.append(header)

    data = get_share()

    for row in data:
        ws.append(row)

    # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=18):
    #     for cell in rows:
    #         cell.fill = PatternFill( fill_type = "solid")

    bold_font = Font(bold=True,size=12)
    for cell in ws["1:1"]:
        cell.font = bold_font
    
    ws.sheet_view.zoomScale = 80

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file    


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def get_share():
    sbe = frappe.get_all("Shares of Business Entry",'mat_no')
    data=[]
    mats = []
    for s in sbe:
        
        if s.mat_no not in mats:
            mats.append(s.mat_no)
    i =1
    for mat in mats:

        shares = frappe.get_all('Shares of Business Entry',{'mat_no':mat},['*'])
        for s in shares:
            
            
            data.append([i,s.supplier_code, s.supplier_name, s.mat_no, s.part_name, s.share_of_business])

            
        i +=1
    
    return data

@frappe.whitelist()
def upload(file):
    file = get_file(file)
    pps = read_xlsx_file_from_attached_file(fcontent=file[1])
    for pp in pps:
        if pp[1] != "Supplier Code":
            pm = frappe.db.exists('Shares of Business Entry',{'supplier_code':pp[1],'mat_no':pp[3]})
            if pm:
                doc = frappe.get_doc('Shares of Business Entry',pm)
                doc.share_of_business = pp[5]
                doc.save(ignore_permissions=True)
                frappe.db.commit()
            else:
                doc = frappe.new_doc("Shares of Business Entry")
                doc.supplier_code = pp[1]
                doc.supplier_name= pp[2]
                doc.mat_no = pp[3]
                doc.part_name = pp[4]
                doc.share_of_business = pp[5]
                doc.save(ignore_permissions=True)
                frappe.db.commit()
    return 'ok'

def test_method():
    sbe = frappe.get_all("Shares of Business Entry",'mat_no')
    mats = []
    for s in sbe:
        if s.mat_no not in mats:
            mats.append(s.mat_no)
    for mat in mats:
        shares = frappe.get_all('Shares of Business Entry',{'mat_no':mat},['*'])
        for share in shares:
            print([mat,share.share_of_business])