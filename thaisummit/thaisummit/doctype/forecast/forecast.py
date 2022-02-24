# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt
import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe.utils.background_jobs import enqueue
import datetime
from datetime import datetime


from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

class Forecast(Document):
    @frappe.whitelist()
    def get_data(self):
        data = """<table class=table table-bordered>
        <tr><td style="background-color:#fafa13; padding:1px; border: 1px solid black; font-size:10px;"><center><b>S.No</b></center></td>
        <td style="background-color:#fafa13; padding:1px; border: 1px solid black; font-size:10px;"><center><b>MAT NO</b></center></td>
        <td style="background-color:#fafa13; padding:1px; border: 1px solid black; font-size:12px;"><center><b>PART NO</b></center></td>
        <td style="background-color:#fafa13; padding:1px; border: 1px solid black; font-size:12px;"><center><b>PART NAME</b></center></td>
        <td style="background-color:#fafa13; padding:1px; border: 1px solid black; font-size:12px;"><center><b>CUS</b></center></td>"""

        no_of_days = date_diff(add_days(self.to_date, 1), self.from_date)
        dates = [add_days(self.from_date, i) for i in range(0, no_of_days)]

        for date in dates:
            date = datetime.strptime(str(date),'%Y-%m-%d')
            d = date.strftime('%d')
            m = date.strftime('%b')
            data += '<td style="background-color:#fafa13; padding:1px; border: 1px solid black; font-size:12px;"><center><b>%s %s</b></center></td>'%(d,m)
        data += "</tr>"
        parts = frappe.get_all('TSAI Part Master',{'mat_type':'FG'},['*'],order_by="mat_no")
        # parts = frappe.db.sql("select * from `tabForecast Data` where date between '%s' and '%s' group by mat_no order by mat_no"%(self.from_date,self.to_date),as_dict=True)
        i = 1
        total_qty = 0
        for p in parts:
            data += """
            <tr>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            """%(i,p.mat_no,p.parts_no,p.parts_name,p.customer)
            # fds = frappe.db.sql("select qty from `tabForecast Data` where mat_no = %s and date between '%s' and '%s' order by date"%(p.mat_no,self.from_date,self.to_date),as_dict=True)
            # for fd in fds:
            #     data += '<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>'%fd.qty or '-'
            for date in dates:
                q = frappe.db.get_value('Forecast Data',{'mat_no':p.mat_no,'date':date},'qty') or '-'
                data += '<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>'%q
                if q != '-':
                    total_qty += q
            data += '</tr>'
            i += 1
        data += '</table>'
        summary = """<table class=table table-bordered>
        <tr><td style="background-color:#fafa13; border: 1px solid black; font-size:15px;"><center><b>Total Mat Nos</b></center></td>
        <td style="border: 1px solid black; font-size:15px;"><b>%s</b></td><tr>
        <tr><td style="background-color:#fafa13; border: 1px solid black; font-size:15px;"><center><b>Total Qty</b></center></td>
        <td style="border: 1px solid black; font-size:15px;"><b>%s</b></td><tr>
        """%(len(parts),total_qty)
        return data, summary


@frappe.whitelist()
def download():
    filename = 'Forecast'
    test = build_xlsx_response(filename)


# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header = ["MAT NO","PART N0","PART NAME","CUSTOMER"]
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    for date in dates:
        date = datetime.strptime(str(date),'%Y-%m-%d')
        d = date.strftime('%d')
        m = date.strftime('%b')
        header.append(d+ '-' +m)
    ws.append(header)

    data = get_parts(dates)

    for row in data:
        ws.append(row)

    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(dates)+4):
        for cell in rows:
            cell.fill = PatternFill(fgColor='fafa13', fill_type = "solid")

    bold_font = Font(bold=True,size=12)
    for cell in ws["1:1"]:
        cell.font = bold_font
    
    ws.sheet_view.zoomScale = 80

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file    


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def get_parts(dates):
    data = []
    parts = frappe.get_all('TSAI Part Master',{'mat_type':'FG'},['*'],order_by="mat_no")
    for p in parts:
        row = [p.mat_no,p.parts_no,p.parts_name,p.customer]
        for date in dates:
            q = frappe.db.get_value('Forecast Data',{'mat_no':p.mat_no,'date':date},'qty') or '-'
            row.append(q)
        data.append(row)
    return data

@frappe.whitelist()
def enqueue_upload(file,from_date,to_date):
    enqueue(upload, queue='default', timeout=6000, event='upload',
    file=file,from_date=from_date,to_date=to_date)

@frappe.whitelist()
def upload(file,from_date,to_date):
    file = get_file(file)
    pps = read_xlsx_file_from_attached_file(fcontent=file[1])
    for pp in pps:
        if pp[0] != 'MAT NO':
            no_of_days = date_diff(add_days(to_date, 1), from_date)
            dates = [add_days(from_date, i) for i in range(0, no_of_days)]
            i = 4
            for date in dates:
                fd = frappe.db.exists('Forecast Data',{'mat_no':pp[0],'date':date})
                if fd:
                    if pp[i] == '-':
                        frappe.db.set_value('Forecast Data',fd,'qty',0)
                    elif not pp[i]:
                        frappe.db.set_value('Forecast Data',fd,'qty',0)
                    else:
                        frappe.db.set_value('Forecast Data',fd,'qty',pp[i])
                else:
                    if pp[i] not in ('-',0):
                        doc = frappe.new_doc('Forecast Data')
                        doc.mat_no = pp[0]
                        doc.part_no = pp[1]
                        doc.part_name = pp[2]
                        doc.customer = pp[3]
                        doc.date = date
                        doc.qty = pp[i]
                        doc.save(ignore_permissions=True)
                        frappe.db.commit()
                i += 1
    return 'ok'