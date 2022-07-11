# Copyright (c) 2022, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import (
    flt,
    cint,
    cstr,
    get_html_format,
    get_url_to_form,
    gzip_decompress,
    format_duration,
)

class FGREQIYM(Document):
   @frappe.whitelist()
   def get_data(self):
        data = """<table class='table table-bordered=1'>
            <tr>
               <td style="background-color:#137FF1; padding:2px; color:white;border: 1px solid black; font-size:15px;">
                  <center><b>S.No</b></center></td>
               <td style="background-color:#137FF1; padding:2px;color:white; border: 1px solid black; font-size:15px;">
                  <center><b>Mat No</b></center></td>
               <td style="background-color:#137FF1; padding:2px; color:white;border: 1px solid black; font-size:15px;">
                  <center><b>Part No</b></center></td>
               <td style="background-color:#137FF1; padding:2px;color:white; border: 1px solid black; font-size:15px;">
                  <center><b>Part Name</b></center></td>
               <td style="background-color:#137FF1;color:white; padding:2px; border: 1px solid black; font-size:15px;">
                  <center><b>Model</b></center></td>
               <td style="background-color:#137FF1; padding:2px; color:white;border: 1px solid black; font-size:15px;">
                  <center><b>Packing</b></center></td>
               <td style="background-color:#137FF1; padding:2px; color:white;border: 1px solid black; font-size:15px;">
                  <center><b>PROD Plan</b></center></td>
               <td style="background-color:#137FF1; padding:2px;color:white; border: 1px solid black; font-size:15px;">
                  <center><b>Tranfer Day</b></center></td>
               <td style="background-color:#137FF1; padding:2px;color:white; border: 1px solid black; font-size:15px;">
                  <center><b>Tranfer Plan</b></center></td>
               <td style="background-color:#137FF1; padding:2px;color:white; border: 1px solid black; font-size:15px;">
                  <center><b>Stock</b></center></td>
               <td style="background-color:#137FF1; padding:2px;color:white; border: 1px solid black; font-size:15px;">
                  <center><b>Coverage Day</b></center></td>
               <td style="background-color:#137FF1; padding:2px;color:white; border: 1px solid black; font-size:15px;">
                  <center><b>Transfer Plan (balance)</b></center></td>	
               <td style="background-color:#137FF1; padding:2px;color:white; border: 1px solid black; font-size:15px;">
                  <center><b>Transfer Actual</b></center></td>
               <td style="background-color:#137FF1; padding:2px;color:white; border: 1px solid black; font-size:15px;">
                  <center><b>Diff</b></center></td> 
               <td style="background-color:#137FF1; padding:2px;color:white; border: 1px solid black; font-size:15px;">
                  <center><b>%</b></center></td>
            </tr>"""


        data += """
                    <tr>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>-</center></td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">-</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">-</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>-</center></td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">-</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>-</center></td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>-</center></td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">-</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">-</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>-</center></td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">-</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>-</center></td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>-</center></td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;">-</td>
                    <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>-</center></td>
                    </tr>
                    """ 
        return data
