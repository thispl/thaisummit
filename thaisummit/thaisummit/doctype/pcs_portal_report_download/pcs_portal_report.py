# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import pandas as pd
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds



import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def download():
	filename = 'PCS Portal Report'
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	ws.append(["Customer: "+ args.customer + "","","","","","","","","","","","","","","","","","","","",""])
	ws.append(["Customer","Vendor Code","Vendor Name","Item Code","Part No","Item Description","Model","Grade","MAT Type","Size Type","RM Length","RM Width","RM Thick","Strip Qty","Gross Weight",'Net Weight',"Scrap Weight","RM Price","Gross WT Cost","Scrap Cost/kg","Scrap Cost","RM Cost","Process Cost","Admin Cost","Transport Cost","Final Part Cost"])
	
	pcs_part_master = frappe.db.sql("""select * from `tabPCS Part Master` where customer = '%s' """%(args.customer),as_dict=1)
	# salary_slips = frappe.get_all("Salary Slip",{'start_date':args.from_date,'end_date':args.to_date,"branch":args.branch,"contractor":args.contractor},['*']) 
	i=1    
	scrap_weight = 0
	scrap_cost = 0
	final_part_cost = 0
	price = 0
	rm_cost = 0
	gross_wt_cost = 0
	for p in pcs_part_master:
		rm_price_iym = frappe.get_single('RM Input').rm_input_table1
		rm_price_re = frappe.get_single('RM Input').rm_input_table2
		if p.customer == "IYM":
			for r in rm_price_iym:
				if p.grade == r.grade:
					if p.size_type == ">100":
						price = r.new1
						# frappe.log_error(title='rm_input',message=r.std_old)
					elif p.size_type == "<100":
						price = r.new2
					else:
						price = 0
		elif p.customer == "RE":
			for r in rm_price_re:
				if p.grade == r.grade:
					if p.mat_type == 'STRIP':
						if p.size_type == ">100":
							price = r.new1
							# frappe.log_error(title='rm_input',message=r.std_old)
						elif p.size_type == "<100":
							price = r.new2
						else:
							price = 0
					elif p.mat_type == 'COIL':
						if p.size_type == ">100":
							price = r.coil_new
							# frappe.log_error(title='rm_input',message=r.std_old)
						elif p.size_type == "<100":
							price = r.coil_new1
						else:
							price = 0
		scrap_master = frappe.db.sql("""select iym from `tabScrap Master` where vendor_code = '%s' """%(p.vendor_code),as_dict=1)
		for s in scrap_master:
			gross_weight = float(p.gross_weight)
			net_weight = float(p.net_weight)
			gross_wt_cost = round((gross_weight * price),3)
			scrap_weight = round((gross_weight - net_weight),3)
			scrap_cost = round((scrap_weight * float(s.iym)),3)
			rm_cost = round((gross_wt_cost - scrap_cost),3)
			final_part_cost = round((float(p.process_cost) + float(p.admin_cost) + float(rm_cost) + float(p.transport_cost)),3)
			ws.append([p.customer,p.vendor_code,p.vendor_name,p.item_code,p.part_no,p.parts_name,p.model,p.grade,p.mat_type,p.size_type,p.rm_length,p.rm__width,p.rm_thick,p.strip_qty,round(float(p.gross_weight),3),p.net_weight,scrap_weight,price,gross_wt_cost,s.iym,scrap_cost,rm_cost,round(float(p.process_cost),3),round(float(p.admin_cost),3),round(float(p.transport_cost),3),final_part_cost])        
		i=1+i
   
		ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=26)
		# ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=19)
		# ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=19)
		# ws.merge_cells(start_row=4,start_column=1,end_row=4,end_column=19)

		# ws.merge_cells(start_row=8,start_column=7,end_row=8,end_column=15)
		# ws.merge_cells(start_row=8,start_column=16,end_row=8,end_column=23)

		bold_font = Font(bold=True)
		for cell in ws["1:1"]:
			cell.font = bold_font
		for cell in ws["2:2"]:
			cell.font = bold_font
		# for cell in ws["3:3"]:
		# 	cell.font = bold_font
		# for cell in ws["4:4"]:
		# 	cell.font = bold_font
		# for cell in ws["5:5"]:
		# 	cell.font = bold_font

		for rows in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=26):
			for cell in rows:
				cell.fill = PatternFill(fgColor="ffdd99", fill_type = "solid")

		# for rows in ws.iter_rows(min_row=5, max_row=5, min_col=1, max_col=19):
		# 	for cell in rows:
		# 		cell.fill = PatternFill(fgColor="ff3333", fill_type = "solid")

		for rows in ws.iter_rows(min_row=3,max_col=26):
			for cell in rows:
				cell.fill = PatternFill(fgColor="fff7e6", fill_type = "solid")
		
		border = Border(left=Side(border_style='thin', color='000000'),
		right=Side(border_style='thin', color='000000'),
		top=Side(border_style='thin', color='000000'),
		bottom=Side(border_style='thin', color='000000'))
	for rows in ws.iter_rows(min_row=1, min_col=1, max_col=26):

		for cell in rows:
			cell.border = border
		

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'