# Copyright (c) 2023, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import frappe
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook

from frappe.utils.csvutils import UnicodeWriter, read_csv_content

import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

class UpdateTagCardFlowMasterinPartMaster(Document):
	pass



@frappe.whitelist()
def download_template():
	filename = 'Template for Tag Card Flow Master'
	test = download_template_tsai_part_master_build_xlsx_response(filename)


# return xlsx file object
def download_template_tsai_part_master_make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.create_sheet(sheet_name, 0)
	header = ["Mat No","Tad Card Flow Master",]
	ws.append(header)

	data = download_template_part_master()

	for row in data:
		ws.append(row)

	for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
		for cell in rows:
			cell.fill = PatternFill(fgColor='60add3', fill_type = "solid")

	bold_font = Font(bold=True,size=12)
	for cell in ws["1:1"]:
		cell.font = bold_font
	
	ws.sheet_view.zoomScale = 80

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file    


def download_template_tsai_part_master_build_xlsx_response(filename):
	xlsx_file = download_template_tsai_part_master_make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def download_template_part_master():
	data = []
	parts = frappe.get_all('TSAI Part Master',['*'])
	for p in parts:
		data.append(["",""])
	return data



@frappe.whitelist()
def enqueue_tag_card_flow_upload(file):
	enqueue(tag_card_flow_upload, queue='default', timeout=6000, event='upload',file=file)
	return 'ok'


@frappe.whitelist()
def tag_card_flow_upload(file):
	file = get_file(file)
	pps = read_xlsx_file_from_attached_file(fcontent=file[1])
	for pp in pps[1:]:
		pm = frappe.db.exists('TSAI Part Master',pp[0])
		if pm:
			doc = frappe.get_doc('TSAI Part Master',pp[0])
			doc.tag_card_flow_master = pp[1]
			doc.save(ignore_permissions=True)
			frappe.db.commit()
	return 'ok'