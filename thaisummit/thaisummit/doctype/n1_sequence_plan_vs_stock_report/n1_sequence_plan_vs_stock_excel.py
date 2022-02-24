from __future__ import unicode_literals
from os import stat
import frappe
from frappe.utils import cstr, add_days, date_diff,format_datetime,format_date
from frappe import _
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from frappe.utils import today

from datetime import date, timedelta, datetime, time

import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from redis import from_url
from six import BytesIO, string_types

@frappe.whitelist()
def download():
    filename = 'N-1 Sequence Plan vs Stock Report'
    build_xlsx_response(filename)

@frappe.whitelist()
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header = add_header_title()
    ws.append(header)
    if args.date:
        from_date = args.date
    else:
        from_date = today()
    # from_date = '2021-08-20'
    to_date = add_days(from_date,1)

    no_of_days = date_diff(add_days(to_date, 1), from_date)
    dates = [add_days(from_date, i) for i in range(0, no_of_days)]

    header = add_header(dates)
    ws.append(header)


    data = get_data(from_date,to_date)

    for row in data:
        ws.append(row)

    ws.sheet_view.zoomScale = 80

    ws.freeze_panes = 'D3'

    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(row=1,column=1).alignment = Alignment(horizontal='center')

    ws['A1'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")
    ws['E1'].fill = PatternFill(fgColor="FA8072", fill_type = "solid")
    ws['H1'].fill = PatternFill(fgColor="52c132", fill_type = "solid")

    for cell in ws["2:2"]:
        cell.fill = PatternFill(fgColor="52c132", fill_type = "solid")

    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)
    ws.cell(row=1,column=5).alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=12)
    ws.cell(row=1,column=8).alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=1, start_column=13, end_row=1, end_column=15)
    ws.cell(row=1,column=13).alignment = Alignment(horizontal='center')

    # ws.merge_cells(start_row=1, start_column=17, end_row=1, end_column=18)
    # ws.cell(row=1,column=17).alignment = Alignment(horizontal='center')

    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
    master = frappe.db.sql("""select DISTINCT mat_no, part_no,part_name from `tabTSA Master` where date between '%s' and '%s' """%(from_date,to_date),as_dict=True)
    for rows in ws.iter_rows(min_row=1, max_row=len(master)+2, min_col=1, max_col=15):
        for cell in rows:
            cell.border = border

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def add_header_title():
    header = ["N-1 SEQUENCE PLAN VS STOCK REPORT",'','','',"DELIVERY PLAN",'','','STOCK','','','','',format_date(today()),'','']
    return header

def add_header(dates):
    d1 = datetime.strptime(dates[0],'%Y-%m-%d')
    d1 = d1.strftime('%d/%m')
    d2 = datetime.strptime(dates[1],'%Y-%m-%d')
    d2 = d2.strftime('%d/%m')
    header = ["MAT NO",'PART NO','PART NAME','DAILY ORDER',d1,d2,'TOTAL PLAN','FG','WELD','HPS','TOTAL STOCK','COVERAGE','SHORT','PROD PLAN','STATUS']
    return header

def get_data(from_date,to_date):
    data = []
    no_of_days = date_diff(add_days(to_date, 1), from_date)
    dates = [add_days(from_date, i) for i in range(0, no_of_days)]
    master = frappe.db.sql("""select DISTINCT mat_no, part_no,part_name from `tabTSA Master` where date between '%s' and '%s' """%(from_date,to_date),as_dict=True)
    for m in master:
        daily_order = frappe.db.sql("""select sum(quantity) as total from `tabTSA Master` where mat_no = '%s' and date between '%s' and '%s' """%(m.mat_no,from_date,to_date),as_dict=True)[0].total or 0
        row = [m.mat_no,m.part_no,m.part_name]
        days = 0
        qty_list = []
        for date in dates:
            qty = frappe.db.get_value('TSA Master',{'mat_no':m.mat_no,'date':date},'quantity') or '-'
            if qty != '-':
                days += 1
            qty_list.append(qty)
        if daily_order > 0:
            row.append(daily_order/days)
        else:
            row.append(0)
        row.extend(qty_list)
        row.append(daily_order)
        # fg = frappe.db.get_value('SAP Outgoing Report',{'part_no':m.mat_no,'wh':'FG'},'quantity') or '-'
        # weld = frappe.db.get_value('SAP Outgoing Report',{'part_no':m.mat_no,'wh':'Weld'},'quantity') or '-'
        # hps = frappe.db.get_value('SAP Outgoing Report',{'part_no':m.mat_no,'wh':'HPS'},'quantity') or '-'
        fg = frappe.db.get_value('TSAI Stock',{'item_no':m.mat_no,'wh':'FG'},'qty') or '-'
        weld = frappe.db.get_value('TSAI Stock',{'item_no':m.mat_no,'wh':'Weld'},'qty') or '-'
        hps = frappe.db.get_value('TSAI Stock',{'item_no':m.mat_no,'wh':'HPS'},'qty') or '-'
        total_stock = 0
        fg_stock = 0
        if type(fg) == int:
            total_stock += fg
            fg_stock += fg
        if type(weld) == int:
            total_stock += weld
        if type(hps) == int:
            total_stock += hps
        if daily_order > 0:
            coverage = round(fg_stock/(daily_order),1)
        else:
            coverage = 0
        short = daily_order - total_stock
        if short < 0:
            short = 0
        prod_plan = frappe.db.get_value('SAP Production Plan',{'product_no':m.mat_no,'order_date':today()},'planned_quantity') or '-'
        status = ''
        if type(prod_plan) == int:
            if prod_plan < int(daily_order):
                status = 'Check'
            else:
                status = 'OK'
        else:
            status = 'Check'
        row.extend([fg,weld,hps,total_stock,coverage,short,prod_plan,status])
        # upcoming_qty1 = frappe.db.get_value('TSA Master',{'mat_no':m.mat_no,'date':add_days(to_date,1)},'quantity') or '-'
        # upcoming_qty2 = frappe.db.get_value('TSA Master',{'mat_no':m.mat_no,'date':add_days(to_date,2)},'quantity') or '-'
        # row.extend([upcoming_qty1,upcoming_qty2])
        data.append(row)
    return data
        
        

    # for date in dates:

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'