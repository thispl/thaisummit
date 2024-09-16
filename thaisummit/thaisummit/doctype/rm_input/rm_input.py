# Copyright (c) 2023, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment,Border,Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
import json
from six import BytesIO, string_types
import pandas as pd
import requests




class RMInput(Document):

	@frappe.whitelist()
	def get_old_iym(self):
		rm_price_iym = frappe.get_single('RM Input').rm_input_table1
		# rm_price_re = frappe.get_single('RM Input').rm_input_table2
		for r in rm_price_iym:
			self.append('old_iym_settings',{
				'customer':r.customer,
				'grade':r.grade,
				'std_old':r.std_old,
				'std_impact':r.std_impact,
				'std_new':r.std_new,
				'old1':r.old1,
				'impact1':r.impact1,
				'new1':r.new1,
				'old2':r.old2,
				'impact2':r.impact2,
				'new2':r.new2,
				'coil_old':r.coil_old,
				'coil_impact':r.coil_impact,
				'coil_new':r.coil_new,
				'coil_old1':r.coil_old1,
				'coil_impact1':r.coil_impact1,
				'coil_new1':r.coil_new1,
				'tube_100_old':r.tube_100_old,	
				'tube_100_impact':r.tube_100_impact,
				'tube_100_new':r.tube_100_new,
				'tube_old':r.tube_old,
				'tube_impact':r.tube_impact,
				'tube_new':r.tube_new
	
			})
		self.save()
		# start

	@frappe.whitelist()
	def get_old_re(self):
		for s in self.rm_input_table2:
			self.append('old_re_settings',{
				'customer':s.customer,
				'grade':s.grade,
				'old1':s.old1,
				'impact1':s.impact1,
				'new1':s.new1,
				'old2':s.old2,
				'impact2':s.impact2,
				'new2':s.new2,
				'coil_old':s.coil_old,
				'coil_impact':s.coil_impact,
				'coil_new':s.coil_new,
				'coil_old1':s.coil_old1,
				'coil_impact1':s.coil_impact1,
				'coil_new1':s.coil_new1,
				'tube_100_old':s.tube_100_old,	
				'tube_100_impact':s.tube_100_impact,
				'tube_100_new':s.tube_100_new,
				'tube_old':s.tube_old,
				'tube_impact':s.tube_impact,
				'tube_new':s.tube_new
			})
		self.save()

@frappe.whitelist()
def make_old_iym_sheet():
    args = frappe.local.form_dict
    filename = args.name
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    doc = frappe.get_doc("RM Input",args.name)
    if doc:
        ws.append(["Customer","Grade","STD Old","STD Impact","STD New",">100 Old",">100 Impact",">100 New","<100 Old","<100 Impact","<100 New"])
        for i in doc.old_iym_settings:
            ws.append([i.customer,i.grade,i.std_old,i.std_impact,i.std_new,i.old1,i.impact1,i.new1,i.old2,i.impact2,i.new2])
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary' 	
 
 
@frappe.whitelist()
def make_old_re_sheet():
    args = frappe.local.form_dict
    filename = args.name
    test = build_xlsx_response_re(filename)

def make_xlsx_file(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    doc = frappe.get_doc("RM Input",args.name)
    if doc:
        ws.append(["Customer","Grade","Strip >100 Old","Strip >100 Impact","Strip >100 New","Strip <100 Old","Strip <100 Impact","Strip <100 New","Coil >100 Old","Coil >100 Impact","Coil >100 New","Coil <100 Old","Coil <100 Impact","Coil <100 New"])
        for i in doc.old_re_settings:
            ws.append([i.customer,i.grade,i.old1,i.impact1,i.new1,i.old2,i.impact2,i.new2,i.coil_old,i.coil_impact,i.coil_new,i.coil_old1,i.coil_impact1,i.coil_new1])
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response_re(filename):
    xlsx_file = make_xlsx_file(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary' 	

def get_live_stock():
    mat_no = '20000610'
    url = "http://apioso.thaisummit.co.th:10401/api/GetItemInventory"
    payload = json.dumps({
        "ItemCode": mat_no,
    })
    headers = {
        'Content-Type': 'application/json',
        'API_KEY': '/1^i[#fhSSDnC8mHNTbg;h^uR7uZe#ninearin!g9D:pos+&terpTpdaJ$|7/QYups;==~w~!AWwb&DU',
    }
    response = requests.request(
        "POST", url, headers=headers, data=payload)
    stock = 0
    if response:
        stocks = json.loads(response.text)
        if stocks:
            ica = frappe.db.sql(
                "select warehouse from `tabInventory Control Area` where iym = 'Y' ", as_dict=True)

            wh_list = [d['warehouse'] for d in ica if 'warehouse' in d]

            df = pd.DataFrame(stocks)
            df = df[df['Warehouse'].isin(wh_list)]
            stock = pd.to_numeric(df["Qty"]).sum()
        print(stock or 0)





