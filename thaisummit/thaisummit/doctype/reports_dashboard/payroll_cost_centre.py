from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def download():
	filename = 'Payroll Cost Center - Report'
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	header_date = title(args)
	ws.append(header_date)
	header_date1=title1(args)
	ws.append(header_date1)
	header_columns=get_column(args)
	ws.append(header_columns)
	department_group = ['Management','Admin','Support','Production']
	mng = frappe.get_all('Department',{'cost_center_group':'Management'})
	admin = frappe.get_all('Department',{'cost_center_group':'Admin'})
	spt = frappe.get_all('Department',{'cost_center_group':'Support'})
	dpt = frappe.get_all('Department',{'cost_center_group':'Production'})
	data = get_data(args)
	for d in data:
		ws.append(d)
	
	
	ws.merge_cells(start_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, start_column=1, end_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, end_column=3)
	ws.merge_cells(start_row=4, start_column=1, end_row=len(mng)+3, end_column=1)
	ws.merge_cells(start_row=len(mng)+4, start_column=1, end_row=len(mng)+len(admin)+3, end_column=1)
	ws.merge_cells(start_row=len(mng)+len(admin)+4, start_column=1, end_row=len(mng)+len(admin)+len(spt)+3, end_column=1)
	ws.merge_cells(start_row=len(mng)+len(admin)+len(spt)+4, start_column=1, end_row=len(mng)+len(admin)+len(spt)+len(dpt)+3, end_column=1)
	for cell in ws["1:1"]:
		cell.alignment = Alignment(horizontal='center')
	for cell in ws["2:2"]:
		cell.alignment = Alignment(horizontal='center')
	for cell in ws["3:3"]:
		cell.alignment = Alignment(horizontal='center')
	for cell in ws["B:B"]:
		cell.alignment = Alignment(horizontal='center')

	for rows in ws.iter_rows(min_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, max_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, min_col=1, max_col=27):
		for cell in rows:
			cell.alignment = Alignment(horizontal='center')

	for rows in ws.iter_rows(min_row=2, max_row=len(mng)+len(admin)+len(spt)+len(dpt)+2, min_col=1, max_col=1):
		for cell in rows:
			cell.alignment = Alignment(vertical='center')

	border = Border(left=Side(border_style='thin', color='000000'),
			right=Side(border_style='thin', color='000000'),
			top=Side(border_style='thin', color='000000'),
			bottom=Side(border_style='thin', color='000000')) 

	for rows in ws.iter_rows(min_row=1, max_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, min_col=1, max_col=26):
		for cell in rows:
			cell.border = border
		
	for header in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=26):
		 for cell in header:
			 cell.fill = PatternFill(fgColor='b9deff', fill_type = "solid")
			 cell.font = Font(bold=True)
	for header in ws.iter_rows(min_row=4, max_row=6, min_col=1, max_col=26):
		 for cell in header:
			 cell.fill = PatternFill(fgColor='b9ffda', fill_type = "solid")
			 

	for header in ws.iter_rows(min_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, max_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, min_col=1, max_col=26):
		 for cell in header:
			 cell.font = Font(bold=True)
			 

	for header in ws.iter_rows(min_row=len(mng)+len(admin)+4, max_row=len(spt)+len(mng)+len(admin)+3, min_col=1, max_col=26):
		 for cell in header:
			 cell.fill = PatternFill(fgColor='ffdab9', fill_type = "solid")
	

	# for header in ws.iter_rows(min_row=len(mng)+3, max_row=len(mng)+len(admin)+3, min_col=1, max_col=27):
	#      for cell in header:
	#          cell.fill = PatternFill(fgColor='ffdab9', fill_type = "solid")
	
	# for header in ws.iter_rows(min_row=len(mng)+3, max_row=len(mng)+3, min_col=1, max_col=26):
	# 	 for cell in header:
	# 		 cell.fill = PatternFill(fgColor='8A9A5B', fill_type = "solid")
	   
	  

	align_center = Alignment(horizontal='center',vertical='center')
	for cell in ws["2:2"]:
		cell.alignment = align_center
		cell.font = Font(bold=True)
	for cell in ws['1:1']:
		cell.alignment = align_center
		cell.font = Font(bold=True)
	border = Border(left=Side(border_style='thin'),
			 right=Side(border_style='thin'),
			 top=Side(border_style='thin'),
			 bottom=Side(border_style='thin'))
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= 26)
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column= 3)
	ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column= 17)
	ws.merge_cells(start_row=2, start_column=18, end_row=2, end_column= 24)
	ws.merge_cells(start_row=2, start_column=25, end_row=2, end_column= 26)

	ws.merge_cells(start_row=len(data)+4, start_column=1, end_row=len(data)+4, end_column= 26)
	
	total1= get_total1(args)
	for val in total1:
		ws.append(val)
		ws.merge_cells(start_row=len(data)+5, start_column=1, end_row=len(data)+5, end_column= 3)
	total2 = get_total2(args)
	for val in total2:
		ws.append(val)
		ws.merge_cells(start_row=len(data)+6, start_column=1, end_row=len(data)+6, end_column= 3)
	total3 = get_total3(args)
	for val in total3:
		ws.append(val)
		ws.merge_cells(start_row=len(data)+7, start_column=1, end_row=len(data)+7, end_column= 3)
	total4 = get_total4(args)
	for val in total4:
		ws.append(val)
		ws.merge_cells(start_row=len(data)+8, start_column=1, end_row=len(data)+8, end_column= 3)
	total= get_total(args)
	for val in total:
		ws.append(val)
		ws.merge_cells(start_row=len(data)+9, start_column=1, end_row=len(data)+9, end_column= 3)
	border = Border(left=Side(border_style='thin', color='000000'),
			right=Side(border_style='thin', color='000000'),
			top=Side(border_style='thin', color='000000'),
			bottom=Side(border_style='thin', color='000000')) 
	align_center = Alignment(horizontal='center',vertical='center')
	# for cell in ws["57:57"]:
	# 	cell.alignment = align_center
	# for cell in ws['58:58']:
	# 	cell.alignment = align_center
	# for cell in ws["59:59"]:
	# 	cell.alignment = align_center
	# for cell in ws['60:60']:
	# 	cell.alignment = align_center
	# for cell in ws['61:61']:
	# 	cell.alignment = align_center
		# cell.font = Font(bold=True)
	border = Border(left=Side(border_style='thin', color='000000'),
			right=Side(border_style='thin', color='000000'),
			top=Side(border_style='thin', color='000000'),
			bottom=Side(border_style='thin', color='000000')) 
	for rows in ws.iter_rows(min_row=len(data)+4,max_row=len(data)+9, min_col=1, max_col=26):
		for cell in rows:
			cell.border = border

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	# write out response as a xlsx type
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def title(args):
	month = datetime.strptime(str(args.to_date),'%Y-%m-%d')
	mon = str(month.strftime('%b') +' '+ str(month.strftime('%Y')))
	data = ["Payroll Cost Center %s"%mon,]
	return data

@frappe.whitelist()
def title1(args):
	data = ['','','','Earnings','','','','','','','','','','','','','','Deductions']
	return data


@frappe.whitelist()
def get_column(args):
	data=["Cost Center Parent","CC","Dept","BAS","HRA","CON","SPA","CH","WA","P ALLO",'ATA','SHT','Additional Allowance','Attendance Bonus','Arrear','Transport allowance','OT AMT','Gross','PF','ESI','CAT','PTAX','LWF','othe Deduc','TOTAL','NET SALARY','Bonus']
	return data

@frappe.whitelist()
def get_data(args):
	data = []
	department_group = ['Management','Admin','Support','Production']
	total_basic  = 0
	total_hra = 0
	total_con = 0
	total_spa = 0
	total_ab=0
	total_aa = 0
	total_a = 0
	total_ch = 0
	total_wa = 0
	total_pa = 0
	total_ata = 0
	total_sht = 0
	total_ta =0
	total_ot = 0
	total_gross = 0
	total_pf = 0
	total_esi = 0
	total_can = 0
	total_ptax = 0
	total_tds = 0
	total_od = 0
	total_total_deduct = 0
	total_net_pay = 0
	total_bonus = 0

	for group in department_group:
		departments = frappe.get_all('Department',{'cost_center_group':group},['cost_centre','name'])
		for dept in departments:
			basic = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as basic from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Basic' and `tabSalary Slip`.employee_type ='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].basic
		  
			hra = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as hra from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'House Rent Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].hra
			
			con = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as con from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Conveyance Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].con
			 
			spa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as spa from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Special Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].spa
			
			# med = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as med from `tabSalary Slip`
			# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Medical Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].med
			
			# lta = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as lta from `tabSalary Slip`
			# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Leave Travel Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].lta
			
			# ce = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ce from `tabSalary Slip`
			# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Children Education' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].ce
			
			ch = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ch from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Children Hostel' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ch

			wa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as wa from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Washing Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].wa

			pa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pa from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Position Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].pa

			ata = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ata from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ata

			sht = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sht from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Shift Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].sht
			
			aa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as aa from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Additional Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].aa

			ab = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ab from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Bonus' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ab


			a = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as a from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Arrear' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].a 

			

			ta = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ta from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Transport Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ta
 
			ot = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ot from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'others' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ot
 

			gross = frappe.db.sql("""select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
			where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].gross

			pf = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pf from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'provident Fund' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].pf
 
			esi = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as esi from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Employee State Insurance' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].esi
			
			can = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as can from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Canteen Charges' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].can
			
			ptax = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ptax from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'professional Tax' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ptax
			
			tds = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as tds from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Tax Deducted at Source' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].tds
			
			od = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as od from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Other Deduction' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].od
			
			total_deduct = frappe.db.sql("""select sum(`tabSalary Slip`.total_deduction) as total_deduct from `tabSalary Slip`
			where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].total_deduct

			
			net_pay = frappe.db.sql("""select sum(`tabSalary Slip`.net_pay) as net_pay from `tabSalary Slip`
			where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].net_pay
			
			bonus = 0
			if(basic):
				bonus = round(basic / 12)

			row = [group,dept.cost_centre,dept.name,basic or 0,hra or 0,con or 0,spa or 0,ch or 0,wa or 0,pa or 0,ata or 0,sht or 0,aa or 0,ab or 0, a or 0,ta or 0,ot or 0,gross or 0,pf or 0,esi or 0,can or 0,ptax or 0,tds or 0,od or 0,total_deduct or 0,net_pay or 0,bonus or 0]
			data.append(row)
			total_basic += basic or 0
			total_hra +=  hra or 0
			total_con += con or 0
			total_spa += spa or 0
			total_ch += ch or 0
			total_wa += wa or 0
			total_pa += pa or 0
			total_ata += ata or 0
			total_sht += sht or 0
			total_aa += aa or 0
			total_ab += ab or 0
			total_a += a or 0
			total_ta += ta or 0
			total_ot += ot or 0
			total_gross += gross or 0
			total_pf += pf or 0
			total_esi += esi or 0
			total_can += can or 0
			total_ptax += ptax or 0
			total_tds += tds or 0
			total_od += od or 0
			total_total_deduct += total_deduct or 0
			total_net_pay += net_pay or 0
			total_bonus += bonus or 0

	row = ['Total','','',total_basic or 0, total_hra or 0,total_con or 0,total_spa or 0,total_ch or 0,total_wa or 0,total_pa or 0,total_ata or 0,total_sht or 0,total_aa or 0,total_ab or 0 ,total_a or 0,total_ta or 0,total_ot or 0,total_gross,total_pf or 0,total_esi or 0,total_can or 0,total_ptax or 0,total_tds or 0,total_od or 0,total_total_deduct or 0,total_net_pay or 0,total_bonus or 0]
	data.append(row)  
			
	return data
   

@frappe.whitelist()
def get_total1(args):
	data=[]
	total_basic  = 0
	total_hra = 0
	total_con = 0
	total_spa = 0
	total_aa = 0
	total_ab = 0

	total_a = 0
	total_ch = 0
	total_wa = 0
	total_pa = 0
	total_ata = 0
	total_sht = 0
	total_ta =0
	total_ot = 0
	total_gross = 0
	total_pf = 0
	total_esi = 0
	total_can = 0
	total_ptax = 0
	total_tds = 0
	total_od = 0
	total_total_deduct = 0
	total_net_pay = 0
	total_bonus = 0
	departments = frappe.get_all('Department',{'cost_center_group':"Management"},['cost_centre','name'])
	for dept in departments:
		basic = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as basic from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Basic' and `tabSalary Slip`.employee_type ='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].basic
		  
		hra = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as hra from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'House Rent Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].hra
		
		con = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as con from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Conveyance Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].con
			
		spa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as spa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Special Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].spa
		
		# med = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as med from `tabSalary Slip`
		# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Medical Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].med
		
		# lta = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as lta from `tabSalary Slip`
		# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Leave Travel Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].lta
		
		# ce = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ce from `tabSalary Slip`
		# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Children Education' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].ce
		
		ch = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ch from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Children Hostel' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ch

		wa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as wa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Washing Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].wa

		pa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Position Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].pa

		ata = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ata from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ata

		sht = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sht from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Shift Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].sht
		
		aa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as aa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Additional Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].aa

		ab = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ab from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Bonus' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ab

		a = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sht from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Arrear' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].a 

		ta = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ta from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Transport Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ta

		ot = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ot from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'others' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ot


		gross = frappe.db.sql("""select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
		where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].gross

		pf = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pf from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'provident Fund' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].pf

		esi = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as esi from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Employee State Insurance' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].esi
		
		can = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as can from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Canteen Charges' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].can
		
		ptax = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ptax from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'professional Tax' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ptax
		
		tds = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as tds from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Tax Deducted at Source' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].tds
		
		od = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as od from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Other Deduction' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].od
		
		total_deduct = frappe.db.sql("""select sum(`tabSalary Slip`.total_deduction) as total_deduct from `tabSalary Slip`
		where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].total_deduct

		
		net_pay = frappe.db.sql("""select sum(`tabSalary Slip`.net_pay) as net_pay from `tabSalary Slip`
		where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].net_pay
		
		bonus = 0
		if(basic):
			bonus = round(basic / 12)	
		total_basic += basic or 0
		total_hra +=  hra or 0
		total_con += con or 0
		total_spa += spa or 0
		total_ch += ch or 0
		total_wa += wa or 0
		total_pa += pa or 0
		total_ata += ata or 0
		total_sht += sht or 0
		total_aa += aa or 0
		total_ab += ab or 0
		total_a += a or 0
		total_ta += ta or 0
		total_ot += ot or 0
		total_gross += gross or 0
		total_pf += pf or 0
		total_esi += esi or 0
		total_can += can or 0
		total_ptax += ptax or 0
		total_tds += tds or 0
		total_od += od or 0
		total_total_deduct += total_deduct or 0
		total_net_pay += net_pay or 0
		total_bonus += bonus or 0
	row = ['Management','','',total_basic or 0, total_hra or 0,total_con or 0,total_spa or 0,total_ch or 0,total_wa or 0,total_pa or 0,total_ata or 0,total_sht or 0,total_aa or 0,total_ab or 0, total_a or 0,total_ta or 0,total_ot or 0,total_gross,total_pf or 0,total_esi or 0,total_can or 0,total_ptax or 0,total_tds or 0,total_od or 0,total_total_deduct or 0,total_net_pay or 0,total_bonus or 0]
	data.append(row)  
			
	return data
@frappe.whitelist()
def get_total2(args):
	data=[]
	total_basic  = 0
	total_hra = 0
	total_con = 0
	total_spa = 0
	total_aa = 0
	total_ab = 0

	total_a = 0
	total_ch = 0
	total_wa = 0
	total_pa = 0
	total_ata = 0
	total_sht = 0
	total_ta =0
	total_ot = 0
	total_gross = 0
	total_pf = 0
	total_esi = 0
	total_can = 0
	total_ptax = 0
	total_tds = 0
	total_od = 0
	total_total_deduct = 0
	total_net_pay = 0
	total_bonus = 0
	departments = frappe.get_all('Department',{'cost_center_group':"Admin"},['cost_centre','name'])
	for dept in departments:
		basic = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as basic from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Basic' and `tabSalary Slip`.employee_type ='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].basic
		  
		hra = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as hra from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'House Rent Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].hra
		
		con = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as con from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Conveyance Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].con
			
		spa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as spa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Special Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].spa
		
		# med = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as med from `tabSalary Slip`
		# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Medical Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].med
		
		# lta = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as lta from `tabSalary Slip`
		# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Leave Travel Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].lta
		
		# ce = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ce from `tabSalary Slip`
		# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Children Education' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].ce
		
		ch = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ch from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Children Hostel' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ch

		wa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as wa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Washing Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].wa

		pa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Position Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].pa

		ata = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ata from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ata

		sht = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sht from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Shift Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].sht
		
		aa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as aa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Additional Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].aa

		ab = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ab from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Bonus' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ab


		a = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sht from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Arrear' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].a 

		ta = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ta from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Transport Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ta

		ot = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ot from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'others' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ot


		gross = frappe.db.sql("""select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
		where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].gross

		pf = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pf from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'provident Fund' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].pf

		esi = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as esi from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Employee State Insurance' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].esi
		
		can = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as can from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Canteen Charges' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].can
		
		ptax = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ptax from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'professional Tax' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ptax
		
		tds = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as tds from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Tax Deducted at Source' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].tds
		
		od = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as od from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Other Deduction' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].od
		
		total_deduct = frappe.db.sql("""select sum(`tabSalary Slip`.total_deduction) as total_deduct from `tabSalary Slip`
		where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].total_deduct

		
		net_pay = frappe.db.sql("""select sum(`tabSalary Slip`.net_pay) as net_pay from `tabSalary Slip`
		where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].net_pay		
		bonus = 0
		if(basic):
			bonus = round(basic / 12)	
		total_basic += basic or 0
		total_hra +=  hra or 0
		total_con += con or 0
		total_spa += spa or 0
		total_ch += ch or 0
		total_wa += wa or 0
		total_pa += pa or 0
		total_ata += ata or 0
		total_sht += sht or 0
		total_aa += aa or 0
		total_ab += ab or 0
		
		total_a += a or 0
		total_ta += ta or 0
		total_ot += ot or 0
		total_gross += gross or 0
		total_pf += pf or 0
		total_esi += esi or 0
		total_can += can or 0
		total_ptax += ptax or 0
		total_tds += tds or 0
		total_od += od or 0
		total_total_deduct += total_deduct or 0
		total_net_pay += net_pay or 0
		total_bonus += bonus or 0
	row = ['Admin','','',total_basic or 0, total_hra or 0,total_con or 0,total_spa or 0,total_ch or 0,total_wa or 0,total_pa or 0,total_ata or 0,total_sht or 0,total_aa or 0,total_ab or 0 ,total_a or 0,total_ta or 0,total_ot or 0,total_gross,total_pf or 0,total_esi or 0,total_can or 0,total_ptax or 0,total_tds or 0,total_od or 0,total_total_deduct or 0,total_net_pay or 0,total_bonus or 0]
	data.append(row)  
	return data
@frappe.whitelist()
def get_total3(args):
	data=[]
	total_basic  = 0
	total_hra = 0
	total_con = 0
	total_spa = 0
	total_aa = 0
	total_a = 0
	total_ch = 0
	total_wa = 0
	total_pa = 0
	total_ata = 0
	total_sht = 0
	total_ab = 0

	total_ta =0
	total_ot = 0
	total_gross = 0
	total_pf = 0
	total_esi = 0
	total_can = 0
	total_ptax = 0
	total_tds = 0
	total_od = 0
	total_total_deduct = 0
	total_net_pay = 0
	total_bonus = 0
	departments = frappe.get_all('Department',{'cost_center_group':"Support"},['cost_centre','name'])
	for dept in departments:
		basic = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as basic from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Basic' and `tabSalary Slip`.employee_type ='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].basic
		  
		hra = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as hra from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'House Rent Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].hra
		
		con = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as con from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Conveyance Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].con
			
		spa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as spa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Special Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].spa
		
		# med = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as med from `tabSalary Slip`
		# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Medical Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].med
		
		# lta = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as lta from `tabSalary Slip`
		# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Leave Travel Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].lta
		
		# ce = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ce from `tabSalary Slip`
		# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Children Education' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].ce
		
		ch = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ch from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Children Hostel' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ch

		wa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as wa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Washing Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].wa

		pa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Position Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].pa

		ata = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ata from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ata

		sht = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sht from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Shift Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].sht
		
		aa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as aa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Additional Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].aa

		ab = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ab from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Bonus' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ab


		a = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sht from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Arrear' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].a 

		ta = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ta from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Transport Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ta

		ot = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ot from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'others' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ot


		gross = frappe.db.sql("""select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
		where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].gross

		pf = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pf from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'provident Fund' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].pf

		esi = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as esi from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Employee State Insurance' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].esi
		
		can = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as can from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Canteen Charges' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].can
		
		ptax = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ptax from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'professional Tax' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ptax
		
		tds = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as tds from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Tax Deducted at Source' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].tds
		
		od = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as od from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Other Deduction' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].od
		
		total_deduct = frappe.db.sql("""select sum(`tabSalary Slip`.total_deduction) as total_deduct from `tabSalary Slip`
		where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].total_deduct

		
		net_pay = frappe.db.sql("""select sum(`tabSalary Slip`.net_pay) as net_pay from `tabSalary Slip`
		where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].net_pay
	
		bonus = 0
		if(basic):
			bonus = round(basic / 12)	
		total_basic += basic or 0
		total_hra +=  hra or 0
		total_con += con or 0
		total_spa += spa or 0
		total_ch += ch or 0
		total_wa += wa or 0
		total_pa += pa or 0
		total_ata += ata or 0
		total_sht += sht or 0
		total_aa += aa or 0
		total_ab += ab or 0
		total_a += a or 0
		total_ta += ta or 0
		total_ot += ot or 0
		total_gross += gross or 0
		total_pf += pf or 0
		total_esi += esi or 0
		total_can += can or 0
		total_ptax += ptax or 0
		total_tds += tds or 0
		total_od += od or 0
		total_total_deduct += total_deduct or 0
		total_net_pay += net_pay or 0
		total_bonus += bonus or 0
	row = ['Support','','',total_basic or 0, total_hra or 0,total_con or 0,total_spa or 0,total_ch or 0,total_wa or 0,total_pa or 0,total_ata or 0,total_sht or 0,total_aa or 0,total_ab or 0 ,total_a or 0,total_ta or 0,total_ot or 0,total_gross,total_pf or 0,total_esi or 0,total_can or 0,total_ptax or 0,total_tds or 0,total_od or 0,total_total_deduct or 0,total_net_pay or 0,total_bonus or 0]
	data.append(row)  
	return data
@frappe.whitelist()
def get_total4(args):
	data=[]
	total_basic  = 0
	total_hra = 0
	total_con = 0
	total_ab = 0

	total_spa = 0
	total_aa = 0
	total_a = 0
	total_ch = 0
	total_wa = 0
	total_pa = 0
	total_ata = 0
	total_sht = 0
	total_ta =0
	total_ot = 0
	total_gross = 0
	total_pf = 0
	total_esi = 0
	total_can = 0
	total_ptax = 0
	total_tds = 0
	total_od = 0
	total_total_deduct = 0
	total_net_pay = 0
	total_bonus = 0
	departments = frappe.get_all('Department',{'cost_center_group':"Production"},['cost_centre','name'])
	for dept in departments:
		basic = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as basic from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Basic' and `tabSalary Slip`.employee_type ='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].basic
		
		hra = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as hra from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'House Rent Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].hra
		
		con = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as con from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Conveyance Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].con
			
		spa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as spa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Special Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].spa
		
		# med = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as med from `tabSalary Slip`
		# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Medical Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].med
		
		# lta = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as lta from `tabSalary Slip`
		# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Leave Travel Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].lta
		
		# ce = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ce from `tabSalary Slip`
		# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Children Education' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].ce
		
		ch = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ch from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Children Hostel' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ch

		wa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as wa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Washing Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].wa

		pa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Position Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].pa

		ata = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ata from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ata

		sht = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sht from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Shift Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].sht
		
		aa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as aa from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Additional Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].aa

		ab = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ab from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Bonus' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ab


		a = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sht from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Arrear' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].a 

		ta = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ta from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Transport Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ta

		ot = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ot from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'others' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ot


		gross = frappe.db.sql("""select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
		where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].gross

		pf = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pf from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'provident Fund' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].pf

		esi = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as esi from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Employee State Insurance' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].esi
		
		can = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as can from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Canteen Charges' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].can
		
		ptax = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ptax from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'professional Tax' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ptax
		
		tds = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as tds from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Tax Deducted at Source' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].tds
		
		od = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as od from `tabSalary Slip`
		left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Other Deduction' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].od
		
		total_deduct = frappe.db.sql("""select sum(`tabSalary Slip`.total_deduction) as total_deduct from `tabSalary Slip`
		where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].total_deduct

		
		net_pay = frappe.db.sql("""select sum(`tabSalary Slip`.net_pay) as net_pay from `tabSalary Slip`
		where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].net_pay
	
		bonus = 0
		if(basic):
			bonus = round(basic / 12)	
		total_basic += basic or 0
		total_hra +=  hra or 0
		total_con += con or 0
		total_spa += spa or 0
		total_ch += ch or 0
		total_wa += wa or 0
		total_pa += pa or 0
		total_ata += ata or 0
		total_sht += sht or 0
		total_aa += aa or 0
		total_ab += ab or 0
		total_a += a or 0
		total_ta += ta or 0
		total_ot += ot or 0
		total_gross += gross or 0
		total_pf += pf or 0
		total_esi += esi or 0
		total_can += can or 0
		total_ptax += ptax or 0
		total_tds += tds or 0
		total_od += od or 0
		total_total_deduct += total_deduct or 0
		total_net_pay += net_pay or 0
		total_bonus += bonus or 0
	row = ['Production','','',total_basic or 0, total_hra or 0,total_con or 0,total_spa or 0,total_ch or 0,total_wa or 0,total_pa or 0,total_ata or 0,total_sht or 0,total_aa or 0,total_ab or 0 ,total_a or 0,total_ta or 0,total_ot or 0,total_gross,total_pf or 0,total_esi or 0,total_can or 0,total_ptax or 0,total_tds or 0,total_od or 0,total_total_deduct or 0,total_net_pay or 0,total_bonus or 0]
	data.append(row)  
	return data
@frappe.whitelist()
def get_total(args):
	data = []
	department_group = ['Management','Admin','Support','Production']
	total_basic  = 0
	total_hra = 0
	total_con = 0
	total_ab = 0

	total_spa = 0
	total_aa = 0
	total_a = 0
	total_ch = 0
	total_wa = 0
	total_pa = 0
	total_ata = 0
	total_sht = 0
	total_ta =0
	total_ot = 0
	total_gross = 0
	total_pf = 0
	total_esi = 0
	total_can = 0
	total_ptax = 0
	total_tds = 0
	total_od = 0
	total_total_deduct = 0
	total_net_pay = 0
	total_bonus = 0

	for group in department_group:
		departments = frappe.get_all('Department',{'cost_center_group':group},['cost_centre','name'])
		for dept in departments:
			basic = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as basic from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Basic' and `tabSalary Slip`.employee_type ='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].basic
		  
			hra = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as hra from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'House Rent Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].hra
			
			con = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as con from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Conveyance Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].con
			 
			spa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as spa from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Special Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].spa
			
			# med = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as med from `tabSalary Slip`
			# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Medical Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].med
			
			# lta = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as lta from `tabSalary Slip`
			# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Leave Travel Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].lta
			
			# ce = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ce from `tabSalary Slip`
			# left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Children Education' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].ce
			
			ch = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ch from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Children Hostel' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ch

			wa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as wa from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Washing Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].wa

			pa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pa from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Position Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].pa

			ata = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ata from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ata

			sht = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sht from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Shift Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].sht
			
			aa = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as aa from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Additional Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].aa

			ab = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ab from `tabSalary Slip`
		    left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Bonus' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ab


			a = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as sht from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Arrear' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].a 

			ta = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ta from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Transport Allowance' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ta
 
			ot = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ot from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'others' and `tabSalary Slip`.employee_type='%s' and`tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ot
 

			gross = frappe.db.sql("""select sum(`tabSalary Slip`.gross_pay) as gross from `tabSalary Slip`
			where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].gross

			pf = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as pf from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'provident Fund' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].pf
 
			esi = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as esi from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Employee State Insurance' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].esi
			
			can = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as can from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Canteen Charges' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].can
			
			ptax = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ptax from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'professional Tax' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].ptax
			
			tds = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as tds from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Tax Deducted at Source' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].tds
			
			od = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as od from `tabSalary Slip`
			left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Other Deduction' and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].od
			
			total_deduct = frappe.db.sql("""select sum(`tabSalary Slip`.total_deduction) as total_deduct from `tabSalary Slip`
			where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].total_deduct

			
			net_pay = frappe.db.sql("""select sum(`tabSalary Slip`.net_pay) as net_pay from `tabSalary Slip`
			where `tabSalary Slip`.department = '%s'  and `tabSalary Slip`.employee_type='%s' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' """%(dept.name,args.employee_type,args.from_date,args.to_date),as_dict=True)[0].net_pay
			
			bonus = 0
			if(basic):
				bonus = round(basic / 12)
			total_basic += basic or 0
			total_hra +=  hra or 0
			total_con += con or 0
			total_spa += spa or 0
			total_ch += ch or 0
			total_wa += wa or 0
			total_pa += pa or 0
			total_ata += ata or 0
			total_sht += sht or 0
			total_aa += aa or 0
			total_ab += ab or 0
			total_a += a or 0
			total_ta += ta or 0
			total_ot += ot or 0
			total_gross += gross or 0
			total_pf += pf or 0
			total_esi += esi or 0
			total_can += can or 0
			total_ptax += ptax or 0
			total_tds += tds or 0
			total_od += od or 0
			total_total_deduct += total_deduct or 0
			total_net_pay += net_pay or 0
			total_bonus += bonus or 0
	row = ['Total','','',total_basic or 0, total_hra or 0,total_con or 0,total_spa or 0,total_ch or 0,total_wa or 0,total_pa or 0,total_ata or 0,total_sht or 0,total_aa or 0,total_ab or 0 ,total_a or 0,total_ta or 0,total_ot or 0,total_gross,total_pf or 0,total_esi or 0,total_can or 0,total_ptax or 0,total_tds or 0,total_od or 0,total_total_deduct or 0,total_net_pay or 0,total_bonus or 0]
	data.append(row)  
			
	return data
   