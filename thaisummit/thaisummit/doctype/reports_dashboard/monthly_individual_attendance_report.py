from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe import _
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from six import BytesIO, string_types


class ReportsDashboard(Document):
    pass

@frappe.whitelist()
def download():
    filename = 'Individual report'
    test = build_xlsx_response(filename)


ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header = add_header(args)   
    ws.append(header)
    header_title = add_header_title()
    heading = add_heading()
    ws.append(header_title)
    ws.append(heading)

    data = get_data(args)
    # data = [['xdhfd'],['gxdxty']]

    for row in data:
        ws.append(row)

    # ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    # ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    # ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=8)
    # ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=14)
    # ws.merge_cells(start_row=1, start_column=15, end_row=1, end_column=20)
    # ws.merge_cells(start_row=1, start_column=21, end_row=1, end_column=26)
    # ws.merge_cells(start_row=1, start_column=27, end_row=1, end_column=32)

    # re = frappe.db.count('Department',{'parent_department':'RE','is_group':0})
    # iym = frappe.db.count('Department',{'parent_department':'IYM','is_group':0})
    # ford = frappe.db.count('Department',{'parent_department':'FORD','is_group':0})
    # support = frappe.db.count('Department',{'parent_department':'SUPPORT','is_group':0})

    # ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    # ws.merge_cells(start_row=3, start_column=1, end_row=re+2, end_column=1)
    # ws.merge_cells(start_row=15, start_column=1, end_row=iym+14, end_column=1)
    # ws.merge_cells(start_row=30, start_column=1, end_row=ford+29, end_column=1)
    # ws.merge_cells(start_row=33, start_column=1, end_row=support+32, end_column=1)


    # shift_1 = ws['C1']
    # shift_1.alignment = Alignment(horizontal='center')
    # shift_2 = ws['I1']
    # shift_2.alignment = Alignment(horizontal='center')
    # shift_3 = ws['O1']
    # shift_3.alignment = Alignment(horizontal='center')
    # shift_pp2 = ws['U1']
    # shift_pp2.alignment = Alignment(horizontal='center')
    # shift_tt = ws['AA1']
    # shift_tt.alignment = Alignment(horizontal='center')
    # re = ws['A3']
    # re.alignment = Alignment(vertical='center',horizontal='center')
    # iym = ws['A15']
    # iym.alignment = Alignment(vertical='center',horizontal='center')
    # ford = ws['A30']
    # ford.alignment = Alignment(vertical='center',horizontal='center')
    # support = ws['A33']
    # support.alignment = Alignment(vertical='center',horizontal='center')


    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def add_header(args):
    employee_name = frappe.db.get_value("Employee",args.employee,"employee_name")
    header = ["ID:","%s"%(args.employee),'','','NAME:',"%s"%(employee_name),'','Dept','','','HR & Admin']
    return header

@frappe.whitelist()
def add_header_title():
    header = ["","","Attendance","","",'','','','',"Overtime",'','','']
    return header


@frappe.whitelist()
def add_heading():
    header = ["Date","Day","Working","In time","Out time","Status",'',"Start time",'','End time','','Total Hrs']
    return header

@frappe.whitelist()
def get_data(args):
    data = []
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
   
    for date in dates:  
        in_time = frappe.db.get_value('Attendance' ,{'employee':args.employee,"attendance_date":date},'in_time') or ''
        out_time = frappe.db.get_value('Attendance' ,{'employee':args.employee,"attendance_date":date},'out_time') or ''
        start_time =frappe.db.get_value("Overtime Request",{'employee':args.employee,'ot_date':date},'from_time') or  ''
        end_time = frappe.db.get_value("Overtime Request",{'employee':args.employee,'ot_date':date},'to_time') or  ''
        total_hours = frappe.db.get_value("Overtime Request",{'employee':args.employee,'ot_date':date},'total_hours') or  ''
        working  = check_holiday(date)  
        status_format = get_status(args,date)
        dt = datetime.strptime(date,'%Y-%m-%d')
        day_format = datetime.date(dt).strftime('%a')
        data.append([dt.strftime('%d-%b'),day_format,working or 'W',in_time,out_time,status_format,'',start_time,'',end_time,'',total_hours])
    return data


def check_holiday(date):
    holiday = frappe.db.sql("""select `tabHoliday`.holiday_date,`tabHoliday`.weekly_off from `tabHoliday List` 
    left join `tabHoliday` on `tabHoliday`.parent = `tabHoliday List`.name where `tabHoliday List`.name = 'Holiday List - 2021' and holiday_date = '%s' """%(date),as_dict=True)
    if holiday:
        if holiday[0].weekly_off == 1:
            return "W/O"
        else:
            return "HH"

def get_status(args,date):
    data = frappe.db.get_value('Attendance' ,{'employee':args.employee,"attendance_date":date},'status','shift','leave_type') 
    status = ''

    if data:        
        if data.status in ("Present","Half Day"):
            status = data.shift
        if data.status == "Absent":
            status = 'A'
        if data.status == "On Leave":
            if data.leave_type == 'Casual Leave':
                status = 'CL' 
            if data.leave_type == 'Sick Leave':
                status = 'SL'
            if data.leave_type == 'Earned Leave':
                status = 'EL'
    return status


    
    
 

