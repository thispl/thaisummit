# # Copyright (c) 2021, TEAMPRO and contributors
# # For license information, please see license.txt

# from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe import _
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from six import BytesIO, string_types
from frappe.utils.background_jobs import enqueue


class ReportsDashboard(Document):
    pass



# @frappe.whitelist()
# def download():
#     filename = 'subash'
#     test = build_xlsx_response(filename)


# ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
# # return xlsx file object
# def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
#     args = frappe.local.form_dict
#     column_widths = column_widths or []
#     if wb is None:
#         wb = openpyxl.Workbook()

#     ws = wb.create_sheet(sheet_name, 0)
#     header = add_header(args)
#     ws.append(header)
#     header_emp_type = add_header_emp_type()
#     ws.append(header_emp_type)

#     data = get_data(args)

#     for row in data:
#         ws.append(row)

#     ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
#     ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
#     ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=8)
#     ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=14)
#     ws.merge_cells(start_row=1, start_column=15, end_row=1, end_column=20)
#     ws.merge_cells(start_row=1, start_column=21, end_row=1, end_column=26)
#     ws.merge_cells(start_row=1, start_column=27, end_row=1, end_column=32)

#     re = frappe.db.count('Department',{'parent_department':'RE','is_group':0})
#     iym = frappe.db.count('Department',{'parent_department':'IYM','is_group':0})
#     ford = frappe.db.count('Department',{'parent_department':'FORD','is_group':0})
#     support = frappe.db.count('Department',{'parent_department':'SUPPORT','is_group':0})

#     ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
#     ws.merge_cells(start_row=3, start_column=1, end_row=re+2, end_column=1)
#     ws.merge_cells(start_row=15, start_column=1, end_row=iym+14, end_column=1)
#     ws.merge_cells(start_row=30, start_column=1, end_row=ford+29, end_column=1)
#     ws.merge_cells(start_row=33, start_column=1, end_row=support+32, end_column=1)


#     shift_1 = ws['C1']
#     shift_1.alignment = Alignment(horizontal='center')
#     shift_2 = ws['I1']
#     shift_2.alignment = Alignment(horizontal='center')
#     shift_3 = ws['O1']
#     shift_3.alignment = Alignment(horizontal='center')
#     shift_pp2 = ws['U1']
#     shift_pp2.alignment = Alignment(horizontal='center')
#     shift_tt = ws['AA1']
#     shift_tt.alignment = Alignment(horizontal='center')
#     re = ws['A3']
#     re.alignment = Alignment(vertical='center',horizontal='center')
#     iym = ws['A15']
#     iym.alignment = Alignment(vertical='center',horizontal='center')
#     ford = ws['A30']
#     ford.alignment = Alignment(vertical='center',horizontal='center')
#     support = ws['A33']
#     support.alignment = Alignment(vertical='center',horizontal='center')


#     xlsx_file = BytesIO()
#     wb.save(xlsx_file)
#     return xlsx_file    


# def build_xlsx_response(filename):
#     xlsx_file = make_xlsx(filename)
#     # write out response as a xlsx type
#     frappe.response['filename'] = filename + '.xlsx'
#     frappe.response['filecontent'] = xlsx_file.getvalue()
#     frappe.response['type'] = 'binary'


# @frappe.whitelist()
# def add_header(args):
#     header = ["Dept","%s"%(args.date),'1','','','','','','2','','','','','','3','','','','','','PP2','','','','','','TOTAL','','','','','',]
#     return header

# @frappe.whitelist()
# def add_header_emp_type():
#     header = ["","",'WC','BC','FT','NT','CL','TT','WC','BC','FT','NT','CL','TT','WC','BC','FT','NT','CL','TT','WC','BC','FT','NT','CL','TT','WC','BC','FT','NT','CL','TT']
#     return header

# def get_data(args):
#     data = []
#     dept_group = ['RE','IYM','FORD','SUPPORT']
#     for dg in dept_group:
#         departments = frappe.get_all('Department',{'parent_department':dg,'is_group':0})
#         for dept in departments:
#             row = [dg,dept.name]
#             shifts = ['1','2','3','PP2']
#             employee_type = ['WC','BC','FT','NT','CL','TT']
#             for shift in shifts:
#                 for emp_type in employee_type:
#                     count = frappe.db.count("QR Checkin",{'department':dept.name,'shift_date':args.date,'qr_shift':shift,'ot':0,'employee_type':emp_type})
#                     row.append(count)
#             data.append(row)
#     return data
