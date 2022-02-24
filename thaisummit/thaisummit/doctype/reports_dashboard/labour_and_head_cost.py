from __future__ import unicode_literals
import frappe
from frappe.monitor import start
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
from thaisummit.thaisummit.doctype.reports_dashboard.bc_salary_register import get_data
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def download():
    filename = 'Labour and Head cost'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name,0)
    ws.append(['MAR/21'])
  

    data = get_data(args)
    for d in data:
        ws.append(d)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
    # ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=17)
    # ws.merge_cells(start_row=2, start_column=18, end_row=2, end_column=26)
    # ws.merge_cells(start_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, start_column=1, end_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, end_column=3)
    # ws.merge_cells(start_row=4, start_column=1, end_row=len(mng)+3, end_column=1)
    # ws.merge_cells(start_row=len(mng)+4, start_column=1, end_row=len(mng)+len(admin)+3, end_column=1)
    # ws.merge_cells(start_row=len(mng)+len(admin)+4, start_column=1, end_row=len(mng)+len(admin)+len(spt)+3, end_column=1)
    # ws.merge_cells(start_row=len(mng)+len(admin)+len(spt)+4, start_column=1, end_row=len(mng)+len(admin)+len(spt)+len(dpt)+3, end_column=1)
    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws["2:2"]:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws["3:3"]:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws["B:B"]:
        cell.alignment = Alignment(horizontal='center')

    for rows in ws.iter_rows(min_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, max_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, min_col=1, max_col=27):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')

    for rows in ws.iter_rows(min_row=2, max_row=len(mng)+len(admin)+len(spt)+len(dpt)+2, min_col=1, max_col=1):
        for cell in rows:
            cell.alignment = Alignment(vertical='center')

    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')) 

    for rows in ws.iter_rows(min_row=1, max_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, min_col=1, max_col=27):
        for cell in rows:
            cell.border = border
            

    for header in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=27):
         for cell in header:
             cell.fill = PatternFill(fgColor='B6D0E2', fill_type = "solid")
             cell.font = Font(bold=True)

    for header in ws.iter_rows(min_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, max_row=len(mng)+len(admin)+len(spt)+len(dpt)+4, min_col=1, max_col=27):
         for cell in header:
             cell.font = Font(bold=True)
             

    for header in ws.iter_rows(min_row=len(mng)+len(admin)+4, max_row=len(spt)+len(mng)+len(admin)+3, min_col=1, max_col=27):
         for cell in header:
             cell.fill = PatternFill(fgColor='ffdab9', fill_type = "solid")

    # for header in ws.iter_rows(min_row=len(mng)+3, max_row=len(mng)+len(admin)+3, min_col=1, max_col=27):
    #      for cell in header:
    #          cell.fill = PatternFill(fgColor='ffdab9', fill_type = "solid")
    
    for header in ws.iter_rows(min_row=len(mng)+3, max_row=len(mng)+3, min_col=1, max_col=27):
         for cell in header:
             cell.fill = PatternFill(fgColor='8A9A5B', fill_type = "solid")
       
      
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
