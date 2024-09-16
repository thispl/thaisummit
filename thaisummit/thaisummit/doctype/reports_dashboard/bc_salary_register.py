from __future__ import unicode_literals
from os import stat
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, touch_file
from frappe import _
from frappe.utils.csvutils import UnicodeWriter, build_csv_response, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime, time

import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment,Border,Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types



@frappe.whitelist()
def download():
    
    filename = 'BC Salary Register'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    header_date = title1 (args)
    ws.append(header_date)
    header_column = get_column(args)
    ws.append(header_column)

    data = get_data(args)
    for row in data:
        ws.append(row)

    align_center = Alignment(horizontal='center',vertical='center')
    for cell in ws["2:2"]:
        cell.alignment = align_center
        cell.font = Font(bold=True)
    for cell in ws['1:1']:
        cell.alignment = align_center
        cell.font = Font(bold=True)

    border = Border(left=Side(border_style='thin', color='000000'),
             right=Side(border_style='thin', color='000000'),
             top=Side(border_style='thin', color='000000'),
             bottom=Side(border_style='thin', color='000000'))
    for rows in ws.iter_rows(min_row=1, max_row=len(get_data(args))+2, min_col=1, max_col=len(get_column(args))):
        for cell in rows:
            cell.border = border

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= 9)
    ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column= 14) 
    ws.merge_cells(start_row=1, start_column=15, end_row=1, end_column= 39)
    ws.merge_cells(start_row=1, start_column=40, end_row=1, end_column= 45)



    for header in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=len(get_column(args))-2):
            for cell in header:
                cell.fill = PatternFill(fgColor='fff808', fill_type = "solid")
    for header in ws.iter_rows(min_row=1, max_row=2, min_col=len(get_column(args))-2, max_col=len(get_column(args))):
            for cell in header:
                cell.fill = PatternFill(fgColor='fff808', fill_type = "solid")

    ws.freeze_panes = 'G3'

    ws.sheet_view.zoomScale = 80 

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def title1(args):
    month = datetime.strptime(str(args.to_date),'%Y-%m-%d')
    mon = str(month.strftime('%b')) + ' ' + str(month.strftime('%Y'))
    data = ["TSAI BC Wages for the Month of %s"%mon,'','','','','','','','','Standard Structure Month','','','','',
    'Earnings Per Month','','','','','','','','','','','','','','','','','','','','','','','','','Deductions','','','','','']
    return data

def get_dates(args):
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    return dates  

@frappe.whitelist()
def get_column(args):
    data = ["SR No","Cost C ","Emp No","Employee_Name","DOJ","Department","Designation",
    "Paid Days 100 %","OT Hours","Basic & DA","House Rent Allow","Welfare","Position All","Gross Wage",
    "Basic","HRA","Welfare","Position All","Additional Bonus","ATA","SHT","ARR","PP Allowance","Transport Allowance","Conveyance Allowance","Special Allowance","Medical Allowance","Leave Travel Allowance","Washing Allowance","Welding Allowance","Temp Allowance","Children Education","Children Hostel","Service Charges","Personal Protective Equipment","Cost to Company","Employer Provident Fund","Additional Allowance","Others",
    "Gross","PF","ESI","Can","P Tax","PPE","Other Deduction","Total","Net Wage","Bonus"]
    
    return data      

@frappe.whitelist() 
def get_data(args):
    data = []
    row = []
    basic_component_amount = earning_component_amount = deduction_component_amount = gross_wage = total_deduction = 0

    earning_comp = ["Basic","House Rent Allowance","Other Allowance","Position Allowance","Attendance Allowance","Attendance Bonus","Shift Allowance","Arrear","PP Allowance","Transport Allowance","Conveyance Allowance","Special Allowance","Medical Allowance","Leave Travel Allowance","Washing Allowance","Welding Allowance","Temp Allowance","Children Education","Children Hostel","Service Charges","Personal Protective Equipment","Cost to Company","Employer Provident Fund","Additional Allowance","Others",]

    dedcution_comp = ["Provident Fund","Employee State Insurance","Canteen Charges","Professional Tax","Personal Protective Equipment","Other Deduction"]

    # if args.department:
    #     salary_slips = frappe.get_all("Salary Slip",{'employee_type':'BC''start_date':args.from_date,'end_date':args.to_date},['*'])	

    if args.employee:
        salary_slips = frappe.get_all("Salary Slip",{'employee_type':'BC','employee':args.employee,'start_date':args.from_date,'end_date':args.to_date},['*'])	
    else:
        salary_slips = frappe.get_all("Salary Slip",{'employee_type':'BC','start_date':args.from_date,'end_date':args.to_date},['*'])	
    i =1
    for ss in salary_slips:
        row = [i,]
        cost_center = frappe.get_value('Department',ss.department,'cost_centre')
        emp = frappe.get_doc("Employee",ss.employee)
        # ot_records = frappe.get_all("Overtime Request",{'employee':ss.employee,'employee_type':'BC','ot_date': ['between', (ss.start_date,ss.end_date)]},['ot_hours'])
        # ot_hours = sum(timedelta_to_hours(record['ot_hours']) for record in ot_records)
        row += [cost_center,emp.name,emp.employee_name,emp.date_of_joining,emp.department,emp.designation,ss.payment_days,ss.total_working_hours,emp.basic,emp.house_rent_allowance,emp.other_allowance,emp.position_allowance,emp.basic+emp.house_rent_allowance+emp.other_allowance+emp.position_allowance]
            
        for ec in earning_comp:
            earning_component_amount = frappe.get_value('Salary Detail',{'salary_component':ec,'parent':ss.name},['amount'])
            if earning_component_amount:
                row.append(earning_component_amount)
            else:
                row.append('')
        row += [ss.gross_pay]
        total_deduction =0
        for dc in dedcution_comp:
            deduction_component_amount = frappe.get_value('Salary Detail',{'salary_component':dc,'parent':ss.name},['amount'])
            if deduction_component_amount:
                row += [deduction_component_amount]
                total_deduction += deduction_component_amount
            else:
                row += ['']
            bonus_value = frappe.get_value('Salary Detail',{'salary_component':'Basic','parent':ss.name},['amount']) or 0
            bonus = round(bonus_value/12)

        row.append(ss.total_deduction)
        row += [ ss.rounded_total ]
        row += [ bonus ]
        # row += [ frappe.get_value('Salary Detail',{'salary_component':'Others','parent':ss.name},['amount']) ]
        # row += [ss.gross_pay]
        data.append(row)
        i+=1
        
    return data

def timedelta_to_hours(td):
    return td.total_seconds() / 60.0 if isinstance(td, timedelta) else td