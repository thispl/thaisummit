# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
from inspect import ArgSpec
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils import cstr, add_days, date_diff, getdate, touch_file,today,get_first_day, get_last_day
from math import floor
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
from thaisummit.thaisummit.report.monthly_attendance_register.monthly_attendance_register import get_columns
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def download(f_date,t_date,emp_type):
    args = {'from_date':f_date,'to_date':t_date,'employee_type':emp_type}
    filename = 'Payroll Cross Check Report'
    build_xlsx_response(filename,args)
    # enqueue(build_xlsx_response, queue='default', timeout=6000, event='build_xlsx_response',filename=filename,args=args)


# return xlsx file object
def make_xlsx(data,args, sheet_name=None, wb=None, column_widths=None):
    # args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header = get_columns()
    ws.append(header)

    data = get_data(args)
    # frappe.log_error(title='xlsx',message=data)
    for row in data:
        ws.append(row)

    ws.sheet_view.zoomScale = 80

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    
    return xlsx_file


def build_xlsx_response(filename,args):
    xlsx_file = make_xlsx(filename,args)
    ret = frappe.get_doc({
            "doctype": "File",
            "attached_to_name": '',
            "attached_to_doctype": 'Reports Dashboard',
            "attached_to_field": 'attach',
            "file_name": filename + '.xlsx',
            "is_private": 0,
            "content": xlsx_file.getvalue(),
            "decode": False
        })
    ret.save(ignore_permissions=True)
    frappe.db.commit()
    frappe.log_error(message=ret)
    attached_file = frappe.get_doc("File", ret.name)
    frappe.db.set_value('Reports Dashboard',None,'attach',attached_file.file_url)

def get_columns():
    columns = ["ID No",
    "Name",
    "Department",
    "DOJ",
    "11",
    "22",
    "33","12","13","21",
    "23",
    "31",
    "32",
    "1L1",
    "2L2",
    "3L3",
    "1L2",
    "1L3",
    "2L1",
    "2L3",
    "3L1",
    "3L2",
    "PP1",
    "PP2",
    "PP1L",
    "PP2L",
    "1M",
    "2M",
    "3M","MM","M1","M2","M3","MPP2","AA","CO","WW","1W","2W","3W","PP2W","HH","1H","2H","3H","PP2H","OD","CL","EL","SL","SPL","LA","LOP","Total","Payable","Shift Allowance",
    "Transport Allowance",
    "OT Hours",
    "OT Amount",
    ]
    return columns

def get_data(args):
    data = []
    dates = get_dates(args)
    emp = frappe.get_all('Employee',{'status':'Active','vacant':0,'employee_type':args['employee_type']},['department','employee_number','employee_name','date_of_joining','designation','boarding_non_boarding','basic'])
    left_emp = frappe.get_all('Employee',{'status':'Left','relieving_date':('>=',args['from_date']),'vacant':0,'employee_type':args['employee_type']},['department','employee_number','employee_name','date_of_joining','designation','boarding_non_boarding','basic'])
    emp.extend(left_emp)

    for i in emp:
        row = []
        count_11 = 0
        count_22 = 0
        count_33 = 0
        count_12 = 0
        count_13 = 0
        count_21 = 0
        count_23 = 0
        count_31 = 0
        count_32 = 0
        count_1l1 = 0
        count_2l2 = 0
        count_3l3 = 0
        count_pp1 = 0
        count_pp2 = 0
        count_pp1l = 0
        count_pp2l = 0
        count_1m = 0
        count_2m = 0
        count_3m = 0
        count_mm =  0
        count_aa = 0
        count_1l2 = 0
        count_1l3 = 0
        count_2l1 = 0
        count_2l3 = 0
        count_3l1 = 0
        count_3l2 = 0
        count_plp = 0
        count_ww = 0
        count_co = 0
        count_m1 =  0
        count_m2 =  0
        count_m3 =  0
        count_mpp2 =  0
        count_hh =  0
        count_od =  0
        count_odw =  0
        count_odh = 0
        count_cl = 0
        count_el = 0
        count_sl = 0
        count_spl = 0
        count_la = 0
        count_lop = 0
        count_1w = 0
        count_2w = 0
        count_3w = 0
        count_pp2w = 0
        count_1h = 0
        count_2h = 0
        count_3h = 0
        count_pp2h = 0
        shift_2 = 0
        shift_3 = 0
        total_ot = timedelta(0,0,0)
        total_ot_hr = 0
        ot_amount = 0
        for date in dates:
            status = frappe.get_value('Attendance',{'employee':i.employee_number,'attendance_date':date},['shift_status','shift','status'])
            if status:
                sts = status[0]
                hh = check_holiday(date)
                if hh:
                    if status[1]:
                        sts = status[1]+hh
                    elif not status[1]:
                        if status[0] == 'OD':
                            sts = status[0]+hh
                        else:
                            sts = hh+hh
                    else:
                        sts = hh+hh
                
                if status[1] in ('2','PP2'):
                    if status[0] not in ('2M','M2'):
                        shift_2 += 1
                elif status[1] == '3':
                    if status[0] not in ('3M','M3'):
                        shift_3 += 1
                
                # elif status[2] == 'Half Day':
                #     if status[1] == '1':
                #         count_11 += 0.5
                #     elif status[1] == '2':
                #         count_22 += 0.5
                #     elif status[1] == '3':
                #         count_33 += 0.5
                #     elif status[1] == 'PP2':
                #         count_pp2 += 0.5
                #     elif status[1] == None:
                #         count_11 += 0.5
                if args['employee_type'] == "WC":
                    shift_status = wc_status_map.get(sts, "")
                else:
                    shift_status = bc_status_map.get(sts, "")
                if shift_status == '11':
                    count_11 += 1
                elif shift_status == '22':
                    count_22 += 1
                elif shift_status == '33':
                    count_33 += 1
                elif shift_status == '12':
                    count_12 += 1
                elif shift_status == '13':
                    count_13 += 1
                elif shift_status == '21':
                    count_21 += 1
                elif shift_status == '23':
                    count_23 += 1
                elif shift_status == '31':
                    count_31 += 1
                elif shift_status == '32':
                    count_32 += 1
                elif shift_status == '1L1':
                    count_1l1 += 1
                elif shift_status == '2L2':
                    count_2l2 += 1
                elif shift_status == '3L3':
                    count_3l3 += 1
                elif shift_status == '1L2':
                    count_1l2 += 1
                elif shift_status == '1L3':
                    count_1l3 += 1
                elif shift_status == '2L1':
                    count_2l1 += 1
                elif shift_status == '2L3':
                    count_2l3 += 1
                elif shift_status == '3L1':
                    count_3l1 += 1
                elif shift_status == '3L2':
                    count_3l2 += 1
                elif shift_status == 'P2LP2':
                    count_plp += 1
                elif shift_status == 'P1P1':
                    count_pp1 += 1
                elif shift_status == 'P2P2':
                    count_pp2 += 1
                elif shift_status == 'PP1L':
                    count_pp1l += 1
                elif shift_status == 'PP2L':
                    count_pp2l += 1
                elif shift_status == '1M':
                    count_1m += 1
                elif shift_status == '2M': 
                    count_2m += 1
                elif shift_status == '3M':
                    count_3m += 1
                elif shift_status == 'MM':
                    count_mm += 1
                elif shift_status == 'M1':
                    count_m1 += 1
                elif shift_status == 'M2':
                    count_m2 += 1
                elif shift_status == 'M3':
                    count_m3 += 1
                elif shift_status == 'MPP2':
                    count_mpp2 += 1
                elif shift_status == 'AA':
                    count_aa += 1
                elif shift_status == '1LM':
                    count_aa += 1
                elif shift_status == '2LM':
                    count_aa += 1
                elif shift_status == '3LM':
                    count_aa += 1
                elif shift_status == 'PP2LM':
                    count_aa += 1
                elif shift_status == 'CO':
                    count_co += 1
                elif shift_status == 'WW':
                    count_ww += 1
                elif shift_status == '1W':
                    count_1w += 1
                elif shift_status == '1LW':
                    count_1w += 1
                elif shift_status == '2W':
                    count_2w += 1
                elif shift_status == '3W':
                    count_3w += 1
                elif shift_status == 'PP2W':
                    count_pp2w += 1
                elif shift_status == 'HH':
                    count_hh += 1
                elif shift_status == '1H':
                    count_1h += 1
                elif shift_status == '2H':
                    count_2h += 1
                elif shift_status == '3H':
                    count_3h += 1
                elif shift_status == 'PP2H':
                    count_pp2h += 1
                elif shift_status == 'OD':
                    count_od += 1
                elif shift_status == 'ODW':
                    count_odw += 1
                elif shift_status == 'ODH':
                    count_odh += 1
                elif shift_status == 'CL':
                    count_cl += 1
                elif shift_status == '0.5CL':
                    count_cl += 0.5
                    count_11 += 0.5
                elif shift_status == 'SL':
                    count_sl += 1
                elif shift_status == '0.5SL':
                    count_sl += 0.5
                    count_11 += 0.5
                elif shift_status == 'EL':
                    count_el += 1
                elif shift_status == '0.5EL':
                    count_el += 0.5
                    count_11 += 0.5
                elif shift_status == 'SPL':
                    count_spl += 1
                elif shift_status == '0.5SPL':
                    count_spl += 0.5
                    count_11 += 0.5
                elif shift_status == 'LA':
                    count_la += 1
                elif shift_status == '0.5LA':
                    count_la += 0.5
                    count_11 += 0.5
                elif shift_status == '0.5LL':
                    count_la += 0.5
                    count_11 += 0.5
                elif shift_status == '0.5LLL':
                    count_la += 0.5
                    count_11 += 0.5
                elif shift_status == 'LOP':
                    count_lop += 1
                elif shift_status == '0.5LOP':
                    count_lop += 0.5
                    count_11 += 0.5

                ot = frappe.db.get_value("Overtime Request",{'ot_date':date,'employee':i.employee_number,'workflow_state':'Approved'},'ot_hours') or ''
                if ot:
                    total_ot += ot
                    day = ot.days * 24
                    hours = day + ot.seconds // 3600
                    minutes = (ot.seconds//60)%60
                    ftr = [3600,60,1]
                    hr = (sum([a*b for a,b in zip(ftr, map(int,str(str(hours) +':'+str(minutes)+':00').split(':')))]))/3600
                day = total_ot.days * 24
                hours = day + total_ot.seconds // 3600
                minutes = (total_ot.seconds//60)%60
                ftr = [3600,60,1]
                total_ot_hr = (sum([a*b for a,b in zip(ftr, map(int,str(str(hours) +':'+str(minutes)+':00').split(':')))]))/3600
                # basic_value = frappe.db.get_value("Employee",{'employee':i.employee_number,'employee_name':i.employee_name},['basic'])
                if args['employee_type'] != 'CL':
                    ot_amount = (((i.basic/26)/8)*2)*total_ot_hr
                else:
                    ot_amount = ((i.basic/8)*2)*total_ot_hr
            
        total = count_11+count_22+count_33+count_12+count_13+count_21+count_23+count_31+count_32+count_1l1+count_2l2+count_3l3+count_1l2+count_1l3+count_plp+count_2l1+count_2l3+count_3l1+count_3l2+count_pp1+count_pp2+count_pp1l+count_pp2l+count_1m+count_2m+count_3m+count_mm+count_m1+count_m2+count_m3+count_mpp2+count_aa+count_co+count_ww+count_1w+count_2w+count_3w+count_pp2w+count_hh+count_1h+count_2h+count_3h+count_pp2h+count_od+count_cl+count_el+count_sl+count_spl+count_la+count_lop
        if i.boarding_non_boarding != 'Boarding':
            if args['employee_type'] == "WC":
                days = add_days(i.dcount_11ate_of_joining,30)
                month_end = add_days(get_first_day(args['to_date']),24)
                if days.date() <= month_end:
                    transport = frappe.db.get_value('Designation',i.designation,'transport_allowance')
                else:
                    transport = (count_11+count_22+count_33+count_1l1+count_2l2+count_3l3+count_pp1+count_pp2+count_od+count_odw+count_odh+count_1w+count_2w+count_3w+count_pp2w+count_1h+count_2h+count_3h+count_pp2h+count_la) * 120
            else:
                transport = (count_11+count_22+count_33+count_1l1+count_2l2+count_3l3+count_pp1+count_pp2+count_od+count_odw+count_odh+count_1w+count_2w+count_3w+count_pp2w+count_1h+count_2h+count_3h+count_pp2h+count_la) * 120
        else:
            transport = 0
        
        if args['employee_type'] == "WC":
            payable = count_11+count_22+count_33+count_1l1+count_2l2+count_3l3+count_plp+count_pp1+count_pp2+count_co+count_ww+count_1w+count_2w+count_3w+count_pp2w+count_hh+count_1h+count_2h+count_3h+count_pp2h+count_od+count_cl+count_el+count_sl+count_spl
        else:
            payable = count_11+count_22+count_33+count_1l1+count_2l2+count_3l3+count_pp1+count_plp+count_pp2+count_co+count_hh+count_1h+count_2h+count_3h+count_pp2h+count_od+count_cl+count_el+count_sl+count_spl
        
        shift_allowance = shift_2 * 10 + shift_3 * 15

        row = [i.employee_number,i.employee_name,i.department,format_date(i.date_of_joining),count_11,count_22,count_33,count_12,count_13,count_21,count_23,count_31,count_32,count_1l1,count_2l2,count_3l3,count_1l2,count_1l3,count_2l1,count_2l3,count_3l1,count_3l2,count_pp1,count_pp2,count_pp1l,count_pp2l,count_1m,count_2m,count_3m,count_mm,count_m1,count_m2,count_m3,count_mpp2,count_aa,count_co,count_ww,count_1w,count_2w,count_3w,count_pp2w,count_hh,count_1h,count_2h,count_3h,count_pp2h,count_od,count_cl,count_el,count_sl,count_spl,count_la,count_lop,total,payable,shift_allowance,transport,total_ot_hr,floor(ot_amount)]
        data.append(row)
    return data

def check_holiday(date):
    holiday = frappe.db.sql("""select `tabHoliday`.holiday_date,`tabHoliday`.weekly_off from `tabHoliday List` 
    left join `tabHoliday` on `tabHoliday`.parent = `tabHoliday List`.name where `tabHoliday List`.name = 'Holiday List - 2021' and holiday_date = '%s' """%(date),as_dict=True)
    if holiday:
        if holiday[0].weekly_off == 1:
            return "W"
        else:
            return "H"

def get_dates(args):
    no_of_days = date_diff(add_days(args['to_date'], 1), args['from_date'])
    dates = [add_days(args['from_date'], i) for i in range(0, no_of_days)]
    return dates
    
bc_status_map = {
    "Absent": "AA",
    "AA":"AA",
    "Half Day": "HD",
    "Holiday": "HH",
    "Weekly Off": "WW",
    "WW":"WW",
    "1H": "1H",
    "2H": "2H",
    "3H": "3H",
    "1W": "1W",
    "2W": "2W",
    "3W": "3W",
    "PP1W": "PP1W",
    "PP2W": "PP2W",
    "1MW": "1W",
    "2MW": "2W",
    "3MW": "3W",
    "PP1MW": "PP1W",
    "PP2MW": "PP2W",
    "1LH": "1LH",
    "2LH": "2LH",
    "3LH": "3LH",
    "1LW": "1LW",
    "2LW": "2LW",
    "3LW": "3LW",
    "PP1LW": "PP1LW",
    "PP2LW": "PP2LW",
    "MW": "MW",
    "On Leave": "L",
    "Present": "P",
    "Work From Home": "WFH",
    "MM": "AA",
    "11": "11",
    "22": "22",
    "33": "33",
    "PP1PP1": "P1P1",
    "PP2PP2": "P2P2",
    "1L1": "1L1",
    "2L2": "2L2",
    "3L3": "3L3",
    "2L3": "2L3",
    "1L2": "1L2",
    "3L1": "3L1",
    "1LM": "1LM",
    "2LM": "2LM",
    "3LM": "3LM",
    "PP1LPP1": "P1LP1",
    "PP2LPP2": "P2LP2",
    "1M": "1M",
    "2M": "2M",
    "3M": "3M",
    "M1": "M1",
    "M2": "M2",
    "M3": "M3",
    "PP1M": "P1M",
    "PP2M": "P2M",
    "MPP1": "MP1",
    "MPP2": "MP2",
    "11": "11",
    "12": "12",
    "13": "13",
    "1PP1": "1P1",
    "1PP2": "1P2",
    "21": "21",
    "22": "22",
    "23": "23",
    "2PP1": "2P1",
    "2PP2": "2P2",
    "31": "31",
    "32": "32",
    "33": "33",
    "3PP1": "3P1",
    "3PP2": "3P2",
    "PP11": "P11",
    "PP12": "P12",
    "PP13": "P13",
    "PP1PP1": "P1P1",
    "PP1PP2": "P1P2",
    "PP21": "P2P1",
    "PP22": "P2P2",
    "PP23": "P2P3",
    "PP2PP1": "P2P1",
    "PP2PP2": "P2P2",
    "Earned Leave": "EL",
    "Casual Leave": "CL",
    "Sick Leave": "SL",
    "Special Leave": "SPL",
    "OD": "OD",
    "Compensatory Off": "CO",
    "Leave Without Pay": "LOP",
    "Earned LeaveW": "EL",
    "Casual LeaveW": "CL",
    "Sick LeaveW": "SL",
    "Special LeaveW": "SPL",
    "ODW": "ODW",
    "Compensatory OffW": "CO",
    "Leave Without PayW": "L)L",
    "Earned LeaveH": "EL",
    "Casual LeaveH": "CL",
    "Sick LeaveH": "SL",
    "Special LeaveH": "SPL",
    "ODH": "ODH",
    "Compensatory OffH": "CO",
    "Leave Without PayH": "LOP",
    "0.5Earned Leave": "0.5EL",
    "0.5Casual Leave": "0.5CL",
    "0.5Sick Leave": "0.5SL",
    "0.5Special Leave": "0.5SPL",
    "0.5Compensatory Off": "0.5CO",
    "0.5Leave Without PayW": "1W",
    "0.5Leave Without Pay": "0.5LL",
    "LEarned Leave/2": "0.5SL",
    "LCasual Leave/2": "0.5LCL",
    "LSick Leave/2": "0.5LSL",
    "LSpecial Leave/2": "0.5LSPL",
    "LCompensatory Off/2": "0.5LCO",
    "LLeave Without Pay/2": "0.5LLL",
    "AAW": "WW",
    "AAH": "HH",
    "HH": "HH",
    "1WW": "1W",
    "1HH": "1H",
    "2WW": "2W",
    "2HH": "2H",
    "3WW": "3W",
    "3HH": "3H",
    "PP2WW": "PP2W",
    "PP2HH": "PP2H",
    "ODWW": "ODW"
    }

wc_status_map = {
    "Absent": "AA",
    "AA": "AA",
    "Half Day": "HD",
    "Holiday": "HH",
    "WW":"WW",
    "1H": "1H",
    "2H": "2H",
    "3H": "3H",
    "1W": "1W",
    "2W": "2W",
    "3W": "3W",
    "PP1W": "PP1W",
    "PP2W": "PP2W",
    "1W": "1W",
    "2W": "2W",
    "3W": "3W",
    "PP1W": "PP1W",
    "PP2W": "PP2W",
    "1MW": "1W",
    "2MW": "2W",
    "3MW": "3W",
    "PP1MW": "PP1W",
    "PP2MW": "PP2W",
    "PP1H": "P1H",
    "PP2H": "P2H",
    "Weekly Off": "WW",
    "On Leave": "L",
    # "Present": "11",
    "Work From Home": "WFH",
    "OD": "OD",
    "Earned Leave": "EL",
    "Casual Leave": "CL",
    "Sick Leave": "SL",
    "Special Leave": "SPL",
    "Compensatory Off": "CO",
    "Leave Without Pay": "LOP",
    "Earned LeaveW": "EL",
    "Casual LeaveW": "CL",
    "Sick LeaveW": "SL",
    "Special LeaveW": "SPL",
    "ODW": "ODW",
    "Compensatory OffW": "CO",
    "Leave Without PayW": "LOP",
    "Earned LeaveH": "EL",
    "Casual LeaveH": "CL",
    "Sick LeaveH": "SL",
    "Special LeaveH": "SPL",
    "ODH": "ODH",
    "Compensatory OffH": "CO",
    "Leave Without PayH": "LOP",
    "1": "11",
    "2": "22",
    "3": "33",
    " ": "MM",
    "PP1": "P1P1",
    "PP2": "P2P2",
    "1L": "1L1",
    "2L": "2L2",
    "3L": "3L3",
    "PP1L": "P1LP1",
    "PP2L": "P2LP2",
    "1LH": "1LH",
    "2LH": "2LH",
    "3LH": "3LH",
    "1LW": "1W",
    "2LW": "2W",
    "3LW": "3W",
    "PP1LW": "PP1W",
    "PP2LW": "PP2W",
    "1M": "1M",
    "2M": "2M",
    "3M": "3M",
    "M1": "M1",
    "M2": "M2",
    "M3": "M3",
    "PP1M": "P1M",
    "PP2M": "P2M",
    "MPP1": "MP1",
    "MPP2": "MP2",
    "0.5Earned Leave": "0.5EL",
    "0.5Casual Leave": "0.5CL",
    "0.5Sick Leave": "0.5SL",
    "0.5Special Leave": "0.5SPL",
    "0.5Compensatory Off": "0.5CO",
    "0.5Leave Without Pay": "0.5LL",
    "LEarned Leave/2": "0.5SL",
    "LCasual Leave/2": "0.5LCL",
    "LSick Leave/2": "0.5LSL",
    "LSpecial Leave/2": "0.5LSPL",
    "LCompensatory Off/2": "0.5LCO",
    "LLeave Without Pay/2": "0.5LLL",
    "0.5Leave Without PayW": "1W",
    "AAW": "WW",
    "AAH": "HH",
    "HH": "HH",
    "ODWW": "ODW"
    }