from __future__ import unicode_literals
from os import stat
from unicodedata import category
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, touch_file
from frappe import _
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime, time

import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment,Border,Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types



@frappe.whitelist()
def download():
    filename = 'monthly ot amount report'
    test = build_xlsx_response(filename)


ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header_date = add_date(args)   
    ws.append(header_date)
    header_day = add_day(args)
    ws.append(header_day)
   
    data = get_data(args)

    for row in data:
        ws.append(row)
    
    dates = get_dates(args)

    align_center = Alignment(horizontal='center',vertical='center')
    for cell in ws["2:2"]:
        cell.alignment = align_center
        cell.fill = PatternFill(fgColor="ffff00", fill_type = "solid")
        
    align_center = Alignment(horizontal='center',vertical='center')
    for cell in ws["2:2"]:
        cell.alignment = align_center
        cell.font = Font(bold=True)
    for cell in ws['1:1']:
        cell.alignment = align_center
        cell.font = Font(bold=True)
    for cell in ws['A:A']:
        cell.alignment = align_center
        cell.font = Font(bold=True)
    
   

    iym = frappe.db.count('Department',{'parent_department':'IYM','is_group':0})
    re = frappe.db.count('Department',{'parent_department':'RE','is_group':0})
    ford = frappe.db.count('Department',{'parent_department':'FORD','is_group':0})
    support = frappe.db.count('Department',{'parent_department':'SUPPORT','is_group':0})
    departments = (iym+re+ford+support)
    contractors = frappe.db.count('Contractor',{'status':'Active'})

    border = Border(left=Side(border_style='thin', color='000000'),
             right=Side(border_style='thin', color='000000'),
             top=Side(border_style='thin', color='000000'),
             bottom=Side(border_style='thin', color='000000'))
    for rows in ws.iter_rows(min_row=1, max_row=departments+28+contractors, min_col=1, max_col=len(dates)+3):
        for cell in rows:
            cell.border = border


    ws.merge_cells(start_row=3, start_column=1, end_row=iym+5, end_column=1)
    ws.merge_cells(start_row=iym+6, start_column=1, end_row=iym+8+re, end_column=1)
    ws.merge_cells(start_row=iym+9+re, start_column=1, end_row=iym+11+ford+re, end_column=1)
    ws.merge_cells(start_row=iym+12+ford+re, start_column=1, end_row=iym+14+ford+re, end_column=1)
    ws.merge_cells(start_row=iym+15+ford+re, start_column=1, end_row=iym+14+ford+re+support+1, end_column=1)
    ws.merge_cells(start_row=iym+16+ford+re+support, start_column=1, end_row=iym+18+ford+re+support, end_column=1)
    ws.merge_cells(start_row=iym+16+ford+re+support+5, start_column=1, end_row=iym+18+ford+re+support+7, end_column=1)
    ws.merge_cells(start_row=iym+16+ford+re+support+12, start_column=1, end_row=iym+18+ford+re+support+9+contractors, end_column=1)


    

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column= 1)
    ws.merge_cells(start_row=1, start_column=len(dates)+3, end_row=2, end_column= len(dates)+3)



    for dept_name in ws.iter_rows(min_row=3, max_row=iym+2, min_col=2, max_col=2):
            for cell in dept_name:
                cell.fill = PatternFill(fgColor='ffff00', fill_type = "solid")

    for dept_name in ws.iter_rows(min_row=iym+2, max_row=iym+re+5, min_col=2, max_col=2):
            for cell in dept_name:
                cell.fill = PatternFill(fgColor='ffff00', fill_type = "solid")
    for dept_name in ws.iter_rows(min_row=iym+re+3, max_row=iym+8+re+ford, min_col=2, max_col=2):
            for cell in dept_name:
                cell.fill = PatternFill(fgColor='ffff00', fill_type = "solid")
    for dept_name in ws.iter_rows(min_row=iym+re+ford+3, max_row=iym+14+re+ford+support, min_col=2, max_col=2):
            for cell in dept_name:
                cell.fill = PatternFill(fgColor='ffff00', fill_type = "solid")
    
    


    for rows in ws.iter_rows(min_row=iym+3, max_row=iym+3, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='8ECA50', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym+6+re, max_row=iym+6+re, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='8ECA50', fill_type = "solid")
            cell.font = Font(bold=True)
    
    for rows in ws.iter_rows(min_row=iym+9+re+ford, max_row=iym+9+re+ford, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='8ECA50', fill_type = "solid")
            cell.font = Font(bold=True)
        
    for rows in ws.iter_rows(min_row=iym+12+re+ford, max_row=iym+12+re+ford, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='8ECA50', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym+15+re+ford+support, max_row=iym+15+re+ford+support, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='8ECA50', fill_type = "solid")
            cell.font = Font(bold=True)
    
    
    for rows in ws.iter_rows(min_row=iym+4, max_row=iym+4, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='D0D5CC', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym+7+re, max_row=iym+7+re, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='D0D5CC', fill_type = "solid")
            cell.font = Font(bold=True)
    
    for rows in ws.iter_rows(min_row=iym+10+re+ford, max_row=iym+10+re+ford, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='D0D5CC', fill_type = "solid")
            cell.font = Font(bold=True)
        
    for rows in ws.iter_rows(min_row=iym+13+re+ford, max_row=iym+13+re+ford, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='D0D5CC', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym+17+re+ford+support, max_row=iym+17+re+ford+support, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='D0D5CC', fill_type = "solid")
            cell.font = Font(bold=True)
    for total in ws.iter_rows(min_row=iym+16+re+ford+support, max_row=iym+16+re+ford+support, min_col=2, max_col=len(dates)+3):
        for cell in total:
            cell.fill = PatternFill(fgColor='4A934E', fill_type = "solid")
            cell.font = Font(bold=True)

    
    for rows in ws.iter_rows(min_row=iym+5, max_row=iym+5, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='46A1E9', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym+8+re, max_row=iym+8+re, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='46A1E9', fill_type = "solid")
            cell.font = Font(bold=True)
    
    for rows in ws.iter_rows(min_row=iym+11+re+ford, max_row=iym+11+re+ford, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='46A1E9', fill_type = "solid")
            cell.font = Font(bold=True)
        
    for rows in ws.iter_rows(min_row=iym+14+re+ford, max_row=iym+14+re+ford, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='46A1E9', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym+18+re+ford+support, max_row=iym+18+re+ford+support, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='46A1E9', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym+16+ford+re+support+5, max_row=iym+16+ford+re+support+9, min_col=2, max_col=2):
        for cell in rows:
            cell.fill = PatternFill(fgColor='ffff00', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym+16+ford+re+support+10, max_row=iym+16+ford+re+support+10, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='46A1E9', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym+16+ford+re+support+12, max_row=iym+16+ford+re+support+11+contractors, min_col=2, max_col=2):
        for cell in rows:
            cell.fill = PatternFill(fgColor='ffff00', fill_type = "solid")
            cell.font = Font(bold=True)
    
    for rows in ws.iter_rows(min_row=iym+16+ford+re+support+contractors+12, max_row=iym+16+ford+re+support+12+contractors, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='46A1E9', fill_type = "solid")
            cell.font = Font(bold=True)
    
    
    ws.freeze_panes='C3'
    ws.sheet_view.zoomScale = 80 

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def add_day(args):
    data = ['',"Day",]
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    for date in dates:
        dt = datetime.strptime(date,'%Y-%m-%d')
        day_format = datetime.date(dt).strftime('%a')
        data.append(day_format)
    return data
    
@frappe.whitelist()
def add_date(args):
    data = ['#',"Date",]
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    for date in dates:
        dt = datetime.strptime(date,'%Y-%m-%d')
        day_format = datetime.date(dt).strftime('%d-%m')
        data.append(day_format)
    data.append("Actual Amt")
    return data

@frappe.whitelist()
def get_data(args):
    data = []
    dept_group = ['IYM','RE','FORD',]
    mfg_dept_list = '('
    mfg_ot_amt_header = ['MFG','OT Amount']
    mfg_ot_percent_header=['','OT%']
    mfg_ot_sales_amt = ['',"Sales Amount"]
      
    for dg in dept_group:
        ot_amt_header = ['','OT Amount']
        ot_percent_header = ['','OT%']
        ot_sales_amt = ['',"Sales Amount"]
        departments = frappe.get_all("Department",{'parent_department':dg,"is_group":"0"},)
        dept_list = '('
        for dept in departments:
            dept_list = dept_list + '"' + str(dept.name) + '",'
            mfg_dept_list = mfg_dept_list + '"' + str(dept.name) + '",'
        dept_list = dept_list[:-1]
        dept_list = dept_list + ')'
        for dept in departments:
            row = [dg,dept.name]
            dates = get_dates(args)
            total_ot_amt = 0
            for date in dates:
                if args.employee_type:
                    ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
                    where `tabOvertime Request`.employee_type = '%s' and `tabOvertime Request`.department = '%s' and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(args.employee_type,dept.name,date),as_dict=True)[0].ot_amount or 0
                else:
                    ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
                    where `tabOvertime Request`.department = '%s' and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(dept.name,date),as_dict=True)[0].ot_amount or 0
                total_ot_amt += round(ot)
                row.append(round(ot))
            row.append(round(total_ot_amt))
                
            data.append(row)
        dates = get_dates(args)
        total_ot_amt = 0
        for date in dates:
            if args.employee_type:
                ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
                where `tabOvertime Request`.employee_type = '%s' and `tabOvertime Request`.department in %s and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(args.employee_type,dept_list,date),as_dict=True)[0].ot_amount or 0
            else:
                ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
                where `tabOvertime Request`.department in %s and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(dept_list,date),as_dict=True)[0].ot_amount or 0
            ot_amt_header.append(ot)
            total_ot_amt += round(ot)
            ot_percent_header.append("0")
            ot_sales_amt.append("0")
        ot_amt_header.append(total_ot_amt)

        data.append(ot_amt_header)
        data.append(ot_sales_amt)
        data.append(ot_percent_header)
    total_mfg_ot_amt = 0
    mfg_dept_list = mfg_dept_list[:-1]
    mfg_dept_list = mfg_dept_list + ')'
    for date in dates:
        if args.employee_type:
            mfg_ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
            where `tabOvertime Request`.employee_type = '%s' and `tabOvertime Request`.department in %s and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(args.employee_type,mfg_dept_list,date),as_dict=True)[0].ot_amount or 0
        else:
            mfg_ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
            where `tabOvertime Request`.department in %s and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(mfg_dept_list,date),as_dict=True)[0].ot_amount or 0
        total_mfg_ot_amt += mfg_ot
        mfg_ot_amt_header.append(mfg_ot)
        mfg_ot_percent_header.append("0")
        mfg_ot_sales_amt.append("0")
    mfg_ot_amt_header.append(total_mfg_ot_amt)
    # mfg_ot_percent_header.append("0")
    # mfg_ot_sales_amt.append("0")

    data.append(mfg_ot_amt_header)
    data.append(mfg_ot_sales_amt)
    data.append(mfg_ot_percent_header)

    support = get_support(args)
    
    for s in support:
        data.append(s)
    tsai = get_support_total(args)
    for ts in tsai:
        data.append(ts)

    return data
    
def get_dates(args):
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    return dates

def get_support(args):
    departments = frappe.get_all("Department",{'parent_department':'SUPPORT',"is_group":"0"},)
    support_data = []
    for dept in departments:
        row = ['SUPPORT',dept.name]
        dates = get_dates(args)
        total_ot_amt =0
        for date in dates:
            if args.employee_type:
                ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
                where `tabOvertime Request`.employee_type = '%s' and `tabOvertime Request`.department = '%s' and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(args.employee_type,dept.name,date),as_dict=True)[0].ot_amount or 0
            else:
                ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
                where `tabOvertime Request`.department = '%s' and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(dept.name,date),as_dict=True)[0].ot_amount or 0
            total_ot_amt += round(ot)
            row.append(round(ot))
        row.append(round(total_ot_amt))
        support_data.append(row)
    return support_data

def get_support_total(args):
    data = []
    dates = get_dates(args)
    support_total_header = ['','OT Amount']
    support_amount_header = ['','Sales Amount']
    support_percent_header = ['','OT %']
    dept_list = '('
    departments = frappe.get_all("Department",{'parent_department':'SUPPORT',"is_group":"0"},)
    for dept in departments:
        dept_list = dept_list + '"' + str(dept.name) + '",'
    dept_list = dept_list[:-1]
    dept_list = dept_list + ')'
    total_ot_amount = 0
    for date in dates:
        if args.employee_type:
            ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
            where `tabOvertime Request`.employee_type = '%s' and `tabOvertime Request`.department in %s and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(args.employee_type,dept_list,date),as_dict=True)[0].ot_amount or 0
        else:
            ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
            where `tabOvertime Request`.department in %s and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(dept_list,date),as_dict=True)[0].ot_amount or 0
        total_ot_amount += ot
        support_total_header.append(ot)
        support_amount_header.append("0")
        support_percent_header.append("0")
    support_total_header.append(total_ot_amount)
    data.append(support_total_header)
    data.append(support_amount_header)
    data.append(support_percent_header)

    tsai_total_header = ['TSAI','Grand Total']
    # tsai_amount_header = ['','Sales Amount']
    # tasi_percent_header = ['','OT %']

    dates = get_dates(args)
    row = []
    total_ot_amt =0
    for date in dates:
        if args.employee_type:
            ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
            where `tabOvertime Request`.employee_type = '%s' and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(args.employee_type,date),as_dict=True)[0].ot_amount or 0
        else:
            ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
            where `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(date),as_dict=True)[0].ot_amount or 0
        total_ot_amt += round(ot)
        tsai_total_header.append(round(ot))
        # tsai_amount_header.append("0")
        # tasi_percent_header.append("0")
    tsai_total_header.append(round(total_ot_amt))
    data.append(tsai_total_header)

    data.append([])
    data.append([])
    for emp_type in ['WC','BC','FT','NT','CL']:
        category_wise = ['Category',emp_type]
        cat_wise_actual = 0
        for date in dates:
            ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
                where `tabOvertime Request`.employee_type = '%s' and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(emp_type,date),as_dict=True)[0].ot_amount or 0
            category_wise.append(round(ot))
            cat_wise_actual += round(ot)
        category_wise.append(cat_wise_actual)
        data.append(category_wise)
    
    category_wise_total = ['','Total']
    cat_wise_actual_total = 0
    for date in dates:
        ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
            where `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(date),as_dict=True)[0].ot_amount or 0
        category_wise_total.append(round(ot))
        cat_wise_actual_total += round(ot)
    category_wise_total.append(cat_wise_actual_total)
    data.append(category_wise_total)

    data.append([])

    contractors = frappe.get_all('Contractor',{'status':'Active'})
    for con in contractors:
        contractor_wise = ['Contractor',con.name]
        con_wise_actual = 0
        for date in dates:
            ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
                where `tabOvertime Request`.employee_type = 'CL' and `tabOvertime Request`.contractor = '%s' and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(con.name,date),as_dict=True)[0].ot_amount or 0
            contractor_wise.append(round(ot))
            con_wise_actual += round(ot)
        contractor_wise.append(con_wise_actual)
        data.append(contractor_wise)
    
    contractor_wise_total = ['','Total']
    con_wise_actual_total = 0
    for date in dates:
        ot = frappe.db.sql("""select sum(ot_amount) as ot_amount from `tabOvertime Request`
                where `tabOvertime Request`.employee_type = 'CL' and `tabOvertime Request`.ot_date = '%s' and `tabOvertime Request`.workflow_state in ('Draft','Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """%(date),as_dict=True)[0].ot_amount or 0
        contractor_wise_total.append(round(ot))
        con_wise_actual_total += round(ot)
    contractor_wise_total.append(con_wise_actual_total)
    data.append(contractor_wise_total)

    return data
