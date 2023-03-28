import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def download(f_date,t_date):
    args = {'from_date':f_date,'to_date':t_date}
    filename = 'Monthly Manpower Plan vs Actual Report'
    # test = build_xlsx_response(filename)
    enqueue(build_xlsx_response, queue='default', timeout=6000, event='build_xlsx_response',filename=filename,args=args)



# return xlsx file object
def make_xlsx(data,args, sheet_name=None, wb=None, column_widths=None):
    # args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)

    report_title = ['Monthly Manpower Plan vs Actual Report']
    ws.append(report_title)

    header = add_header(args)
    ws.append(header)

    header = add_day_header(args)
    ws.append(header)

    header = add_header_emp_type(args)
    ws.append(header)

    data = get_data(args)

    for row in data:
        ws.append(row)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    dates = get_dates(args)
    i = 2
    for n in range(len(dates)):
        j = i + 2
        ws.merge_cells(start_row=2, start_column=i, end_row=2, end_column=j)
        ws.merge_cells(start_row=3, start_column=i, end_row=3, end_column=j)
        i += 3
    
    ws['A1'].fill = PatternFill(fgColor="27ae60", fill_type = "solid")

    align_center = Alignment(horizontal='center',vertical='center')
    for cell in ws["1:1"]:
        cell.alignment = align_center
    for cell in ws["2:2"]:
        cell.alignment = align_center
        cell.fill = PatternFill(fgColor="FFC300", fill_type = "solid")
    for cell in ws["3:3"]:
        cell.alignment = align_center
        cell.fill = PatternFill(fgColor="FFC300", fill_type = "solid")
    for cell in ws["4:4"]:
        cell.alignment = align_center
        cell.fill = PatternFill(fgColor="FFFF00", fill_type = "solid")

    for cell in ws["1:1"]:
        cell.font = Font(bold=True,size=23)
    for cell in ws["2:2"]:
        cell.font = Font(bold=True)
    for cell in ws["3:3"]:
        cell.font = Font(bold=True)
    for cell in ws["4:4"]:
        cell.font = Font(bold=True)

    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
    
    departments = frappe.db.count('Department',{'is_group':0})

    for rows in ws.iter_rows(min_row=1, max_row=departments+15, min_col=1, max_col=(len(dates)*3)+1):
        for cell in rows:
            cell.border = border
            # cell.alignment = Alignment(wrapText=True,horizontal='center')

    iym_direct = frappe.db.count('Department',{'is_group':0,'parent_department':'IYM','direct':1})
    re_direct = frappe.db.count('Department',{'is_group':0,'parent_department':'RE','direct':1})
    ford_direct = frappe.db.count('Department',{'is_group':0,'parent_department':'FORD','direct':1})
    support = frappe.db.count('Department',{'is_group':0,'parent_department':'SUPPORT'})

    iym_support = frappe.db.count('Department',{'is_group':0,'parent_department':'IYM','direct':0})
    re_support = frappe.db.count('Department',{'is_group':0,'parent_department':'RE','direct':0})
    ford_support = frappe.db.count('Department',{'is_group':0,'parent_department':'FORD','direct':0})

    for rows in ws.iter_rows(min_row=iym_direct+5, max_row=iym_direct+5, min_col=1, max_col=len(dates)*3+1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='d5dbdb', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+6, max_row=iym_direct+iym_support+6, min_col=1, max_col=len(dates)*3+1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='d5dbdb', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+7, max_row=iym_direct+iym_support+7, min_col=1, max_col=len(dates)*3+1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='58d68d', fill_type = "solid")
    
    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+8, max_row=iym_direct+iym_support+re_direct+8, min_col=1, max_col=len(dates)*3+1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='d5dbdb', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+9, max_row=iym_direct+iym_support+re_direct+re_support+9, min_col=1, max_col=len(dates)*3+1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='d5dbdb', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+10, max_row=iym_direct+iym_support+re_direct+re_support+10, min_col=1, max_col=len(dates)*3+1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='58d68d', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+11, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+11, min_col=1, max_col=len(dates)*3+1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='d5dbdb', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+12, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+12, min_col=1, max_col=len(dates)*3+1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='d5dbdb', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+13, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+13, min_col=1, max_col=len(dates)*3+1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='58d68d', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+14, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+14, min_col=1, max_col=len(dates)*3+1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='d5dbdb', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+15, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+15, min_col=1, max_col=len(dates)*3+1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='CCCCFF', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+16, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+16, min_col=1, max_col=len(dates)*3+1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='FFA07A', fill_type = "solid")

    ws.sheet_view.zoomScale = 80

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file    


def build_xlsx_response(filename,args):
    xlsx_file = make_xlsx(filename,args)
    # frappe.response['filename'] = filename + '.xlsx'
    # frappe.response['filecontent'] = xlsx_file.getvalue()
    # frappe.response['type'] = 'binary'
    ret = frappe.get_doc({
            "doctype": "File",
            "attached_to_name": '',
            "attached_to_doctype": 'Reports Dashboard',
            "attached_to_field": 'attach',
            "file_name": filename + '.xlsx',
            "is_private": 0,
            "content": xlsx_file.getvalue(),
            "decode": False
        })
    ret.save(ignore_permissions=True)
    # frappe.log_error(title='res',message=ret)
    frappe.db.commit()
    attached_file = frappe.get_doc("File", ret.name)
    frappe.db.set_value('Reports Dashboard',None,'attach',attached_file.file_url)


@frappe.whitelist()
def add_header(args):
    header = ["Date"]
    dates = get_dates(args)
    for date in dates:
        date = datetime.strptime(date,'%Y-%m-%d')
        date = date.strftime('%d-%b-%Y')
        header.extend([date,'',''])
    return header

@frappe.whitelist()
def add_day_header(args):
    day_header = ["Day"]
    dates = get_dates(args)
    for date in dates:
        date = datetime.strptime(date,'%Y-%m-%d')
        day = datetime.date(date).strftime('%a')
        day_header.extend([day,'',''])
    return day_header

@frappe.whitelist()
def add_header_emp_type(args):
    header = ["DEPT"]
    emp_type = ['Plan','Actual','Gap']
    dates = get_dates(args)
    for i in range(len(dates)):
        header.extend(emp_type)
    return header

def get_dates(args):
    no_of_days = date_diff(add_days(args['to_date'], 1), args['from_date'])
    dates = [add_days(args['from_date'], i) for i in range(0, no_of_days)]
    return dates

def get_data(args):
    data = []
    dept_group = ['IYM','RE','FORD']
    for dg in dept_group:
        i = 1
        for n in range(2):
            departments = frappe.get_all('Department',{'parent_department':dg,'is_group':0,'direct':i})
            for dept in departments:
                row = [dept.name]
                dates = get_dates(args)
                for date in dates:
                    plan = frappe.db.count("Shift Assignment",{'department':dept.name,'start_date':date,'employee_type':('!=','WC'),'docstatus':1})
                    actual = frappe.db.count("QR Checkin",{'department':dept.name,'shift_date':date,'employee_type':('!=','WC'),'ot':0})
                    row.extend([plan,actual,(plan-actual)])
                data.append(row)

            if i == 0:
                sub_total = ['TOTAL SUPPORT - ' + dg]
            else:
                sub_total = ['TOTAL DIRECT - ' + dg]


            dates = get_dates(args)
            for date in dates:
                plan = frappe.db.sql("""select count(*) as count from `tabShift Assignment` left join `tabDepartment` on `tabShift Assignment`.department = `tabDepartment`.name where `tabShift Assignment`.start_date = '%s' and `tabDepartment`.parent_department = '%s' and `tabDepartment`.direct = %s and `tabShift Assignment`.docstatus = 1 and `tabShift Assignment`.employee_type != 'WC' """%(date,dg,i),as_dict=True)[0].count or 0
                actual = frappe.db.sql("""select count(*) as count from `tabQR Checkin` left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabDepartment`.parent_department =  '%s' and `tabDepartment`.direct = %s and `tabQR Checkin`.ot = 0 """%(date,dg,i),as_dict=True)[0].count or 0
                sub_total.extend([plan,actual,plan - actual])
            data.append(sub_total)
            i = 0
            
        total_row = ['TOTAL - ' +dg]

        for date in dates:
            plan = frappe.db.sql("""select count(*) as count from `tabShift Assignment` left join `tabDepartment` on `tabShift Assignment`.department = `tabDepartment`.name where `tabShift Assignment`.start_date = '%s' and `tabDepartment`.parent_department = '%s' and `tabShift Assignment`.docstatus = 1 and `tabShift Assignment`.employee_type != 'WC' """%(date,dg),as_dict=True)[0].count or 0
            actual = frappe.db.sql("""select count(*) as count from `tabQR Checkin` left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabDepartment`.parent_department = '%s' and `tabQR Checkin`.ot = 0 """%(date,dg),as_dict=True)[0].count or 0
            total_row.extend([plan,actual,plan-actual])
        data.append(total_row)
    
    total_direct = get_total_direct(args)
    for t in total_direct:
        data.append(t)

    support_depts = get_support_depts(args)
    for s in support_depts:
        data.append(s)

    grand_total = get_grand_total(args)
    for g in grand_total:
        data.append(g)

    return data

def get_support_depts(args):
    departments = frappe.get_all('Department',{'parent_department':'SUPPORT','is_group':0})
    data = []
    dates = get_dates(args)
    for dept in departments:
        row = [dept.name]
        for date in dates:
            plan = frappe.db.sql("""select count(*) as count from `tabShift Assignment` left join `tabDepartment` on `tabShift Assignment`.department = `tabDepartment`.name where `tabShift Assignment`.start_date = '%s' and `tabDepartment`.name = '%s' and `tabShift Assignment`.docstatus = 1 and `tabShift Assignment`.employee_type != 'WC' """%(date,dept.name),as_dict=True)[0].count or 0
            actual = frappe.db.sql("""select count(*) as count from `tabQR Checkin` left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabDepartment`.name = '%s' and `tabQR Checkin`.ot = 0 """%(date,dept.name),as_dict=True)[0].count or 0
            row.extend([plan,actual,plan - actual])

        data.append(row)

    sub_total = ['TOTAL SUPPORT']
    for date in dates:
        plan = frappe.db.sql("""select count(*) as count from `tabShift Assignment` left join `tabDepartment` on `tabShift Assignment`.department = `tabDepartment`.name where `tabShift Assignment`.start_date = '%s' and `tabDepartment`.parent_department = 'SUPPORT' and `tabShift Assignment`.docstatus = 1 and `tabShift Assignment`.employee_type != 'WC' """%(date),as_dict=True)[0].count or 0
        actual = frappe.db.sql("""select count(*) as count from `tabQR Checkin` left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabDepartment`.parent_department = 'SUPPORT' and `tabQR Checkin`.ot = 0 """%(date),as_dict=True)[0].count or 0
        sub_total.extend([plan,actual,plan - actual])
    
    data.append(sub_total)
    return data


def get_grand_total(args):
    data = []
    row = ['GRAND TOTAL']
    dates = get_dates(args)
    for date in dates:
        plan = frappe.db.sql("""select count(*) as count from `tabShift Assignment` left join `tabDepartment` on `tabShift Assignment`.department = `tabDepartment`.name where `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.docstatus = 1 and `tabShift Assignment`.employee_type != 'WC' and `tabDepartment`.is_group = 0 and `tabDepartment`.parent_department in ('IYM','RE','FORD','SUPPORT') """%(date),as_dict=True)[0].count or 0
        actual = frappe.db.sql("""select count(*) as count from `tabQR Checkin` left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.ot = 0 """%(date),as_dict=True)[0].count or 0
        row.extend([plan,actual,plan - actual])
    
    data.append(row)
    return data

def get_total_direct(args):
    data = []
    row = ['TOTAL DIRECT (IYM+RE+FORD)']
    dates = get_dates(args)
    for date in dates:
        plan = frappe.db.sql("""select count(*) as count from `tabShift Assignment` left join `tabDepartment` on `tabShift Assignment`.department = `tabDepartment`.name where `tabShift Assignment`.start_date = '%s' and `tabDepartment`.parent_department in ('IYM','FORD','RE') and `tabShift Assignment`.docstatus = 1 and `tabShift Assignment`.employee_type != 'WC' """%(date),as_dict=True)[0].count or 0
        actual = frappe.db.sql("""select count(*) as count from `tabQR Checkin` left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabDepartment`.parent_department in ('IYM','FORD','RE') and `tabQR Checkin`.ot = 0 """%(date),as_dict=True)[0].count or 0
        row.extend([plan,actual,plan-actual])
    data.append(row)
    return data