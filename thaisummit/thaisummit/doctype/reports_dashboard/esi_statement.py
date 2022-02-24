# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def download():
    filename = 'ESI Statement'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    ws.append(["ESI STATEMENT",])
    ws.append(["For the month of October",])
    ws.append(['S.No','category','ID No','ESI No','Name','Paid Days','Wages','ESI','ER Cout','Total'])
    salary_slips = frappe.get_all("Salary Slip",{'employee_type':'WC','start_date':args.from_date,'end_date':args.to_date},['*'])
    i = 1
    for ss in salary_slips:
        esi_no = frappe.db.get_value('Employee',ss.employee,'esi_no')
        esi_amount = frappe.get_value('Salary Detail',{'salary_component':'Employee State insurance','parent':ss.name },['amount']) or 0
        er_cont = ss.gross_pay * 0.0325
        total = esi_amount + er_cont
        ws.append([i,ss.employee_type,ss.employee,esi_no,ss.employee_name,ss.payment_days,ss.gross_pay,esi_amount,er_cont,total])
        i += 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal='center')

    for cell in ws["2:2"]:
        cell.alignment = Alignment(horizontal='center')
    
    for cell in ws["3:3"]:
        cell.alignment = Alignment(horizontal='center')


    for cell in ws["A:A"]:
        cell.alignment = Alignment(horizontal='center')

    for cell in ws["B:B"]:
        cell.alignment = Alignment(horizontal='center')

    for cell in ws["C:C"]:
        cell.alignment = Alignment(horizontal='center')

    for cell in ws["D:D"]:
        cell.alignment = Alignment(horizontal='center')

    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))

    for rows in ws.iter_rows(min_row=1, max_row=len(salary_slips)+3, min_col=1, max_col=10):
        for cell in rows:
            cell.border = border

    for header in ws.iter_rows(min_row=2, max_row=2, min_col=8, max_col=13):
        for cell in header:
            cell.alignment = Alignment(horizontal='center')
            # cell.fill = PatternFill(fgColor='cec8ef', fill_type = "solid")

    for header in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=10):
        for cell in header:
            cell.font = Font(bold=True)


    ws = wb.create_sheet(sheet_name, 0)
    ws.append(["ESI STATEMENT",])
    ws.append(["For the month of October",])
    ws.append(['S.No','category','ID No','ESI No','Name','Paid Days','Wages','ESI','ER Cout','Total'])
    salary_slips = frappe.get_all("Salary Slip",{'employee_type':'BC'},['*'])
    i = 1
    for ss in salary_slips:
        esi_no = frappe.db.get_value('Employee',ss.employee,'esi_no')
        esi_amount = frappe.get_value('Salary Detail',{'salary_component':'Employee State insurance','parent':ss.name },['amount']) or 0
        er_cont = ss.gross_pay * 0.0325
        total = esi_amount + er_cont
        ws.append([i,ss.employee_type,ss.employee,esi_no,ss.employee_name,ss.payment_days,ss.gross_pay,esi_amount,er_cont,total])
        i += 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal='center')

    for cell in ws["2:2"]:
        cell.alignment = Alignment(horizontal='center')
    
    for cell in ws["3:3"]:
        cell.alignment = Alignment(horizontal='center')


    for cell in ws["A:A"]:
        cell.alignment = Alignment(horizontal='center')

    for cell in ws["B:B"]:
        cell.alignment = Alignment(horizontal='center')

    for cell in ws["C:C"]:
        cell.alignment = Alignment(horizontal='center')

    for cell in ws["D:D"]:
        cell.alignment = Alignment(horizontal='center')

    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))

    for rows in ws.iter_rows(min_row=1, max_row=len(salary_slips)+3, min_col=1, max_col=10):
        for cell in rows:
            cell.border = border

    for header in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=10):
        for cell in header:
            cell.fill = PatternFill(fgColor='cec8ef', fill_type = "solid")

    for header in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=10):
        for cell in header:
            cell.font = Font(bold=True)

    ws = wb.create_sheet(sheet_name, 0)
    ws.append(["ESI STATEMENT",])
    ws.append(["For the month of October",])
    ws.append(['S.No','category','ID No','ESI No','Name','Paid Days','Wages','ESI','ER Cout','Total'])
    salary_slips = frappe.get_all("Salary Slip",{'employee_type':'FT'},['*'])
    i = 1
    for ss in salary_slips:
        esi_no = frappe.db.get_value('Employee',ss.employee,'esi_no')
        esi_amount = frappe.get_value('Salary Detail',{'salary_component':'Employee State insurance','parent':ss.name },['amount']) or 0
        er_cont = ss.gross_pay * 0.0325
        total = esi_amount + er_cont
        ws.append([i,ss.employee_type,ss.employee,esi_no,ss.employee_name,ss.payment_days,ss.gross_pay,esi_amount,er_cont,total])
        i += 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal='center')

    for cell in ws["2:2"]:
        cell.alignment = Alignment(horizontal='center')
    
    for cell in ws["3:3"]:
        cell.alignment = Alignment(horizontal='center')


    for cell in ws["A:A"]:
        cell.alignment = Alignment(horizontal='center')

    for cell in ws["B:B"]:
        cell.alignment = Alignment(horizontal='center')

    for cell in ws["C:C"]:
        cell.alignment = Alignment(horizontal='center')

    for cell in ws["D:D"]:
        cell.alignment = Alignment(horizontal='center')

    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))

    for rows in ws.iter_rows(min_row=1, max_row=len(salary_slips)+3, min_col=1, max_col=10):
        for cell in rows:
            cell.border = border

    for header in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=10):
        for cell in header:
            cell.fill = PatternFill(fgColor='cec8ef', fill_type = "solid")
    
    for header in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=10):
        for cell in header:
            cell.font = Font(bold=True)
               
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'