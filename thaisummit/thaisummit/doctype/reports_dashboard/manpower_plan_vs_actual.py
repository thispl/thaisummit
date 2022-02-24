from __future__ import unicode_literals
from inspect import ArgSpec
import frappe
from frappe.exceptions import RetryBackgroundJobError
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from six import BytesIO, string_types

@frappe.whitelist()
def download():
    filename = 'Manpower Plan vs Actual Report'
    test = build_xlsx_response(filename)


# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)

    report_title = [format_date(args.date),'Manpower Plan vs Actual Report']
    ws.append(report_title)

    header = [args.shift,'WC','BC','FT','NT','CL','TOTAL','BC+FT+ NT+CL Plan','BC+FT+ NT+CL Actual','DIFF','%','OT PERSON','BC PLAN','BC ACTUAL','DIFF','%','NT PLAN','NT ACTUAL','DIFF','%','FT PLAN','FT ACTUAL','DIFF','%','CL PLAN','CL ACTUAL','DIFF','%']
    ws.append(header)

    data = get_data(args)

    for row in data:
        ws.append(row)
    
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=12)

    align_center = Alignment(horizontal='center',vertical='center')

    for cell in ws["1:1"]:
        cell.font = Font(bold=True,size=20)
        cell.alignment = align_center
        
    for cell in ws["2:2"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrapText=True,horizontal='center')

    for cell in ws["A:A"]:
        cell.font = Font(bold=True)

    ws['A1'].alignment = align_center
    ws['A2'].alignment = align_center
    ws['A1'].fill = PatternFill(fgColor="85c1e9", fill_type = "solid")
    ws['B1'].fill = PatternFill(fgColor="f9ff33", fill_type = "solid")
    ws['A2'].fill = PatternFill(fgColor="a9cce3", fill_type = "solid")

    iym_direct = frappe.db.count('Department',{'is_group':0,'parent_department':'IYM','direct':1})
    re_direct = frappe.db.count('Department',{'is_group':0,'parent_department':'RE','direct':1})
    ford_direct = frappe.db.count('Department',{'is_group':0,'parent_department':'FORD','direct':1})
    support = frappe.db.count('Department',{'is_group':0,'parent_department':'SUPPORT'})

    iym_support = frappe.db.count('Department',{'is_group':0,'parent_department':'IYM','direct':0})
    re_support = frappe.db.count('Department',{'is_group':0,'parent_department':'RE','direct':0})
    ford_support = frappe.db.count('Department',{'is_group':0,'parent_department':'FORD','direct':0})

    i = 11
    j = 1
    for n in range(5):
        for rows in ws.iter_rows(min_row=2, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+12, min_col=i, max_col=i):
            for cell in rows:
                cell.number_format = FORMAT_PERCENTAGE
        if j == 1:
            i += 5
            j = 0
        else:
            i += 4
    
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
    
    for rows in ws.iter_rows(min_row=1, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+13, min_col=1, max_col=28):
        for cell in rows:
            cell.border = border

    for rows in ws.iter_rows(min_row=iym_direct+3, max_row=iym_direct+3, min_col=1, max_col=1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+4, max_row=iym_direct+iym_support+5, min_col=1, max_col=1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+6, max_row=iym_direct+iym_support+re_direct+6, min_col=1, max_col=1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+7, max_row=iym_direct+iym_support+re_direct+re_support+8, min_col=1, max_col=1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, min_col=1, max_col=1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+10, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+11, min_col=1, max_col=1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+12, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+13, min_col=1, max_col=1):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")

    
    for rows in ws.iter_rows(min_row=iym_direct+3, max_row=iym_direct+3, min_col=2, max_col=7):
        for cell in rows:
            cell.fill = PatternFill(fgColor='27ae60', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+4, max_row=iym_direct+iym_support+5, min_col=2, max_col=7):
        for cell in rows:
            cell.fill = PatternFill(fgColor='27ae60', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+6, max_row=iym_direct+iym_support+re_direct+6, min_col=2, max_col=7):
        for cell in rows:
            cell.fill = PatternFill(fgColor='27ae60', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+7, max_row=iym_direct+iym_support+re_direct+re_support+8, min_col=2, max_col=7):
        for cell in rows:
            cell.fill = PatternFill(fgColor='27ae60', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, min_col=2, max_col=7):
        for cell in rows:
            cell.fill = PatternFill(fgColor='27ae60', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+10, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+11, min_col=2, max_col=7):
        for cell in rows:
            cell.fill = PatternFill(fgColor='27ae60', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+12, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+13, min_col=2, max_col=7):
        for cell in rows:
            cell.fill = PatternFill(fgColor='27ae60', fill_type = "solid")
            cell.font = Font(bold=True)
            

    for rows in ws.iter_rows(min_row=iym_direct+3, max_row=iym_direct+3, min_col=8, max_col=12):
        for cell in rows:
            cell.fill = PatternFill(fgColor='aed6f1', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+4, max_row=iym_direct+iym_support+5, min_col=8, max_col=12):
        for cell in rows:
            cell.fill = PatternFill(fgColor='aed6f1', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+6, max_row=iym_direct+iym_support+re_direct+6, min_col=8, max_col=12):
        for cell in rows:
            cell.fill = PatternFill(fgColor='aed6f1', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+7, max_row=iym_direct+iym_support+re_direct+re_support+8, min_col=8, max_col=12):
        for cell in rows:
            cell.fill = PatternFill(fgColor='aed6f1', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, min_col=8, max_col=12):
        for cell in rows:
            cell.fill = PatternFill(fgColor='aed6f1', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+10, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+11, min_col=8, max_col=12):
        for cell in rows:
            cell.fill = PatternFill(fgColor='aed6f1', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+12, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+13, min_col=8, max_col=12):
        for cell in rows:
            cell.fill = PatternFill(fgColor='aed6f1', fill_type = "solid")
            cell.font = Font(bold=True)

    
    for rows in ws.iter_rows(min_row=iym_direct+3, max_row=iym_direct+3, min_col=13, max_col=16):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+4, max_row=iym_direct+iym_support+5, min_col=13, max_col=16):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+6, max_row=iym_direct+iym_support+re_direct+6, min_col=13, max_col=16):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+7, max_row=iym_direct+iym_support+re_direct+re_support+8, min_col=13, max_col=16):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, min_col=13, max_col=16):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+10, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+11, min_col=13, max_col=16):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+12, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+13, min_col=13, max_col=16):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")
            cell.font = Font(bold=True)


    for rows in ws.iter_rows(min_row=iym_direct+3, max_row=iym_direct+3, min_col=17, max_col=20):
        for cell in rows:
            cell.fill = PatternFill(fgColor='fcbfa6', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+4, max_row=iym_direct+iym_support+5, min_col=17, max_col=20):
        for cell in rows:
            cell.fill = PatternFill(fgColor='fcbfa6', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+6, max_row=iym_direct+iym_support+re_direct+6, min_col=17, max_col=20):
        for cell in rows:
            cell.fill = PatternFill(fgColor='fcbfa6', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+7, max_row=iym_direct+iym_support+re_direct+re_support+8, min_col=17, max_col=20):
        for cell in rows:
            cell.fill = PatternFill(fgColor='fcbfa6', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, min_col=17, max_col=20):
        for cell in rows:
            cell.fill = PatternFill(fgColor='fcbfa6', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+10, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+11, min_col=17, max_col=20):
        for cell in rows:
            cell.fill = PatternFill(fgColor='fcbfa6', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+12, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+13, min_col=17, max_col=20):
        for cell in rows:
            cell.fill = PatternFill(fgColor='fcbfa6', fill_type = "solid")
            cell.font = Font(bold=True)


    for rows in ws.iter_rows(min_row=iym_direct+3, max_row=iym_direct+3, min_col=21, max_col=24):
        for cell in rows:
            cell.fill = PatternFill(fgColor='CD5C5C', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+4, max_row=iym_direct+iym_support+5, min_col=21, max_col=24):
        for cell in rows:
            cell.fill = PatternFill(fgColor='CD5C5C', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+6, max_row=iym_direct+iym_support+re_direct+6, min_col=21, max_col=24):
        for cell in rows:
            cell.fill = PatternFill(fgColor='CD5C5C', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+7, max_row=iym_direct+iym_support+re_direct+re_support+8, min_col=21, max_col=24):
        for cell in rows:
            cell.fill = PatternFill(fgColor='CD5C5C', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, min_col=21, max_col=24):
        for cell in rows:
            cell.fill = PatternFill(fgColor='CD5C5C', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+10, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+11, min_col=21, max_col=24):
        for cell in rows:
            cell.fill = PatternFill(fgColor='CD5C5C', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+12, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+13, min_col=21, max_col=24):
        for cell in rows:
            cell.fill = PatternFill(fgColor='CD5C5C', fill_type = "solid")
            cell.font = Font(bold=True)


    for rows in ws.iter_rows(min_row=iym_direct+3, max_row=iym_direct+3, min_col=25, max_col=28):
        for cell in rows:
            cell.fill = PatternFill(fgColor='f5b041', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+4, max_row=iym_direct+iym_support+5, min_col=25, max_col=28):
        for cell in rows:
            cell.fill = PatternFill(fgColor='f5b041', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+6, max_row=iym_direct+iym_support+re_direct+6, min_col=25, max_col=28):
        for cell in rows:
            cell.fill = PatternFill(fgColor='f5b041', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+7, max_row=iym_direct+iym_support+re_direct+re_support+8, min_col=25, max_col=28):
        for cell in rows:
            cell.fill = PatternFill(fgColor='f5b041', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+9, min_col=25, max_col=28):
        for cell in rows:
            cell.fill = PatternFill(fgColor='f5b041', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+10, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+11, min_col=25, max_col=28):
        for cell in rows:
            cell.fill = PatternFill(fgColor='f5b041', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+12, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+13, min_col=25, max_col=28):
        for cell in rows:
            cell.fill = PatternFill(fgColor='f5b041', fill_type = "solid")
            cell.font = Font(bold=True)



    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=7):
        for cell in rows:
            cell.fill = PatternFill(fgColor='27ae60', fill_type = "solid")
            cell.font = Font(bold=True)
    
    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=8, max_col=12):
        for cell in rows:
            cell.fill = PatternFill(fgColor='aed6f1', fill_type = "solid")
            cell.font = Font(bold=True)
    
    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=13, max_col=16):
        for cell in rows:
            cell.fill = PatternFill(fgColor='e6b0aa', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=17, max_col=20):
        for cell in rows:
            cell.fill = PatternFill(fgColor='fcbfa6', fill_type = "solid")
            cell.font = Font(bold=True)
    
    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=21, max_col=24):
        for cell in rows:
            cell.fill = PatternFill(fgColor='CD5C5C', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=25, max_col=28):
        for cell in rows:
            cell.fill = PatternFill(fgColor='f5b041', fill_type = "solid")
            cell.font = Font(bold=True)

    
    for rows in ws.iter_rows(min_row=3, max_row=iym_direct+iym_support+re_direct+re_support+ford_direct+ford_support+support+13, min_col=2, max_col=28):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'B3'
    ws.sheet_view.zoomScale = 80

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


def get_data(args):
    data = []
    dept_group = ['IYM','RE','FORD']
    for dg in dept_group:
        i = 1
        dg_list = []
        depts = frappe.get_all('Department',{'parent_department':dg,'is_group':0})
        for dp in depts:
            dg_list.append(dp.name)
        for n in range(2):
            departments = frappe.get_all('Department',{'parent_department':dg,'is_group':0,'direct':i})
            dept_list = []
            for dp in departments:
                dept_list.append(dp.name)
            for dept in departments:
                row = [dept.name]
                employee_type = ['WC','BC','FT','NT','CL','TT']
                total_count = 0
                for emp_type in employee_type:
                    if emp_type != 'TT':
                        if emp_type == 'WC':
                            if args.shift == 'Full Day':
                                count = frappe.db.count("Attendance",{'department':dept.name,'attendance_date': args.date,'employee_type':emp_type,'status':('in',('Present','Half Day'))})
                            else:
                                count = frappe.db.count("Attendance",{'department':dept.name,'attendance_date': args.date,'employee_type':emp_type,'shift':args.shift,'status':('in',('Present','Half Day'))})
                        else:
                            if args.shift == 'Full Day':
                                count = frappe.db.count("QR Checkin",{'department':dept.name,'shift_date': args.date,'employee_type':emp_type,'ot':0})
                            else:
                                count = frappe.db.count("QR Checkin",{'department':dept.name,'shift_date': args.date,'employee_type':emp_type,'qr_Shift':args.shift,'ot':0})
                        if count == 0:
                            row.append('')
                        else:
                            row.append(count)
                        total_count += count
                    else:
                        if total_count == 0:
                            row.append('')
                        else:
                            row.append(total_count)
                if args.shift == 'Full Day':
                    plan = frappe.db.count("Shift Assignment",{'department':dept.name,'start_date':args.date,'employee_type':('in',('BC','FT','NT','CL'))})
                    actual = frappe.db.count("QR Checkin",{'department':dept.name,'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':0})
                    actual_ot = frappe.db.count("QR Checkin",{'department':dept.name,'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':1})
                else:
                    plan = frappe.db.count("Shift Assignment",{'department':dept.name,'start_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'shift_type':args.shift})
                    actual = frappe.db.count("QR Checkin",{'department':dept.name,'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':0,'qr_shift':args.shift})
                    actual_ot = frappe.db.count("QR Checkin",{'department':dept.name,'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':1,'qr_shift':args.shift})
                diff = actual - plan
                try:
                    percent = round((diff/plan),2)
                except:
                    percent = 0
                
                if plan == 0:
                    row.append('')
                else:
                    row.append(plan)
                if actual == 0:
                    row.append('')
                else:
                    row.append(actual)
                if diff == 0:
                    row.append('')
                else:
                    row.append(diff)
                if percent == 0:
                    row.append('')
                else:
                    row.append(percent)
                if actual_ot == 0:
                    row.append('')
                else:
                    row.append(actual_ot)

                diffs = get_diff_count(args,dept.name)
                for d in diffs:
                    if d == 0:
                        row.append('')
                    else:
                        row.append(d)

                data.append(row)
            if i == 0:
                sub_total = ['TOTAL SUPPORT - ' + dg]
            else:
                sub_total = ['TOTAL DIRECT - ' + dg]

            employee_type = ['WC','BC','FT','NT','CL','TT']
            total_count = 0
            for emp_type in employee_type:
                if emp_type != 'TT':
                    if emp_type == 'WC':
                        if args.shift == 'Full Day':
                            count = frappe.db.sql("""select count(*) as count from `tabAttendance` 
                            left join `tabDepartment` on `tabAttendance`.department = `tabDepartment`.name 
                            where `tabAttendance`.attendance_date = '%s' and `tabAttendance`.status in ('Present','Half Day') and `tabAttendance`.employee_type = '%s' and `tabDepartment`.parent_department = '%s' and `tabDepartment`.direct = '%s' """%(args.date,emp_type,dg,i),as_dict=True)
                        else:
                            count = frappe.db.sql("""select count(*) as count from `tabAttendance` 
                            left join `tabDepartment` on `tabAttendance`.department = `tabDepartment`.name 
                            where `tabAttendance`.attendance_date = '%s' and `tabAttendance`.status in ('Present','Half Day') and `tabAttendance`.employee_type = '%s' and `tabDepartment`.parent_department = '%s' and `tabDepartment`.direct = '%s' and `tabAttendance`.shift = '%s' """%(args.date,emp_type,dg,i,args.shift),as_dict=True)
                    else:
                        if args.shift == 'Full Day':
                            count = frappe.db.sql("""select count(*) as count from `tabQR Checkin` 
                            left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name 
                            where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type = '%s' and `tabDepartment`.parent_department = '%s' and `tabDepartment`.direct = '%s' and `tabQR Checkin`.ot = 0 """%(args.date,emp_type,dg,i),as_dict=True)
                        else:
                            count = frappe.db.sql("""select count(*) as count from `tabQR Checkin` 
                            left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name 
                            where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type = '%s' and `tabDepartment`.parent_department = '%s' and `tabDepartment`.direct = '%s' and `tabQR Checkin`.qr_shift = %s and `tabQR Checkin`.ot = 0 """%(args.date,emp_type,dg,i,args.shift),as_dict=True)
                    if count:
                        count = count[0].count
                    else:
                        count = 0
                    sub_total.append(count)
                    total_count += count
                else:
                    sub_total.append(total_count)
            if args.shift == 'Full Day':      
                plan = frappe.db.sql("""select count(*) as count from `tabShift Assignment` 
                        left join `tabDepartment` on `tabShift Assignment`.department = `tabDepartment`.name 
                        where `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type in ('BC','FT','NT','CL') and `tabDepartment`.parent_department = '%s' and `tabDepartment`.direct = '%s' """%(args.date,dg,i),as_dict=True)
            else:
                plan = frappe.db.sql("""select count(*) as count from `tabShift Assignment` 
                    left join `tabDepartment` on `tabShift Assignment`.department = `tabDepartment`.name 
                    where `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type in ('BC','FT','NT','CL') and `tabDepartment`.parent_department = '%s' and `tabDepartment`.direct = '%s' and `tabShift Assignment`.shift_type = %s """%(args.date,dg,i,args.shift),as_dict=True)
            if plan:
                plan = plan[0].count
            else:
                plan = 0
            if args.shift == 'Full Day':
                actual = frappe.db.sql("""select count(*) as count from `tabQR Checkin`
                left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type in ('BC','FT','NT','CL') and `tabDepartment`.parent_department = '%s' and `tabDepartment`.direct = '%s' and `tabQR Checkin`.ot = 0 """%(args.date,dg,i),as_dict=True)
            else:
                actual = frappe.db.sql("""select count(*) as count from `tabQR Checkin`
                left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type in ('BC','FT','NT','CL') and `tabDepartment`.parent_department = '%s' and `tabDepartment`.direct = '%s' and `tabQR Checkin`.qr_shift = %s and `tabQR Checkin`.ot = 0 """%(args.date,dg,i,args.shift),as_dict=True)
            if actual:
                actual = actual[0].count
            else:
                actual = 0
            if args.shift == 'Full Day':
                actual_ot = frappe.db.sql("""select count(*) as count from `tabQR Checkin`
                left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' 
                and `tabQR Checkin`.employee_type in ('BC','FT','NT','CL') and `tabDepartment`.parent_department = '%s' and `tabQR Checkin`.ot = 1 and `tabDepartment`.direct = '%s' """%(args.date,dg,i),as_dict=True)
            else:
                actual_ot = frappe.db.sql("""select count(*) as count from `tabQR Checkin`
                left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' 
                and `tabQR Checkin`.employee_type in ('BC','FT','NT','CL') and `tabDepartment`.parent_department = '%s' and `tabQR Checkin`.ot = 1 and `tabDepartment`.direct = '%s' and `tabQR Checkin`.qr_shift = %s """%(args.date,dg,i,args.shift),as_dict=True)
            if actual_ot:
                actual_ot = actual_ot[0].count
            else:
                actual_ot = 0
            diff = actual - plan
            try:
                percent = round((diff/plan),2)
            except:
                percent = 0
            sub_total.append(plan)
            sub_total.append(actual)
            sub_total.append(diff)
            sub_total.append(percent)
            sub_total.append(actual_ot)

            diffs = get_sub_total_diff_count(args,dg,i)
            for d in diffs:
                sub_total.append(d)
            data.append(sub_total)
            i = 0

        total_row = ['GRAND TOTAL - ' +dg]
        total_count = 0
        employee_type = ['WC','BC','FT','NT','CL','TT']
        for emp_type in employee_type:
            if emp_type != 'TT':
                if emp_type == 'WC':
                    if args.shift == 'Full Day':
                        count = frappe.db.sql("""select count(*) as count from `tabAttendance` left join `tabDepartment` on `tabAttendance`.department = `tabDepartment`.name where `tabAttendance`.attendance_date = '%s' and `tabAttendance`.status in ('Present','Half Day') and `tabAttendance`.employee_type = '%s' and `tabDepartment`.parent_department =  '%s' """%(args.date,emp_type,dg),as_dict=True)
                    else:
                        count = frappe.db.sql("""select count(*) as count from `tabAttendance` left join `tabDepartment` on `tabAttendance`.department = `tabDepartment`.name where `tabAttendance`.attendance_date = '%s' and `tabAttendance`.status in ('Present','Half Day') and `tabAttendance`.employee_type = '%s' and `tabDepartment`.parent_department =  '%s' and `tabAttendance`.shift = %s """%(args.date,emp_type,dg,args.shift),as_dict=True)
                else:
                    if args.shift == 'Full Day':
                        count = frappe.db.sql("""select count(*) as count from `tabQR Checkin` left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type = '%s' and `tabDepartment`.parent_department =  '%s' and `tabQR Checkin`.ot = 0 """%(args.date,emp_type,dg),as_dict=True)
                    else:
                        count = frappe.db.sql("""select count(*) as count from `tabQR Checkin` left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type = '%s' and `tabDepartment`.parent_department =  '%s' and `tabQR Checkin`.qr_shift = %s and `tabQR Checkin`.ot = 0 """%(args.date,emp_type,dg,args.shift),as_dict=True)
                if count:
                    count = count[0].count
                else:
                    count = 0
                if count == 0:
                    total_row.append('')
                else:
                    total_row.append(count)
                total_count += count
            else:
                total_row.append(total_count)
        if args.shift == 'Full Day':
            plan = frappe.db.count("Shift Assignment",{'department':('in',dg_list),'start_date':args.date,'employee_type':('in',('BC','FT','NT','CL'))})
            actual = frappe.db.count("QR Checkin",{'department':('in',dg_list),'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':0})
            actual_ot = frappe.db.count("QR Checkin",{'department':('in',dg_list),'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':1})
        else:
            plan = frappe.db.count("Shift Assignment",{'department':('in',dg_list),'start_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'shift_type':args.shift})
            actual = frappe.db.count("QR Checkin",{'department':('in',dg_list),'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':0,'qr_shift':args.shift})
            actual_ot = frappe.db.count("QR Checkin",{'department':('in',dg_list),'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':1,'qr_shift':args.shift})
        diff = actual - plan
        try:
            percent = round((diff/plan),2)
        except:
            percent = 0
        total_row.append(plan)
        total_row.append(actual)
        total_row.append(diff)
        total_row.append(percent)
        total_row.append(actual_ot)
        diffs = get_grand_total_diff_count(args,dg_list)
        for d in diffs:
            total_row.append(d)
        data.append(total_row)

    support = get_support_depts(args)
    for s in support:
        data.append(s)
    return data

def get_support_depts(args):
    departments = frappe.get_all('Department',{'parent_department':'SUPPORT','is_group':0})
    dept_list = []
    data = []
    for dp in departments:
        dept_list.append(dp.name)
    for dept in departments:
        row = [dept.name]
        employee_type = ['WC','BC','FT','NT','CL','TT']
        total_count = 0
        # for emp_type in employee_type:
        #     if emp_type != 'TT':
        #         count = frappe.db.count("Shift Assignment",{'department':dept.name,'start_date':args.date,'employee_type':emp_type})
        #         if count == 0:
        #             row.append('')
        #         else:
        #             row.append(count)
        #         total_count += count
        #     else:
        #         row.append(total_count)
        for emp_type in employee_type:
            if emp_type != 'TT':
                if emp_type == 'WC':
                    if args.shift == 'Full Day':
                        count = frappe.db.sql("""select count(*) as count from `tabAttendance` where `tabAttendance`.attendance_date = '%s' and `tabAttendance`.status in ('Present','Half Day') and `tabAttendance`.employee_type = '%s' and department = '%s' """%(args.date,emp_type,dept.name),as_dict=True)
                    else:
                        count = frappe.db.sql("""select count(*) as count from `tabAttendance` where `tabAttendance`.attendance_date = '%s' and `tabAttendance`.status in ('Present','Half Day') and `tabAttendance`.employee_type = '%s' and `tabAttendance`.shift = %s and department = '%s' """%(args.date,emp_type,args.shift,dept.name),as_dict=True)
                else:
                    if args.shift == 'Full Day':
                        count = frappe.db.sql("""select count(*) as count from `tabQR Checkin` where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type = '%s' and `tabQR Checkin`.ot = 0 and department = '%s' """%(args.date,emp_type,dept.name),as_dict=True)
                    else:
                        count = frappe.db.sql("""select count(*) as count from `tabQR Checkin` where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type = '%s' and `tabQR Checkin`.qr_shift = %s and `tabQR Checkin`.ot = 0 and department = '%s' """%(args.date,emp_type,args.shift,dept.name),as_dict=True)
                if count:
                    count = count[0].count
                else:
                    count = 0
                if count == 0:
                    row.append('')
                else:
                    row.append(count)
                total_count += count
            else:
                row.append(total_count)
        if args.shift == 'Full Day':
            plan = frappe.db.count("Shift Assignment",{'department':dept.name,'start_date':args.date,'employee_type':('in',('BC','FT','NT','CL'))})
            actual = frappe.db.count("QR Checkin",{'department':dept.name,'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':0})
            actual_ot = frappe.db.count("QR Checkin",{'department':dept.name,'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':1})
        else:
            plan = frappe.db.count("Shift Assignment",{'department':dept.name,'start_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'shift_type':args.shift})
            actual = frappe.db.count("QR Checkin",{'department':dept.name,'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':0,'qr_shift':args.shift})
            actual_ot = frappe.db.count("QR Checkin",{'department':dept.name,'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':1,'qr_shift':args.shift})
        diff = actual - plan
        try:
            percent = round((diff/plan),2)
        except:
            percent = 0
        row.append(plan)
        row.append(actual)
        row.append(diff)
        row.append(percent)
        row.append(actual_ot)
        diffs = get_diff_count(args,dept.name)
        for d in diffs:
            row.append(d)
        data.append(row)
    sub_total = ['SUPPORT']


    employee_type = ['WC','BC','FT','NT','CL','TT']
    total_count = 0
    for emp_type in employee_type:
        if emp_type != 'TT':
            if emp_type == 'WC':
                if args.shift == 'Full Day':
                    count = frappe.db.sql("""select count(*) as count from `tabAttendance` left join `tabDepartment` on `tabAttendance`.department = `tabDepartment`.name where `tabAttendance`.attendance_date = '%s' and `tabAttendance`.status in ('Present','Half Day') and `tabAttendance`.employee_type = '%s' and `tabDepartment`.parent_department =  'SUPPORT' """%(args.date,emp_type),as_dict=True)
                else:
                    count = frappe.db.sql("""select count(*) as count from `tabAttendance` left join `tabDepartment` on `tabAttendance`.department = `tabDepartment`.name where `tabAttendance`.attendance_date = '%s' and `tabAttendance`.status in ('Present','Half Day') and `tabAttendance`.employee_type = '%s' and `tabDepartment`.parent_department =  'SUPPORT' and `tabAttendance`.shift = %s """%(args.date,emp_type,args.shift),as_dict=True)
            else:
                if args.shift == 'Full Day':
                    count = frappe.db.sql("""select count(*) as count from `tabQR Checkin` left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type = '%s' and `tabDepartment`.parent_department =  'SUPPORT' and `tabQR Checkin`.ot = 0 """%(args.date,emp_type),as_dict=True)
                else:
                    count = frappe.db.sql("""select count(*) as count from `tabQR Checkin` left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type = '%s' and `tabDepartment`.parent_department =  'SUPPORT' and `tabQR Checkin`.qr_shift = %s and `tabQR Checkin`.ot = 0 """%(args.date,emp_type,args.shift),as_dict=True)
            if count:
                count = count[0].count
            else:
                count = 0
            if count == 0:
                sub_total.append('')
            else:
                sub_total.append(count)
            total_count += count
        else:
            sub_total.append(total_count)
    # for emp_type in employee_type:
    #     if emp_type != 'TT':
    #         if args.shift == 'Full Day':
    #             count = frappe.db.count("Shift Assignment",{'department':('in',dept_list),'start_date':args.date,'employee_type':emp_type})
    #         else:
    #             count = frappe.db.count("Shift Assignment",{'department':('in',dept_list),'start_date':args.date,'employee_type':emp_type,'shift_type':args.shift})
    #         sub_total.append(count)
    #         total_count += count
    #     else:
    #         sub_total.append(total_count)
    if args.shift == 'Full Day':
        plan = frappe.db.count("Shift Assignment",{'department':('in',dept_list),'start_date':args.date,'employee_type':('in',('BC','FT','NT','CL'))})
        actual = frappe.db.count("QR Checkin",{'department':('in',dept_list),'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':0})
        actual_ot = frappe.db.count("QR Checkin",{'department':('in',dept_list),'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':1})
    else:
        plan = frappe.db.count("Shift Assignment",{'department':('in',dept_list),'start_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'shift_type':args.shift})
        actual = frappe.db.count("QR Checkin",{'department':('in',dept_list),'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':0,'qr_shift':args.shift})
        actual_ot = frappe.db.count("QR Checkin",{'department':('in',dept_list),'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':1,'qr_shift':args.shift})
    diff = actual - plan
    try:
        percent = round((diff/plan),2)
    except:
        percent = 0
    sub_total.append(plan)
    sub_total.append(actual)
    sub_total.append(diff)
    sub_total.append(percent)
    sub_total.append(actual_ot)

    diffs = get_sub_total_diff_count(args,'SUPPORT','0')
    for d in diffs:
        sub_total.append(d)
    data.append(sub_total)

    total_row = ['TSAI']
    total_count = 0
    for emp_type in employee_type:
        if emp_type != 'TT':
            if emp_type == 'WC':
                if args.shift == 'Full Day':
                    count = frappe.db.sql("""select count(*) as count from `tabAttendance` left join `tabDepartment` on `tabAttendance`.department = `tabDepartment`.name where `tabAttendance`.attendance_date = '%s' and `tabAttendance`.status in ('Present','Half Day') and `tabAttendance`.employee_type = '%s' and `tabDepartment`.parent_department in ('IYM','RE','FORD','SUPPORT') """%(args.date,emp_type),as_dict=True)
                else:
                    count = frappe.db.sql("""select count(*) as count from `tabAttendance` left join `tabDepartment` on `tabAttendance`.department = `tabDepartment`.name where `tabAttendance`.attendance_date = '%s' and `tabAttendance`.status in ('Present','Half Day') and `tabAttendance`.employee_type = '%s' and `tabDepartment`.parent_department in ('IYM','RE','FORD','SUPPORT') and `tabAttendance`.shift = %s """%(args.date,emp_type,args.shift),as_dict=True)
            else:
                if args.shift == 'Full Day':
                    count = frappe.db.sql("""select count(*) as count from `tabQR Checkin` where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type = '%s' and ot = 0 """%(args.date,emp_type),as_dict=True)
                else:
                    count = frappe.db.sql("""select count(*) as count from `tabQR Checkin` where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type = '%s' and `tabQR Checkin`.qr_shift = %s and ot = 0"""%(args.date,emp_type,args.shift),as_dict=True)
            if count:
                count = count[0].count
            else:
                count = 0
            if count == 0:
                total_row.append('')
            else:
                total_row.append(count)
            total_count += count
        else:
            total_row.append(total_count)
    if args.shift == 'Full Day':
        plan = frappe.db.count("Shift Assignment",{'start_date':args.date,'employee_type':('in',('BC','FT','NT','CL'))})
        actual = frappe.db.count("QR Checkin",{'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':0})
        actual_ot = frappe.db.count("QR Checkin",{'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':1})
    else:
        plan = frappe.db.count("Shift Assignment",{'start_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'shift_type':args.shift})
        actual = frappe.db.count("QR Checkin",{'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':0,'qr_shift':args.shift})
        actual_ot = frappe.db.count("QR Checkin",{'shift_date':args.date,'employee_type':('in',('BC','FT','NT','CL')),'ot':1,'qr_shift':args.shift})
    diff = actual - plan
    try:
        percent = round((diff/plan),2)
    except:
        percent = 0
    total_row.append(plan)
    total_row.append(actual)
    total_row.append(diff)
    total_row.append(percent)
    total_row.append(actual_ot)

    diffs = get_tsai_total(args)
    for d in diffs:
        total_row.append(d)

    data.append(total_row)
    return data


def get_diff_count(args,dept):
    emp_types = ['BC','NT','FT','CL']
    row = []
    for emp_type in emp_types:
        if args.shift == 'Full Day':
            plan = frappe.db.count("Shift Assignment",{'department':dept,'start_date':args.date,'employee_type':emp_type})
            actual = frappe.db.count("QR Checkin",{'department':dept,'shift_date':args.date,'employee_type':emp_type,'ot':0})
        else:
            plan = frappe.db.count("Shift Assignment",{'department':dept,'start_date':args.date,'employee_type':emp_type,'shift_type':args.shift})
            actual = frappe.db.count("QR Checkin",{'department':dept,'shift_date':args.date,'employee_type':emp_type,'ot':0,'qr_shift':args.shift})
        diff = actual - plan
        try:
            percent = round((diff/plan),2)
        except:
            percent = 0
        row.append(plan)
        row.append(actual)
        row.append(diff)
        row.append(percent)
    return row

def get_sub_total_diff_count(args,dg,i):
    emp_types = ['BC','NT','FT','CL']
    row = []
    for emp_type in emp_types:
        if args.shift == 'Full Day':
            plan = frappe.db.sql("""select count(*) as count from `tabShift Assignment` 
                left join `tabDepartment` on `tabShift Assignment`.department = `tabDepartment`.name 
                where `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type = '%s' and `tabDepartment`.parent_department = '%s' and `tabDepartment`.direct = '%s' """%(args.date,emp_type,dg,i),as_dict=True)
        else:
            plan = frappe.db.sql("""select count(*) as count from `tabShift Assignment` 
                left join `tabDepartment` on `tabShift Assignment`.department = `tabDepartment`.name 
                where `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type = '%s' and `tabDepartment`.parent_department = '%s' and `tabShift Assignment`.shift_type = %s and `tabDepartment`.direct = '%s' """%(args.date,emp_type,dg,args.shift,i),as_dict=True)
        if plan:
            plan = plan[0].count
        else:
            plan = 0
        if args.shift == 'Full Day':
            actual = frappe.db.sql("""select count(*) as count from `tabQR Checkin`
                left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type = '%s' and `tabDepartment`.parent_department = '%s' and `tabDepartment`.direct = '%s' and `tabQR Checkin`.ot = 0 """%(args.date,emp_type,dg,i),as_dict=True)
        else:
            actual = frappe.db.sql("""select count(*) as count from `tabQR Checkin`
                left join `tabDepartment` on `tabQR Checkin`.department = `tabDepartment`.name where `tabQR Checkin`.shift_date = '%s' and `tabQR Checkin`.employee_type = '%s' and `tabDepartment`.parent_department = '%s' and `tabDepartment`.direct = '%s' and `tabQR Checkin`.qr_shift = %s and `tabQR Checkin`.ot = 0"""%(args.date,emp_type,dg,i,args.shift),as_dict=True)
        if actual:
            actual = actual[0].count
        else:
            actual = 0
        diff = actual - plan
        try:
            percent = round((diff/plan),2)
        except:
            percent = 0
        row.append(plan)
        row.append(actual)
        row.append(diff)
        row.append(percent)
    return row

def get_grand_total_diff_count(args,dg_list):
    emp_types = ['BC','NT','FT','CL']
    row = []
    for emp_type in emp_types:
        if args.shift == 'Full Day':
            plan = frappe.db.count("Shift Assignment",{'department':('in',dg_list),'start_date':args.date,'employee_type':emp_type})
            actual = frappe.db.count("QR Checkin",{'department':('in',dg_list),'shift_date':args.date,'employee_type':emp_type,'ot':0})
        else:
            plan = frappe.db.count("Shift Assignment",{'department':('in',dg_list),'start_date':args.date,'employee_type':emp_type,'shift_type':args.shift})
            actual = frappe.db.count("QR Checkin",{'department':('in',dg_list),'shift_date':args.date,'employee_type':emp_type,'ot':0,'qr_shift':args.shift})
        diff = actual - plan
        try:
            percent = round((diff/plan),2)
        except:
            percent = 0
        row.append(plan)
        row.append(actual)
        row.append(diff)
        row.append(percent)
    return row

def get_tsai_total(args):
    emp_types = ['BC','NT','FT','CL']
    row = []
    for emp_type in emp_types:
        if args.shift == 'Full Day':
            plan = frappe.db.count("Shift Assignment",{'start_date':args.date,'employee_type':emp_type})
            actual = frappe.db.count("QR Checkin",{'shift_date':args.date,'employee_type':emp_type,'ot':0})
        else:
            plan = frappe.db.count("Shift Assignment",{'start_date':args.date,'employee_type':emp_type,'shift_type':args.shift})
            actual = frappe.db.count("QR Checkin",{'shift_date':args.date,'employee_type':emp_type,'ot':0,'qr_shift':args.shift})
        diff = actual - plan
        try:
            percent = round((diff/plan),2)
        except:
            percent = 0
        row.append(plan)
        row.append(actual)
        row.append(diff)
        row.append(percent)
    return row