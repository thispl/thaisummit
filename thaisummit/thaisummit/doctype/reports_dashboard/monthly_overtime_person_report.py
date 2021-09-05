from __future__ import unicode_literals
from os import stat
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, touch_file
from frappe import _
from frappe.utils.csvutils import UnicodeWriter, build_csv_response, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime, time

import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment,Border,Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def download():
    filename = 'monthly ot person report'
    test = build_xlsx_response(filename)


# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    header_date = add_date(args)   
    ws.append(header_date)
    header_day = add_day(args)
    ws.append(header_day)
   
    data = get_data(args)

    for row in data:
        ws.append(row)

    dates = get_dates(args)
    align_center = Alignment(horizontal='center',vertical='center')
    for cell in ws["2:2"]:
        cell.alignment = align_center
        cell.font = Font(bold=True)
    for cell in ws['1:1']:
        cell.alignment = align_center
        cell.font = Font(bold=True)
    for cell in ws['A:A']:
        cell.alignment = align_center
        cell.font = Font(bold=True)

    
        
    iym = frappe.db.count('Department',{'parent_department':'IYM','is_group':0})
    re = frappe.db.count('Department',{'parent_department':'RE','is_group':0})
    ford = frappe.db.count('Department',{'parent_department':'FORD','is_group':0})
    support = frappe.db.count('Department',{'parent_department':'SUPPORT','is_group':0})
    departments = (iym+re+ford+support)

    border = Border(left=Side(border_style='thin', color='000000'),
             right=Side(border_style='thin', color='000000'),
             top=Side(border_style='thin', color='000000'),
             bottom=Side(border_style='thin', color='000000'))
    for rows in ws.iter_rows(min_row=1, max_row=departments+8, min_col=1, max_col=len(dates)+3):
        for cell in rows:
            cell.border = border

   

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column= 1)
    ws.merge_cells(start_row=1, start_column=len(dates)+3, end_row=2, end_column= len(dates)+3)
    ws.merge_cells(start_row=iym+7+re+ford+support, start_column=1, end_row=iym+8+re+ford+support, end_column= 1)


    ws.merge_cells(start_row=3, start_column=1, end_row=iym+3, end_column=1)
    ws.merge_cells(start_row=iym+4, start_column=1, end_row=iym+4+re, end_column=1)
    ws.merge_cells(start_row=iym+5+re, start_column=1, end_row=iym+5+ford+re, end_column=1)
    ws.merge_cells(start_row=iym+6+ford+re, start_column=1, end_row=iym+6+ford+re, end_column=1)
    ws.merge_cells(start_row=iym+7+ford+re, start_column=1, end_row=iym+6+ford+re+support, end_column=1)

    for header in ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=len(dates)+2):
            for cell in header:
                cell.fill = PatternFill(fgColor='ffff00', fill_type = "solid")
                cell.font = Font(bold=True)

    for dept_name in ws.iter_rows(min_row=3, max_row=iym+2, min_col=2, max_col=2):
            for cell in dept_name:
                cell.fill = PatternFill(fgColor='ffff00', fill_type = "solid")
                cell.font = Font(bold=True)

    for dept_name in ws.iter_rows(min_row=iym+4, max_row=iym+re+3, min_col=2, max_col=2):
            for cell in dept_name:
                cell.fill = PatternFill(fgColor='ffff00', fill_type = "solid")
                cell.font = Font(bold=True)
    for dept_name in ws.iter_rows(min_row=iym+re+5, max_row=iym+4+re+ford, min_col=2, max_col=2):
            for cell in dept_name:
                cell.fill = PatternFill(fgColor='ffff00', fill_type = "solid")
                cell.font = Font(bold=True)
    for dept_name in ws.iter_rows(min_row=iym+re+ford+7, max_row=iym+6+re+ford+support, min_col=2, max_col=2):
            for cell in dept_name:
                cell.fill = PatternFill(fgColor='ffff00', fill_type = "solid")
                cell.font = Font(bold=True)
    
    


    for rows in ws.iter_rows(min_row=iym+3, max_row=iym+3, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='8ECA50', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym+4+re, max_row=iym+4+re, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='8ECA50', fill_type = "solid")
            cell.font = Font(bold=True)
    
    for rows in ws.iter_rows(min_row=iym+5+re+ford, max_row=iym+5+re+ford, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='8ECA50', fill_type = "solid")
            cell.font = Font(bold=True)
        
    for rows in ws.iter_rows(min_row=iym+6+re+ford, max_row=iym+6+re+ford, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='8ECA50', fill_type = "solid")
            cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=iym+7+re+ford+support, max_row=iym+7+re+ford+support, min_col=2, max_col=len(dates)+3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='8ECA50', fill_type = "solid")
            cell.font = Font(bold=True)
    
    

    for total in ws.iter_rows(min_row=iym+8+re+ford+support, max_row=iym+8+re+ford+support, min_col=2, max_col=len(dates)+3):
        for cell in total:
            cell.fill = PatternFill(fgColor='4A934E', fill_type = "solid")
            cell.font = Font(bold=True)
            
    ws.freeze_panes = 'C3'

    ws.sheet_view.zoomScale = 80 

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def add_day(args):
    data = ['',"Day"]
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    for date in dates:
        dt = datetime.strptime(date,'%Y-%m-%d')
        day_format = datetime.date(dt).strftime('%a')
        data.append(day_format)
    
    return data
    
@frappe.whitelist()
def add_date(args):
    data = ['#',"Date",]
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    for date in dates:
        dt = datetime.strptime(date,'%Y-%m-%d')
        day_format = datetime.date(dt).strftime('%d-%m')
        data.append(day_format)
    data.append('AVG')
    return data

def get_dates(args):
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    return dates

@frappe.whitelist()
def get_data(args):
    data = []
    dept_group = ['IYM','RE','FORD',]
    mfg_dept_list=[]
    for dg in dept_group:
        sum_header =['','SUM']
        mfg_header =['MFG','SUM']
        departments = frappe.get_all("Department",{'parent_department':dg,"is_group":"0"},)
        dept_list=[]
        for dept in departments:
            dept_list.append(dept.name)
            mfg_dept_list.append(dept.name)
            row = [dg,dept.name]
            dates = get_dates(args)
            avg =0
            # value=[]
            for date in dates:
                person = frappe.db.count("Overtime Request",{'department':dept.name,"ot_date":date,"workflow_state":"Approved"})
                row.append(person)
                avg += person
            total_avg = avg/len(dates)
            row.append(total_avg)
            data.append(row)
        
        avg=0
        for date in dates:
            total_sum = 0
            person_sum =frappe.db.count("Overtime Request",{'department':('in',(dept_list)),"ot_date":date,"workflow_state":"Approved"})
            total_sum += person_sum
            avg+=person_sum
            sum_header.append(total_sum)
        total_avg =avg/len(dates)
        sum_header.append(total_avg)
        data.append(sum_header)
    avg =0
    for date in dates:
        total_mfg_sum = 0
        mf_person_sum =frappe.db.count("Overtime Request",{'department':('in',(mfg_dept_list)),"ot_date":date,"workflow_state":"Approved"})
        total_mfg_sum+=mf_person_sum
        mfg_header.append(total_mfg_sum)
        avg+=mf_person_sum
    total_avg = avg/len(dates)
    mfg_header.append(total_avg)
    data.append(mfg_header)
    support = get_support(args)
    for s in support:
        data.append(s)
    grand = grand_total(args)
    for g in grand:
        data.append(g)
    return data
def get_support(args):
    departments = frappe.get_all("Department",{'parent_department':'SUPPORT',"is_group":"0"},)
    support_data = []
    support_dept_list = []
    tsai_sum_header=['TSAI','SUM']
    for dept in departments:
        row = ['SUPPORT',dept.name]
        dates = get_dates(args)
        support_dept_list.append(dept.name)
        avg=0
        for date in dates:
            support_person = frappe.db.count("Overtime Request",{"department":dept.name,"ot_date":date,"workflow_state":"Approved"})
            row.append(support_person) 
            avg+=support_person
        total_avg =avg/len(dates)
        row.append(total_avg)
        support_data.append(row)
        
    avg=0
    for date in dates:
        total_sum=0
        tsai_sum = frappe.db.count("Overtime Request",{"department":('in',(support_dept_list)),"ot_date":date,"workflow_state":"Approved"})
        total_sum += tsai_sum
        avg += tsai_sum
        tsai_sum_header.append(total_sum)
    total_avg =avg/len(dates)
    tsai_sum_header.append(total_avg)
    support_data.append(tsai_sum_header)
    
    return support_data

def grand_total(args):
    data = []
    dept_group = ['IYM','RE','FORD','SUPPORT']
    grand_total_header = ['','Grand Total']
    dept_list=[]
    for dg in dept_group:
        departments = frappe.get_all("Department",{'parent_department':dg,"is_group":"0"},)
        for dept in departments:
            dept_list.append(dept.name)
    dates = get_dates(args)
    avg=0
    for date in dates:
        total_sum = 0
        person = frappe.db.count("Overtime Request",{"department":('in',(dept_list)),"ot_date":date,"workflow_state":"Approved"})
        total_sum += person
        avg +=person
        grand_total_header.append(total_sum)
    total_avg =avg/len(dates)
    grand_total_header.append(total_avg)
    data.append(grand_total_header)
    return data



            

