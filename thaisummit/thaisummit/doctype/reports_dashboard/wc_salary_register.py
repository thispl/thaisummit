from __future__ import unicode_literals
from os import stat
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, touch_file
from frappe import _
from frappe.utils.csvutils import UnicodeWriter, build_csv_response, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime, time

import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment,Border,Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def download():
    filename = 'WC Salary Register'
    test = build_xlsx_response(filename)


# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    header_date = title1(args)   
    ws.append(header_date)
    header_day = title2(args)
    ws.append(header_day)
    header_column = get_column1()
    ws.append(header_column)
    header_column = get_column2()
    ws.append(header_column)
   
    data = get_data(args)
    for row in data:
        ws.append(row)

    # dates = get_dates(args)
    align_center = Alignment(horizontal='center',vertical='center')
    for cell in ws["2:2"]:
        cell.alignment = align_center
        cell.font = Font(bold=True)
    for cell in ws['1:1']:
        cell.alignment = align_center
        cell.font = Font(bold=True)
    for cell in ws['3:3']:
        cell.alignment = align_center
        cell.font = Font(bold=True)

    
    border = Border(left=Side(border_style='thin', color='000000'),
             right=Side(border_style='thin', color='000000'),
             top=Side(border_style='thin', color='000000'),
             bottom=Side(border_style='thin', color='000000'))
    for rows in ws.iter_rows(min_row=3, max_row=len(get_data(args))+4, min_col=1, max_col=len(get_column2())+4):
        for cell in rows:
            cell.border = border

   

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= 6)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column= 6)
    for i in range(9):
        i += 1
        ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column= i)
    ws.merge_cells(start_row=3, start_column=10, end_row=3, end_column= 19)
    ws.merge_cells(start_row=3, start_column=20, end_row=3, end_column= 36)
    ws.merge_cells(start_row=3, start_column=37, end_row=3, end_column= 46)
    # ws.merge_cells(start_row=3, start_column=47, end_row=4, end_column= 47)
    ws.merge_cells(start_row=3, start_column=48, end_row=4, end_column= 48)
    ws.merge_cells(start_row=3, start_column=49, end_row=4, end_column= 49)
    ws.merge_cells(start_row=3, start_column=50, end_row=4, end_column= 50)

    for header in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=len(get_column2())-2):
            for cell in header:
                cell.fill = PatternFill(fgColor='f9e1ed', fill_type = "solid")
    for header in ws.iter_rows(min_row=3, max_row=3, min_col=len(get_column2())-2, max_col=len(get_column2())):
            for cell in header:
                cell.fill = PatternFill(fgColor='fff808', fill_type = "solid")
                # cell.font = Font(bold=True)
    for header in ws.iter_rows(min_row=3, max_row=len(get_data(args))+4, min_col=10, max_col=19):
            for cell in header:
                cell.fill = PatternFill(fgColor='f9e1ed', fill_type = "solid")
    for header in ws.iter_rows(min_row=3, max_row=3, min_col=48, max_col=52):
            for cell in header:
                cell.fill = PatternFill(fgColor='fff808', fill_type = "solid")

            
    ws.freeze_panes = 'F5'

    ws.sheet_view.zoomScale = 80 

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def title1(args):
    data = ["THAI SUMMIT AUTO PARTS INDIA PVT LTD"]
    return data
    
@frappe.whitelist()
def title2(args):
    date = datetime.strptime(args.to_date,'%Y-%m-%d')
    data = ["SALARY STATEMENT FOR THE MONTH OF %s %s"%(date.strftime('%B'),date.strftime('%y'))]
    return data

def get_dates(args):
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    return dates


@frappe.whitelist()
def get_column2():
    data = ["","","","","","","","","","Basic Salary","HRA","Conveyance","Special Allowance","Medical Allowance",
    "LTA","Children Allowance","Chidren Hostel","Washing","Gross Salary","BAS","HRA","CON","SPA","MED","LTA","Children Allowance","Chidren Hostel","Washing","Position Allowance","ARR"," Other ARR","ATA","SHT","Transport Allowance","Additional Allowance","Other Allowance","Other(Extra hrs)","GROSS","PF","ESI",
    "CAN","PTAX","LWF","TDS","Arrears TDS","TEL EXP","Other Deduction","TOTAL"]
    return data

@frappe.whitelist()
def get_column1():
    data = ["SR No","Emp No","Cost Centre No","Cost Centre Name","DOJ","Employee_Name","Designation",
    "Payable","CTC","STANDARD","","","","","","","","","","EARNING","","","","","","","","","","","","","","","","","DEDUCTION","","","","","","","","","","","Total Deduction","Earned & Net Salary","Bonus","Other(Extra hrs)","For P Tax Earnings"]
    return data
@frappe.whitelist()
def get_data(args):
    data = []
    row = []
    basic_component_amount = earning_component_amount = deduction_component_amount = gross_wage = total_deduction = 0

    salary_comp = ['Basic','House Rent Allowance','Conveyance Allowance','Special Allowance','Medical Allowance','Leave Travel Allowance','Children Education','Children Hostel','Washing Allowance']

    earning_comp = ['Basic','House Rent Allowance','Conveyance Allowance','Special Allowance','Medical Allowance','Leave Travel Allowance','Children Education','Children Hostel','Washing Allowance','Position Allowance','Arrear','Other Arrear','Attendance Bonus','Shift Allowance','Transport Allowance','Additional Allowance','Other Allowance','Others']

    dedcution_comp = ['Provident Fund','Employee State Insurance','Canteen Charges','Professional Tax','Labor Welfare Fund','Tax Deducted at Source','Arrear TDS','TEL EXP','Other Deduction']

    if args.department:
        salary_slips = frappe.get_all("Salary Slip",{'employee_type':'WC','department':args.department,'start_date':args.from_date,'end_date':args.to_date},['*'])	

    if args.employee:
        salary_slips = frappe.get_all("Salary Slip",{'employee_type':'WC','employee':args.employee,'start_date':args.from_date,'end_date':args.to_date},['*'])	
    
    else:
        salary_slips = frappe.get_all("Salary Slip",{'employee_type':'WC','start_date':args.from_date,'end_date':args.to_date},['*'])	
    i =1
    for ss in salary_slips:
        row = [i,]
        cost_center = frappe.get_value('Department',ss.department,'cost_centre')
        emp = frappe.get_doc("Employee",ss.employee)
        row += [emp.name,cost_center,emp.department,emp.date_of_joining,emp.employee_name,emp.designation,ss.payment_days,round(emp.ctc),emp.basic,emp.house_rent_allowance,emp.conveyance_allowance,emp.special_allowance,emp.medical_allowance,emp.leave_travel_allowance,emp.children_education,emp.children_hostel,emp.washing_allowance,emp.basic+emp.house_rent_allowance+emp.conveyance_allowance+emp.special_allowance+emp.medical_allowance+emp.leave_travel_allowance+emp.children_education+emp.children_hostel+emp.washing_allowance]
            
        for ec in earning_comp:
            earning_component_amount = frappe.get_value('Salary Detail',{'salary_component':ec,'parent':ss.name},['amount'])
            if earning_component_amount:
                row.append(earning_component_amount)
            else:
                row.append('')
        row += [ss.gross_pay]
        total_deduction =0
        for dc in dedcution_comp:
            deduction_component_amount = frappe.get_value('Salary Detail',{'salary_component':dc,'parent':ss.name},['amount'])
            if deduction_component_amount:
                row += [deduction_component_amount]
                total_deduction += deduction_component_amount
            else:
                row += ['']
            bonus_value = frappe.get_value('Salary Detail',{'salary_component':'Basic','parent':ss.name},['amount']) or 0
            bonus = round(bonus_value/12)

        row.append(ss.total_deduction)
        row += [ ss.net_pay ]
        row += [ bonus ]
        row += [ frappe.get_value('Salary Detail',{'salary_component':'Others','parent':ss.name},['amount']) ]
        row += [ss.gross_pay]
        data.append(row)
        i+=1
        
    return data


            

