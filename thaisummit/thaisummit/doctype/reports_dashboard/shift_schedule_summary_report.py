# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

class ShiftScheduleSummaryReport(Document):
    pass

@frappe.whitelist()
def enqueue_download():
    enqueue(download, queue='default', timeout=6000, event='download')

@frappe.whitelist()
def download():
    filename = 'Shift Schedule Status Report'
    test = build_xlsx_response(filename)


# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header = add_header(args)
    ws.append(header)
    header_emp_type = add_header_emp_type()
    ws.append(header_emp_type)

    data = get_data(args)

    for row in data:
        ws.append(row)

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=8)
    ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=14)
    ws.merge_cells(start_row=1, start_column=15, end_row=1, end_column=20)
    ws.merge_cells(start_row=1, start_column=21, end_row=1, end_column=26)
    ws.merge_cells(start_row=1, start_column=27, end_row=1, end_column=32)

    contractors = frappe.get_all('Contractor',{'status':'Active'})
    i = 33
    for n in range(len(contractors)+2):
        j = i + 4
        ws.merge_cells(start_row=1, start_column=i, end_row=1, end_column=j)
        ws.cell(row=1,column=i).alignment = Alignment(horizontal='center')
        i += 5

    re = frappe.db.count('Department',{'parent_department':'RE','is_group':0})
    iym = frappe.db.count('Department',{'parent_department':'IYM','is_group':0})
    ford = frappe.db.count('Department',{'parent_department':'FORD','is_group':0})
    support = frappe.db.count('Department',{'parent_department':'SUPPORT','is_group':0})


    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=3, start_column=1, end_row=iym+3, end_column=1)
    ws.merge_cells(start_row=iym+4, start_column=1, end_row=re+iym+4, end_column=1)
    ws.merge_cells(start_row=iym+re+5, start_column=1, end_row=re+iym+ford+5, end_column=1)
    ws.merge_cells(start_row=iym+re+ford+6, start_column=1, end_row=iym+re+ford+support+6, end_column=1)

    align_center = Alignment(horizontal='center')
    for cell in ws["1:1"]:
        cell.alignment = align_center

    iym_group = ws['A3']
    iym_group.alignment = Alignment(vertical='center',horizontal='center')
    re_group = ws['A19']
    re_group.alignment = Alignment(vertical='center',horizontal='center')
    ford_group = ws['A32']
    ford_group.alignment = Alignment(vertical='center',horizontal='center')
    support_group = ws['A36']
    support_group.alignment = Alignment(vertical='center',horizontal='center')

    ws.freeze_panes = 'C3'

    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font
    for cell in ws["2:2"]:
        cell.font = bold_font
    for cell in ws["A:A"]:
        cell.font = bold_font
    for cell in ws["B:B"]:
        cell.font = bold_font

    
    ws['B1'].fill = PatternFill(fgColor="85c1e9", fill_type = "solid")

    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=3, max_col=32):
        for cell in rows:
            cell.fill = PatternFill(fgColor="f6fc2d", fill_type = "solid")

    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=3, max_col=32):
        for cell in rows:
            cell.fill = PatternFill(fgColor="52be80", fill_type = "solid")

    for rows in ws.iter_rows(min_row=1, max_row=1, min_col=33, max_col=(len(contractors)*5)+42):
        for cell in rows:
            cell.fill = PatternFill(fgColor="f8c471", fill_type = "solid")

    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=33, max_col=(len(contractors)*5)+42):
        for cell in rows:
            cell.fill = PatternFill(fgColor="97ea8b", fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym+3, max_row=iym+3, min_col=2, max_col=(len(contractors)*5)+42):
        for cell in rows:
            cell.fill = PatternFill(fgColor="d5d8dc", fill_type = "solid")
    
    for rows in ws.iter_rows(min_row=iym+re+4, max_row=iym+re+4, min_col=2, max_col=(len(contractors)*5)+42):
        for cell in rows:
            cell.fill = PatternFill(fgColor="d5d8dc", fill_type = "solid")

    for rows in ws.iter_rows(min_row=iym+re+ford+5, max_row=iym+re+ford+5, min_col=2, max_col=(len(contractors)*5)+42):
        for cell in rows:
            cell.fill = PatternFill(fgColor="d5d8dc", fill_type = "solid")
    
    for rows in ws.iter_rows(min_row=iym+re+ford+support+6, max_row=iym+re+ford+support+6, min_col=2, max_col=(len(contractors)*5)+42):
        for cell in rows:
            cell.fill = PatternFill(fgColor="d5d8dc", fill_type = "solid")
    
    for rows in ws.iter_rows(min_row=iym+re+ford+support+7, max_row=iym+re+ford+support+7, min_col=2, max_col=(len(contractors)*5)+42):
        for cell in rows:
            cell.fill = PatternFill(fgColor="FFA07A", fill_type = "solid")

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file    


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def add_header(args):
    header = ["Dept","%s"%(format_date(args.date)),'1','','','','','','2','','','','','','3','','','','','','PP2','','','','','','TOTAL','','','','','',]
    contractors = frappe.get_all('Contractor',{'status':'Active'})
    for contractor in contractors:
        header.extend([contractor.name,'','','',''])
    
    total = ['TOTAL','','','','','VACANT','','','','']
    header.extend(total)
    return header

@frappe.whitelist()
def add_header_emp_type():
    header = ["","",'WC','BC','FT','NT','CL','TT','WC','BC','FT','NT','CL','TT','WC','BC','FT','NT','CL','TT','WC','BC','FT','NT','CL','TT','WC','BC','FT','NT','CL','TT']
    shifts = ['1','2','3','PP2','TT']
    contractors = frappe.get_all('Contractor',{'status':'Active'})
    for i in range(len(contractors)+2):
        header.extend(shifts)
    return header

def get_data(args):
    data = []
    dept_group = ['IYM','RE','FORD','SUPPORT']
    for dg in dept_group:
        departments = frappe.get_all('Department',{'parent_department':dg,'is_group':0})
        dept_list = []
        for dp in departments:
            dept_list.append(dp.name)
        for dept in departments:
            row = [dg,dept.name]
            shifts = ['1','2','3','PP2','TOTAL']
            employee_type = ['WC','BC','FT','NT','CL','TT']
            contractors = frappe.get_all('Contractor',{'status':'Active'})
            for shift in shifts:
                if shift != 'TOTAL':
                    total_count = 0
                    for emp_type in employee_type:
                        if emp_type != 'TT':
                            count = frappe.db.count("Shift Assignment",{'department':dept.name,'start_date':args.date,'shift_type':shift,'employee_type':emp_type,'docstatus':'1'})
                            row.append(count)
                            total_count += count
                        else:
                            row.append(total_count)
                else:
                    total_count = 0
                    for emp_type in employee_type:
                        if emp_type != 'TT':
                            count = frappe.db.count("Shift Assignment",{'department':dept.name,'start_date':args.date,'employee_type':emp_type,'docstatus':'1'})
                            total_count += count
                            row.append(count)
                        else:
                            row.append(total_count)

            for contractor in contractors:
                shifts = ['1','2','3','PP2','TT']
                total_count = 0
                for shift in shifts:
                    if shift != 'TT':
                        count = frappe.db.sql("select count(*) as count from `tabShift Assignment` left join `tabEmployee` on `tabShift Assignment`.employee = `tabEmployee`.name where `tabEmployee`.contractor = '%s' and `tabShift Assignment`.shift_type = '%s' and `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type = 'CL' and `tabShift Assignment`.department = '%s' "%(contractor.name,shift,args.date,dept.name),as_dict=True)
                        if count:
                            count = count[0].count
                        else:
                            count = 0
                        row.append(count)
                        total_count += count
                    else:
                        row.append(total_count)
            total_count = 0
            for shift in shifts:
                if shift != 'TT':
                    count = frappe.db.sql("select count(*) as count from `tabShift Assignment` left join `tabEmployee` on `tabShift Assignment`.employee = `tabEmployee`.name where `tabShift Assignment`.shift_type = '%s' and `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type = 'CL' and `tabShift Assignment`.department = '%s' "%(shift,args.date,dept.name),as_dict=True)
                    if count:
                        count = count[0].count
                    else:
                        count = 0
                    row.append(count)
                    total_count += count
                else:
                    row.append(total_count)
            total_count = 0
            for shift in shifts:
                if shift != 'TT':
                    count = frappe.db.sql("select count(*) as count from `tabShift Assignment` left join `tabEmployee` on `tabShift Assignment`.employee = `tabEmployee`.name where `tabShift Assignment`.shift_type = '%s' and `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type = 'CL' and `tabShift Assignment`.department = '%s' and `tabEmployee`.vacant = 1 "%(shift,args.date,dept.name),as_dict=True)
                    if count:
                        count = count[0].count
                    else:
                        count = 0
                    row.append(count)
                    total_count += count
                else:
                    row.append(total_count)
            data.append(row)

        row = [dg,'TOTAL ' + dg]
        shifts = ['1','2','3','PP2','TOTAL']
        employee_type = ['WC','BC','FT','NT','CL','TT']
        for shift in shifts:
            if shift != 'TOTAL':
                total_count = 0
                for emp_type in employee_type:
                    if emp_type != 'TT':
                        count = frappe.db.count("Shift Assignment",{'department':('in',dept_list),'start_date':args.date,'shift_type':shift,'employee_type':emp_type,'docstatus':'1'})
                        total_count += count
                        row.append(count)
                    else:
                        row.append(total_count)
            else:
                total_count = 0
                for emp_type in employee_type:
                    if emp_type != 'TT':
                        count = frappe.db.count("Shift Assignment",{'department':('in',dept_list),'start_date':args.date,'employee_type':emp_type,'docstatus':'1'})
                        total_count += count
                        row.append(count)
                    else:
                        row.append(total_count)
        for contractor in contractors:
                shifts = ['1','2','3','PP2','TT']
                total_count = 0
                depts = str(dept_list).replace('[','(')
                depts = str(depts).replace(']',')')
                for shift in shifts:
                    if shift != 'TT':
                        count = frappe.db.sql("select count(*) as count from `tabShift Assignment` left join `tabEmployee` on `tabShift Assignment`.employee = `tabEmployee`.name where `tabEmployee`.contractor = '%s' and `tabShift Assignment`.shift_type = '%s' and `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type = 'CL' and `tabShift Assignment`.department in %s "%(contractor.name,shift,args.date,depts),as_dict=True)
                        if count:
                            count = count[0].count
                        else:
                            count = 0
                        row.append(count)
                        total_count += count
                    else:
                        row.append(total_count)
        total_count = 0
        for shift in shifts:
            if shift != 'TT':
                count = frappe.db.sql("select count(*) as count from `tabShift Assignment` left join `tabEmployee` on `tabShift Assignment`.employee = `tabEmployee`.name where `tabShift Assignment`.shift_type = '%s' and `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type = 'CL' and `tabShift Assignment`.department in %s "%(shift,args.date,depts),as_dict=True)
                if count:
                    count = count[0].count
                else:
                    count = 0
                row.append(count)
                total_count += count
            else:
                row.append(total_count)
        total_count = 0
        for shift in shifts:
            if shift != 'TT':
                count = frappe.db.sql("select count(*) as count from `tabShift Assignment` left join `tabEmployee` on `tabShift Assignment`.employee = `tabEmployee`.name where `tabShift Assignment`.shift_type = '%s' and `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type = 'CL' and `tabShift Assignment`.department in %s and `tabEmployee`.vacant = 1 "%(shift,args.date,depts),as_dict=True)
                if count:
                    count = count[0].count
                else:
                    count = 0
                row.append(count)
                total_count += count
            else:
                row.append(total_count)
        data.append(row)

    row = ['','TOTAL ']
    shifts = ['1','2','3','PP2','TOTAL']
    employee_type = ['WC','BC','FT','NT','CL','TT']
    for shift in shifts:
        if shift != 'TOTAL':
            total_count = 0
            for emp_type in employee_type:
                if emp_type != 'TT':
                    count = frappe.db.count("Shift Assignment",{'start_date':args.date,'shift_type':shift,'employee_type':emp_type,'docstatus':'1'})
                    total_count += count
                    row.append(count)
                else:
                    row.append(total_count)
        else:
            total_count = 0
            for emp_type in employee_type:
                if emp_type != 'TT':
                    count = frappe.db.count("Shift Assignment",{'start_date':args.date,'employee_type':emp_type,'docstatus':'1'})
                    total_count += count
                    row.append(count)
                else:
                    row.append(total_count)
    for contractor in contractors:
        shifts = ['1','2','3','PP2','TT']
        total_count = 0
        for shift in shifts:
            if shift != 'TT':
                count = frappe.db.sql("select count(*) as count from `tabShift Assignment` left join `tabEmployee` on `tabShift Assignment`.employee = `tabEmployee`.name where `tabEmployee`.contractor = '%s' and `tabShift Assignment`.shift_type = '%s' and `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type = 'CL' "%(contractor.name,shift,args.date),as_dict=True)
                if count:
                    count = count[0].count
                else:
                    count = 0
                row.append(count)
                total_count += count
            else:
                row.append(total_count)
    total_count = 0
    for shift in shifts:
        if shift != 'TT':
            count = frappe.db.sql("select count(*) as count from `tabShift Assignment` left join `tabEmployee` on `tabShift Assignment`.employee = `tabEmployee`.name where `tabShift Assignment`.shift_type = '%s' and `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type = 'CL' "%(shift,args.date),as_dict=True)
            if count:
                count = count[0].count
            else:
                count = 0
            row.append(count)
            total_count += count
        else:
            row.append(total_count)
    total_count = 0
    for shift in shifts:
        if shift != 'TT':
            count = frappe.db.sql("select count(*) as count from `tabShift Assignment` left join `tabEmployee` on `tabShift Assignment`.employee = `tabEmployee`.name where `tabShift Assignment`.shift_type = '%s' and `tabShift Assignment`.start_date = '%s' and `tabShift Assignment`.employee_type = 'CL' and `tabEmployee`.vacant = 1 "%(shift,args.date),as_dict=True)
            if count:
                count = count[0].count
            else:
                count = 0
            row.append(count)
            total_count += count
        else:
            row.append(total_count)
    data.append(row)
    return data
