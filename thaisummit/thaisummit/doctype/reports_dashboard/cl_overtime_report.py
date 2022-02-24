import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, touch_file
from frappe import _
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime, time

import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment,Border,Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def download():
    filename = 'CL Overtime Report (Shift Continue)'
    # args = {'from_date':f_date,'to_date':t_date}
    build_xlsx_response(filename)
    # enqueue(build_xlsx_response, queue='default', timeout=6000, event='build_xlsx_response',filename=filename,args=args)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)

    ws.append(['CL Overtime Report (Shift Continue)'])
    date_header = ['','DATE']
    shift_header = ['','Shift']

    dates = get_dates(args)
    for date in dates:
        dt = datetime.strptime(date,'%Y-%m-%d')
        d = datetime.date(dt).strftime('%d-%b')
        date_header.extend([d,''])
        shift_header.extend(['2','3'])
    date_header.extend(['Total'])
    shift_header.extend(['Total Shift 2','Total Shift 3'])
    ws.append(date_header)
    ws.append(shift_header)
    total_dept = 0
    dept_group = ['IYM','RE','FORD','SUPPORT']
    for dept in dept_group:
        depts = frappe.get_all('Department',{'parent_department':dept})
        for d in depts:
            row = [dept,d.name]
            total_dept += 0
            dates = get_dates(args)
            total_shift_2 = 0
            total_shift_3 = 0 
            for date in dates:
                shift_2 = frappe.db.count('QR Checkin',{'department':d.name,'ot':1,'shift_date':date,'qr_shift':'2'})
                shift_3 = frappe.db.count('QR Checkin',{'department':d.name,'ot':1,'shift_date':date,'qr_shift':'3'})
                total_shift_2 += shift_2
                total_shift_3 += shift_3
                row.extend([shift_2,shift_3])
            row.extend([total_shift_2,total_shift_3])
            ws.append(row)
    
    row = ['','Total']
    dates = get_dates(args)
    total_shift_2 = 0
    total_shift_3 = 0 
    for date in dates:
        shift_2 = frappe.db.count('QR Checkin',{'ot':1,'shift_date':date,'qr_shift':'2'})
        shift_3 = frappe.db.count('QR Checkin',{'ot':1,'shift_date':date,'qr_shift':'3'})
        row.extend([shift_2,shift_3])
        total_shift_2 += shift_2
        total_shift_3 += shift_3 
    row.extend([total_shift_2,total_shift_3])
    ws.append(row)

       

    ws.freeze_panes='C4'
    ws.sheet_view.zoomScale = 90 

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.merge_cells(start_row=2, start_column=len(date_header), end_row=2, end_column=len(date_header)+1)
    i = 3
    for date in dates:
        ws.merge_cells(start_row=2, start_column=i, end_row=2, end_column=i+1)
        i += 2

    c = 4
    e = 3
    for d in dept_group:
        count =frappe.db.count('Department',{'parent_department':d})
        e += count
        ws.merge_cells(start_row=c, start_column=1, end_row=e, end_column=1)
        c += count

    for rows in ws.iter_rows(min_row = 2, max_row = 2, min_col=len(date_header), max_col=len(date_header)+1):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')

    for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=5):
         for cell in header:
             cell.fill = PatternFill(fgColor='ADFF2F', fill_type = "solid")
    for header in ws.iter_rows(min_row=4, max_row=total_dept, min_col=2, max_col=2):
         for cell in header:
             cell.fill = PatternFill(fgColor='fefe33', fill_type = "solid")
    for header in ws.iter_rows(min_row=3, max_row=3, min_col=3, max_col=len(shift_header)):
         for cell in header:
             cell.fill = PatternFill(fgColor='FFA07A', fill_type = "solid")

    for cell in ws["A:A"]:
        cell.alignment = Alignment(vertical='center')
    for cell in ws["2:2"]:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws["2:2"]:
        cell.font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    for cell in ws["3:3"]:
        cell.font = Font(bold=True)
    for cell in ws["A:A"]:
        cell.font = Font(bold=True)
    for cell in ws["54:54"]:
        cell.font = Font(bold=True)


    # border = Border(left=Side(border_style='thin', color='000000'),
    #     right=Side(border_style='thin', color='000000'),
    #     top=Side(border_style='thin', color='000000'),
    #     bottom=Side(border_style='thin', color='000000'))

    # for rows in ws.iter_rows(min_row=2, max_row = total_dept, min_col=1, max_col=len(shift_header)):
    #     for cell in rows:
    #         cell.border = border


    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

# def build_xlsx_response(filename,args):
#     xlsx_file = make_xlsx(filename,args)
#     ret = frappe.get_doc({
#             "doctype": "File",
#             "attached_to_name": '',
#             "attached_to_doctype": 'Reports Dashboard',
#             "attached_to_field": 'attach',
#             "file_name": filename + '.xlsx',
#             "is_private": 0,
#             "content": xlsx_file.getvalue(),
#             "decode": False
#         })
#     ret.save(ignore_permissions=True)
#     frappe.db.commit()
#     frappe.log_error(message=ret)
#     attached_file = frappe.get_doc("File", ret.name)
#     frappe.db.set_value('Reports Dashboard',None,'attach',attached_file.file_url)


def get_dates(args):
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    return dates