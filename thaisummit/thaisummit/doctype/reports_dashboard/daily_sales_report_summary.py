# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types



@frappe.whitelist()
def download():
    filename = 'Daily Sales Report'
    test = build_xlsx_response(filename)


# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header = ["DAILY SALES REPORT"]
    ws.append(header)

    group_header = add_header()
    ws.append(group_header)

    data = get_data(args)

    for row in data:
        ws.append(row)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    align_center = Alignment(horizontal='center',vertical='center')
    for cell in ws["2:2"]:
        cell.alignment = align_center
    for cell in ws["3:3"]:
        cell.alignment = align_center
    for cell in ws["A:A"]:
        cell.alignment = align_center
    for cell in ws["B:B"]:
        cell.alignment = align_center

    ws['A1'].alignment = Alignment(vertical='center',horizontal='center')

    bold_font = Font(bold=True,size=20)
    for cell in ws["1:1"]:
        cell.font = bold_font
    for cell in ws["2:2"]:
        cell.font = Font(bold=True)

    
    ws['A1'].fill = PatternFill(fgColor="dc7633", fill_type = "solid")

    dates = get_dates(args)
    for rows in ws.iter_rows(min_row=2, max_row=len(dates)+2, min_col=1, max_col=2):
        for cell in rows:
            cell.fill = PatternFill(fgColor="d6dbdf", fill_type = "solid")
    
    ws.merge_cells(start_row=len(dates)+4, start_column=1, end_row=len(dates)+4, end_column=2)

    for rows in ws.iter_rows(min_row=2, max_row=len(dates)+2, min_col=3, max_col=3):
        for cell in rows:
            cell.fill = PatternFill(fgColor='3498db', fill_type = "solid")
    
    for rows in ws.iter_rows(min_row=2, max_row=len(dates)+2, min_col=4, max_col=4):
        for cell in rows:
            cell.fill = PatternFill(fgColor='7dcea0', fill_type = "solid")

    for rows in ws.iter_rows(min_row=2, max_row=len(dates)+2, min_col=5, max_col=5):
        for cell in rows:
            cell.fill = PatternFill(fgColor='edbb99', fill_type = "solid")

    for rows in ws.iter_rows(min_row=2, max_row=len(dates)+2, min_col=6, max_col=6):
        for cell in rows:
            cell.fill = PatternFill(fgColor='d6dbdf', fill_type = "solid")

    for rows in ws.iter_rows(min_row=len(dates)+3, max_row=len(dates)+3, min_col=1, max_col=6):
        for cell in rows:
            cell.fill = PatternFill(fgColor='faf423', fill_type = "solid")

    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
         
    for rows in ws.iter_rows(min_row=1, max_row=len(dates)+3, min_col=1, max_col=6):
        for cell in rows:
            cell.border = border

    ws.sheet_view.zoomScale = 80

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file    


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def get_dates(args):
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    return dates



@frappe.whitelist()
def add_header():
    group_header = ['Date','Day','TOTAL IYM','TOTAL RE','TOTAL FORD','TOTAL']
    return group_header

def get_data(args):
    data = []
    groups = ['IYM','RE','FORD']
    hps_parts = ('10000280','10000291')
    dates = get_dates(args)
    hps_total = 0
    for date in dates:
        row = []
        row_total = 0
        dt = datetime.strptime(date,'%Y-%m-%d')
        d = dt.strftime('%d-%b-%Y')
        day = datetime.date(dt).strftime('%a')
        hps_out = frappe.db.sql(""" select sum(sales_value) as sales_value from `tabSAP Outgoing Report` where ar_invoice_date = '%s' and part_no in %s """%(date,hps_parts),as_dict=True)
        if hps_out[0].sales_value is not None:
            hps_total += hps_out[0].sales_value
            hps_out = hps_out[0].sales_value
        else:
            hps_out = 0
        row.extend([d,day])
        row_total = 0
        for g in groups:
            sap_out = frappe.db.sql(""" select sum(sales_value) as sales_value from `tabSAP Outgoing Report` where ar_invoice_date = '%s' and customer_group = '%s' """%(date,g),as_dict=True)
            if sap_out[0].sales_value is not None:
                row_total += sap_out[0].sales_value+hps_out
                row.append(sap_out[0].sales_value+hps_out)
            else:
                row.append('-')
            hps_out = 0
        if row_total == 0:
            row_total = '-'
        row.append(row_total)

        data.append(row)
    total = ['TOTAL','']
    group_total = 0
    for g in groups:
        sap_out = frappe.db.sql(""" select sum(sales_value) as sales_value from `tabSAP Outgoing Report` where ar_invoice_date in %s and customer_group = '%s' """%(tuple(dates),g),as_dict=True)
        if sap_out[0].sales_value is not None:
            group_total += sap_out[0].sales_value+hps_total
            total.append(sap_out[0].sales_value+hps_total)
        else:
            total.append('-')
        hps_total = 0
    if group_total == 0:
        total.append('-')
    total.append(group_total)
    data.append(total)

    return data
