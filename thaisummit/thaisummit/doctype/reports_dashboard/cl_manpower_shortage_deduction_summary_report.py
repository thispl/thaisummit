from __future__ import unicode_literals
import math
from os import SCHED_RESET_ON_FORK
from babel import dates
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
from pytz import AmbiguousTimeError
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from six import BytesIO, string_types


@frappe.whitelist()
def download():
    filename = 'cl plan shortage'
    test = build_xlsx_response(filename)
    # enqueue(build_xlsx_response, queue='default', timeout=6000, event='build_xlsx_response',filename=filename)

    
# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    header = add_day_header(args)
    ws.append(header)

    header = add_header(args)
    ws.append(header)

    header = add_header_emp_type(args)
    ws.append(header)

    data = get_data(args)

    for row in data:
        ws.append(row)

    dates = get_dates(args)
    i = 2
    for n in range(len(dates)+1):
        j = i + 4
        ws.merge_cells(start_row=1, start_column=i, end_row=1, end_column=j)
        ws.cell(row=1,column=i).alignment = Alignment(horizontal='center')

        ws.merge_cells(start_row=2, start_column=i, end_row=2, end_column=j)
        ws.cell(row=2,column=i).alignment = Alignment(horizontal='center')
        i += 5
    
    contractors = frappe.get_all('Contractor',{'status':'Active'})
    i = 5
    for n in range(len(dates)+1):
        for rows in ws.iter_rows(min_row=4, max_row=len(contractors)+4, min_col=i, max_col=i):
            for cell in rows:
                cell.number_format = FORMAT_PERCENTAGE
        i += 5
    

    ws.freeze_panes = 'B4'

    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font
        cell.fill = PatternFill(fgColor="f1fc04", fill_type = "solid")
    for cell in ws["2:2"]:
        cell.font = bold_font
        cell.fill = PatternFill(fgColor="f1fc04", fill_type = "solid")
    for cell in ws["3:3"]:
        cell.font = bold_font
        cell.fill = PatternFill(fgColor="f1fc04", fill_type = "solid")
    for cell in ws["A:A"]:
        cell.font = bold_font

    contractors = frappe.get_all('Contractor',{'status':'Active'})
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
         
    for rows in ws.iter_rows(min_row=1, max_row=len(contractors)+4, min_col=1, max_col=(len(dates)*5)+6):
        for cell in rows:
            cell.border = border

    for rows in ws.iter_rows(min_row=1, max_row=len(contractors)+4, min_col=1, max_col=(len(dates)*5)+6):
        for cell in rows:
            cell.alignment = Alignment(wrapText=True,horizontal='center')

    # for rows in ws.iter_rows(min_row=4, max_row=len(contractors)+4, min_col=2, max_col=(len(dates)*5)+6):
    #     for cell in rows:
    #         cell.alignment = Alignment(horizontal='center')

    for rows in ws.iter_rows(min_row=len(contractors)+4, max_row=len(contractors)+4, min_col=2, max_col=(len(dates)*5)+6):
        for cell in rows:
            cell.font = bold_font

    ws.sheet_view.zoomScale = 70

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file 

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def add_header(args):
    header = ['']
    dates = get_dates(args)
    for date in dates:
        date = datetime.strptime(date,'%Y-%m-%d')
        date = date.strftime('%d-%b-%Y')
        header.extend([date,'','','',''])
    header.append('OVERALL TOTAL')
    return header

@frappe.whitelist()
def add_day_header(args):
    day_header = ['']
    dates = get_dates(args)
    for date in dates:
        date = datetime.strptime(date,'%Y-%m-%d')
        day = datetime.date(date).strftime('%a')
        day_header.extend([day,'','','',''])
    return day_header

@frappe.whitelist()
def add_header_emp_type(args):
    header = [""]
    emp_type = ['PLAN','ACTUAL','DIFF','Shortage','Amount to be Debited']
    dates = get_dates(args)
    for i in range(len(dates)):
        header.extend(emp_type)
    header.extend(emp_type)
    return header


@frappe.whitelist()
def get_data(args):
    data = []
    contractors = frappe.get_all('Contractor',{'status':'Active'})
    for contractor in contractors:
        row_plan = 0
        row_actual = 0
        side_diff = 0
        row = [contractor.name]
        dates =get_dates(args)
        for date in dates:
            sum=0
            if frappe.db.exists('CL Head Count Plan',{'contractor':contractor.name,'date':date}):
                plan = frappe.get_doc('CL Head Count Plan',{'contractor':contractor.name,'date':date})
                sum = plan.shift_1 + plan.shift_2+plan.shift_3+plan.shift_pp1+plan.shift_pp2
                
            count = frappe.db.sql("""select count(*) as count from `tabQR Checkin` 
                                    left join `tabEmployee` on `tabQR Checkin`.employee = `tabEmployee`.name
                                     where `tabEmployee`.contractor='%s' and  `tabQR Checkin`.shift_date='%s' and ot = 0
                                    """%(contractor.name,date),as_dict=True)
            if count:
                count = count[0].count
            else:
                count = 0
            diff = count - sum
            try:
                shortage = diff /sum
            except ZeroDivisionError:
                shortage = 0
            short_percent = round(shortage,2)
            if diff < 0:
                deduction = frappe.db.get_single_value('Deduction Settings', 'cl_deduction')
                amt_debt = diff * deduction
            else:
                amt_debt = 0
            row.append(sum)
            row.append(count)
            row.append(diff)
            row.append(short_percent)
            row.append(amt_debt)
            row_plan += sum
            row_actual += count
            if (count - sum) < 0:
                side_diff += count - sum
        try:
            shortage = (side_diff/row_plan)
        except ZeroDivisionError:   
            shortage = 0
        short_percent = round(shortage,2)
        if side_diff < 0:
            deduction = frappe.db.get_single_value('Deduction Settings', 'cl_deduction')
            amount_debt = side_diff * deduction
        else:
            amount_debt = 0
        row.append(row_plan)
        row.append(row_actual)
        row.append(side_diff)
        row.append(short_percent)
        row.append(amount_debt)
        data.append(row)
    tot = total(args)
    data.append(tot)
    
    return data

def get_dates(args):
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    return dates

def total(args):
    data = ['TOTAL',]
    dates = get_dates(args)
    contractors = frappe.get_all('Contractor',{'status':'Active'})
    overall_plan = 0
    overall_actual = 0
    overall_diff = 0
    overall_shortage = 0
    overall_amt_debt = 0
    for date in dates:
        total_plan = 0
        total_actual = 0
        total_diff = 0
        total_shortage = 0
        total_amt_debt = 0
        for contractor in contractors:
            sum=0
            if frappe.db.exists('CL Head Count Plan',{'contractor':contractor.name,'date':date}):
                plan = frappe.get_doc('CL Head Count Plan',{'contractor':contractor.name,'date':date})
                sum = plan.shift_1 + plan.shift_2+plan.shift_3+plan.shift_pp1+plan.shift_pp2
                
            count = frappe.db.sql("""select count(*) as count from `tabQR Checkin` 
                                    left join `tabEmployee` on `tabQR Checkin`.employee = `tabEmployee`.name
                                     where `tabEmployee`.contractor='%s' and  `tabQR Checkin`.shift_date='%s' and ot = 0
                                    """%(contractor.name,date),as_dict=True)
            if count:
                count = count[0].count
            else:
                count = 0
            total_actual += count
            overall_actual += count
            total_plan += sum
            overall_plan += sum
            diff = count - sum
            if diff < 0:
                total_diff += diff
                overall_diff += diff
                try:
                    total_shortage += (diff /sum)
                    overall_shortage += (diff /sum)
                except ZeroDivisionError:
                    overall_shortage += 0
                deduction = frappe.db.get_single_value('Deduction Settings', 'cl_deduction')
                total_amt_debt += diff * deduction
                overall_amt_debt += diff * deduction
        short_percent = round(total_shortage,2)
        overall_short_percent = round(overall_shortage,2)
        data.append(total_plan)
        data.append(total_actual)
        data.append(total_diff)
        data.append(short_percent)
        data.append(total_amt_debt)
    data.append(overall_plan)
    data.append(overall_actual)
    data.append(overall_diff)
    data.append(overall_short_percent)
    data.append(overall_amt_debt)
    return data