from __future__ import unicode_literals
from os import stat
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, touch_file
from frappe import _
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment,Border,Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import  formatdate
from frappe import _, bold
from calendar import month, monthrange
from datetime import date, timedelta, datetime,time
from frappe.utils import cstr, cint, getdate, get_last_day, get_first_day, add_days
from frappe.utils import cstr, add_days, date_diff, getdate, format_date

@frappe.whitelist()
def download():
    filename = 'CL Salary Register'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    header_date = title (args)
    ws.append(header_date)

    header_date = title1(args)
    ws.append(header_date)

    header = add_header(args)
    ws.append(header)

    header_column = get_column(args)
    ws.append(header_column)
    
    data = get_data(args)
    for row in data:
        ws.append(row)

    align_center = Alignment(horizontal='center',vertical='center')
    for cell in ws["2:2"]:
        cell.alignment = align_center
        cell.font = Font(bold=True)
    for cell in ws['1:1']:
        cell.alignment = align_center
        cell.font = Font(bold=True)
    for cell in ws['3:3']:
        cell.alignment = align_center
        cell.font = Font(bold=True)

    border = Border(left=Side(border_style='thin'),
             right=Side(border_style='thin'),
             top=Side(border_style='thin'),
             bottom=Side(border_style='thin'))

    for rows in ws.iter_rows(min_row=1, max_row=len(get_data(args))+4, min_col=1, max_col=len(title1(args))):
        for cell in rows:
            cell.border = border  

    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column= 6)
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column= len(add_header(args)))
    ws.merge_cells(start_row=1, start_column=len(add_header(args))+1, end_row=3, end_column= len(add_header(args))+8)
    ws.merge_cells(start_row=1, start_column=len(add_header(args))+9, end_row=3, end_column= len(add_header(args))+11)
    ws.merge_cells(start_row=1, start_column=len(add_header(args))+12, end_row=3, end_column= len(add_header(args))+28)
    ws.merge_cells(start_row=1, start_column=len(add_header(args))+29, end_row=3, end_column= len(add_header(args))+36)
    ws.merge_cells(start_row=1, start_column=len(add_header(args))+37, end_row=3, end_column= len(add_header(args))+51)

    for header in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=6):
            for cell in header:
                cell.fill = PatternFill(fgColor='92d14f', fill_type = "solid")
                cell.font = Font(bold=True)
                cell.alignment = align_center
    for header in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=6):
            for cell in header:
                cell.fill = PatternFill(fgColor='feedcc', fill_type = "solid")
                cell.font = Font(bold=True)
                cell.alignment = align_center
    for header in ws.iter_rows(min_row=1, max_row=4, min_col=7, max_col=len(add_header(args))):
            for cell in header:
                cell.fill = PatternFill(fgColor='f8c02b', fill_type = "solid")
                cell.font = Font(bold=True)
                cell.alignment = align_center
    for header in ws.iter_rows(min_row=1, max_row=len(get_data(args))+4, min_col=len(add_header(args))+1, max_col=len(add_header(args))+8):
            for cell in header:
                cell.fill = PatternFill(fgColor='f3af84', fill_type = "solid")
                cell.font = Font(bold=True)
                cell.alignment = align_center
    for header in ws.iter_rows(min_row=1, max_row=len(get_data(args))+4, min_col=len(add_header(args))+9, max_col=len(add_header(args))+11):
            for cell in header:
                cell.fill = PatternFill(fgColor='b3c6e7', fill_type = "solid")
                cell.font = Font(bold=True)
                cell.alignment = align_center
    for header in ws.iter_rows(min_row=4, max_row=4, min_col=len(add_header(args))+12, max_col=len(add_header(args))+28):
            for cell in header:
                cell.fill = PatternFill(fgColor='5ea95e', fill_type = "solid")
                cell.font = Font(bold=True)
    for header in ws.iter_rows(min_row=4, max_row=4, min_col=len(add_header(args))+29, max_col=len(add_header(args))+36):
            for cell in header:
                cell.font = Font(bold=True)
                cell.alignment = align_center
    for header in ws.iter_rows(min_row=4, max_row=4, min_col=len(add_header(args))+37, max_col=len(add_header(args))+51):
            for cell in header:
                cell.fill = PatternFill(fgColor='5ea95e', fill_type = "solid")
                cell.font = Font(bold=True)

    ws.freeze_panes = 'G5'

    ws.sheet_view.zoomScale = 100 

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def title(args):
    month = datetime.strptime(str(args.to_date),'%Y-%m-%d')
    mon = str(month.strftime('%B') +''+ str(month.strftime('%Y')))
    data = ["CL WAGE - "+mon,"",
            "",
            "",
            "",
            "","Attendance"]
    
    return data

@frappe.whitelist()
def title1(args):
    data = ["",
            "",
            "",
            "",
            "",
            "",'Type']
    dates = get_dates(args)
    for date in dates:
        holiday = frappe.db.sql("""select `tabHoliday`.holiday_date,`tabHoliday`.weekly_off from `tabHoliday List` 
        left join `tabHoliday` on `tabHoliday`.parent = `tabHoliday List`.name where  holiday_date = '%s' """%(date),as_dict=True)
        status = ''
        if holiday :
            if holiday[0].weekly_off == 1:
                status = "WW"     
            else:
                status = "HH"
        else:
            status = "W"
        data.extend([status])
    data.extend(["","","","Attendance Summary","","","","","","Overtime Summary","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","","",""])
    return data

@frappe.whitelist()
def add_header(args):
    header = ["","","","","","","Day"]
    dates = get_dates(args)
    for date in dates:
        date = datetime.strptime(date,'%Y-%m-%d')
        day = datetime.date(date).strftime('%a')
        header.extend([day])
    return header

@frappe.whitelist()
def get_column(args):
    data = []
    data += ['S No','Contractor',"ID","Name","Category","DOJ","Department/Date"]
    dates = get_dates(args)
    for date in dates:
        date = datetime.strptime(date,'%Y-%m-%d')
        date = date.strftime('%d')
        data.extend([date])
    data.extend(["11","22","33","PP1","PP2","Total","Fixed Basic","Total Amount","Total OT Hours","OT Amount PH","Total OT Amount"])
    data.extend(["Status","Holiday Working","Working Day","Wage Amount","OT Amount","Shift","Transport","Attendance","Canteen","Deduction","Insurence","PPE","Service Charge","Sourcing Fee","Working Day","Total","%"])
    data.extend(["Basic","HRA","Gross","PPE","ESI","EPF","Service Charge","CTC"])
    data.extend(["Basic","HRA","Gross","PPE","ESI","EPF","Service Charge","Total","OT Amount","Shift","Transport","Attendance","Total","EPF","ESI"])
    return data  

def get_dates(args):
    no_of_days = date_diff(add_days(args['to_date'], 1), args['from_date'])
    dates = [add_days(args['from_date'], i) for i in range(0, no_of_days)]
    return dates

@frappe.whitelist() 
def get_data(args):
    data = []
    row = []
    # basic_component_amount = earning_component_amount = deduction_component_amount = gross_wage = total_deduction = 0

    # earning_comp = ["Basic","House Rent Allowance","Welding Allowance","Attendance Bonus","Additional Allowance","Shift Allowance","Arrear","PP Allowance","Transport Allowance","Others",]

    # dedcution_comp = ["Provident Fund","Employee State Insurance","Canteen Charges","Professional Tax","Labor Welfare Fund","Tel EXP","Personal Protective Equipment","Advance" ]

    if args.employee:
        salary_slips = frappe.get_all("Salary Slip",{'employee_type':'CL','employee':args.employee,'start_date':args.from_date,'end_date':args.to_date},['*'])	

    else:
        salary_slips = frappe.get_all("Salary Slip",{'employee_type':'CL','start_date':args.from_date,'end_date':args.to_date},['*'])	
    i =1
    for ss in salary_slips:
        row = [i,]
        cost_center = frappe.get_value('Department',ss.department,'cost_centre')
        emp = frappe.get_doc("Employee",ss.employee)
        row += [emp.contractor,emp.name,emp.employee_name,emp.designation,formatdate(emp.date_of_joining),emp.department]
        dates = get_dates(args)
        s1 = 0
        s2 = 0
        s3 = 0
        pp1 = 0
        pp2 = 0
        hd_wd = 0
        wd = 0
        twd = 0
        for date in dates:
            att = frappe.db.get_value("Attendance",{'attendance_date':date,'employee':emp.name,'docstatus':('!=','2')},['status','shift','shift_status','qr_shift','attendance_date']) or ''
            if att:
                if att[0] == "Present":
                    if att[1]:
                        row.append(att[1])
                        if att[1] == "1" :
                            s1 += 1
                        if att[1] == "2" :
                            s2 += 1
                        if att[1] == "3" :
                            s3 += 1
                        if att[1] == "PP1" :
                            pp1 += 1
                        if att[1] == "PP2" :
                            pp2 += 1
                    else:
                        if att[3]:
                            row.append(att[3])
                            if att[3] == "1" :
                                s1 += 1
                            if att[3] == "2" :
                                s2 += 1
                            if att[3] == "3" :
                                s3 += 1
                            if att[3] == "PP1" :
                                pp1 += 1
                            if att[3] == "PP2" :
                                pp2 += 1
                    hh = check_holiday(date,emp.name)
                    if hh :
                        if hh == 'WW':
                            hd_wd +=1
                            wd +=1
                        elif hh == 'HH': 
                            hd_wd +=1  
                            wd +=1
                    else:
                        wd +=1
                else:
                    row.append(att[2])
                if att[0]:
                    hh = check_holiday(date,emp.name)
                    if hh :
                        if hh == 'WW':
                            twd +=1
                        elif hh == 'HH': 
                            twd +=1
                    else:
                        twd +=1
            else:
                hh = check_holiday(date,emp.name)
                if hh:
                    if hh == 'WW':
                        row.append(hh)
                    elif hh == 'HH':
                        row.append(hh)
                    elif hh == 'NJ':
                        row.append("-")
                else:
                    row.append("-")
        total_ot = timedelta(0,0,0)
        for date in dates:
            ot = frappe.db.get_value("Overtime Request",{'ot_date':date,'employee':emp.name,'workflow_state':'Approved'},'ot_hours') or ''
            if ot:
                total_ot += ot
                day = ot.days * 24
                hours = day + ot.seconds // 3600
                minutes = (ot.seconds//60)%60
                ftr = [3600,60,1]
                hr = (sum([a*b for a,b in zip(ftr, map(int,str(str(hours) +':'+str(minutes)+':00').split(':')))]))/3600
        day = total_ot.days * 24
        hours = day + total_ot.seconds // 3600
        minutes = (total_ot.seconds//60)%60
        ftr = [3600,60,1]
        total_ot_hr = (sum([a*b for a,b in zip(ftr, map(int,str(str(hours) +':'+str(minutes)+':00').split(':')))]))/3600
        if emp.designation == 'Skilled':
            basic = frappe.db.get_single_value('HR Time Settings','skilled_amount_per_hour')
            ot_amount = basic * total_ot_hr
        elif emp.designation == 'Un Skilled':
            basic = frappe.db.get_single_value('HR Time Settings','unskilled_amount_per_hour')
            ot_amount = basic * total_ot_hr  
        row.extend([
                s1,s2,s3,pp1,pp2,
                (s1+s2+s3+pp1+pp2),
                round(emp.ctc,0),
                round(((s1+s2+s3+pp1+pp2)*round(emp.ctc,0)),0),
                total_ot_hr or 0,
                basic,
                ot_amount,
                "-",
                hd_wd or '-',
                wd or '-',
                round(((s1+s2+s3+pp1+pp2)*round(emp.ctc,0)),0),
                ot_amount,
                int(frappe.get_value('Salary Detail',{'salary_component':"Shift Allowance",'parent':ss.name},["amount"]) or 0),
				int(frappe.get_value('Salary Detail',{'salary_component':"Transport Allowance",'parent':ss.name},["amount"]) or 0),
				int(frappe.get_value('Salary Detail',{'salary_component':"Attendance Bonus",'parent':ss.name},["amount"]) or 0),
				int(frappe.get_value('Salary Detail',{'salary_component':"Canteen Charges",'parent':ss.name},["amount"]) or 0),
				int(frappe.get_value('Salary Detail',{'salary_component':"Other Deduction",'parent':ss.name},["amount"]) or 0),
				int(frappe.get_value('Salary Detail',{'salary_component':"Employee State Insurance",'parent':ss.name},["amount"]) or 0),
				int(frappe.get_value('Salary Detail',{'salary_component':"Personal Protective Equipment",'parent':ss.name},["amount"]) or 0),
				int(frappe.get_value('Salary Detail',{'salary_component':"Service Charges",'parent':ss.name},["amount"]) or 0),
                "",
                (s1+s2+s3+pp1+pp2),
                twd,
                str(float(int(s1+s2+s3+pp1+pp2)/twd)*100) + "%",
                emp.basic or 0,
                emp.house_rent_allowance or 0,
                emp.gross or 0,
                emp.ppe or 0,
                emp.esi_er or 0,
                emp.epf_er or 0,
                emp.service_charge or 0,
                emp.ctc ,
                int(frappe.get_value('Salary Detail',{'salary_component':"Basix",'parent':ss.name},["amount"]) or 0),
                int(frappe.get_value('Salary Detail',{'salary_component':"House Rent Allowance",'parent':ss.name},["amount"]) or 0),
                " ",
                int(frappe.get_value('Salary Detail',{'salary_component':"Personal Protective Equipment",'parent':ss.name},["amount"]) or 0),
                int(frappe.get_value('Salary Detail',{'salary_component':"",'parent':ss.name},["amount"]) or 0),
                int(frappe.get_value('Salary Detail',{'salary_component':"",'parent':ss.name},["amount"]) or 0),
                int(frappe.get_value('Salary Detail',{'salary_component':"",'parent':ss.name},["amount"]) or 0),
                    ])
        data.append(row)   
        i+=1
        
    return data

def check_holiday(date,emp):
    holiday_list = frappe.db.get_value('Employee',{'name':emp},'holiday_list')
    holiday = frappe.db.sql("""select `tabHoliday`.holiday_date,`tabHoliday`.weekly_off from `tabHoliday List` 
    left join `tabHoliday` on `tabHoliday`.parent = `tabHoliday List`.name where `tabHoliday List`.name = '%s' and holiday_date = '%s' """%(holiday_list,date),as_dict=True)
    doj= frappe.db.get_value("Employee",{'name':emp},"date_of_joining")
    status = ''
    if holiday :
        if doj < holiday[0].holiday_date:
            if holiday[0].weekly_off == 1:
                status = "WW"     
            else:
                status = "HH"
        else:
            status = 'NJ'
    return status


              
