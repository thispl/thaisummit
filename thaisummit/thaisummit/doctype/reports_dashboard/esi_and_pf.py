# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def download():
    filename = 'ESI and PF'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
        
    ws = wb.create_sheet(sheet_name, 0)
    ws.append(["S No","Emp No","Cost Center No","DOJ","Employee Name","Designation","Payable Days","ESI","","","","","","EPF","","","","","","","",""])  
    ws.append(["","","","","","","","Standard Gross","Earned Gross","ESI Wages","ESI @ 0,75%","ESI @ 3.25%","Total","PF Wages","EPF Wages","EDLI Wages","EE @ 12%","EPS @ 8.33%","Diff @ 3.67%","ER @ 12%","Total"])  
    salary_slips = frappe.get_all("Salary Slip",{'employee_type':'WC','start_date':args.from_date,'end_date':args.to_date},['*'])
    # cc = frappe.get_all('Department',{'employee_type':'WC'},['cost_centre','name'])
    i = 1
    total_standard_gross = 0
    total_earned_gross = 0
    total_esi_wages = 0
    esi_a = 0
    esi_b = 0
    e_total = 0
    total_pf_wages = 0
    total_eps_wages = 0
    total_edli_wages = 0
    total_ee = 0
    total_eps = 0
    total_diff = 0
    total_er = 0
    p_total = 0

    for ss in salary_slips:
        # cst_ctr = frappe.db.get_value('Employee',ss.employee,'cost_center')
        pf_wages = frappe.db.get_value('Employee',ss.employee,'basic')
        ee = frappe.db.get_value('Employee',ss.employee,'epf_er')
        eps = frappe.db.get_value('Employee',ss.employee,'medical_allowance')
        diff = ee-eps
        total_2 = ee + eps + diff
        standard_gross = frappe.db.get_value('Employee',ss.employee,'gross')
        cost = frappe.db.get_value('Department',ss.department,'cost_centre')
       
        if(pf_wages > 15000):
            eps_wages = 15000
            edli_wages = 15000
        if(pf_wages < 15000):
            eps_wages = pf_wages
            edli_wages = pf_wages
        if(standard_gross < 21000):
            esi_wages = ss.gross_pay
            esi_1 = frappe.get_value('Salary Detail',{'salary_component':'Employee State insurance','parent':ss.name },['amount']) or 0
            esi_2 = esi_wages * 0.0325
            total = esi_1 + esi_2
        else:
            esi_wages = 0
            esi_1 = 0
            esi_2 = 0
            total = 0
        ws.append([i,ss.employee,cost,ss.date_of_join,ss.employee_name,ss.designation,ss.payment_days,standard_gross,ss.gross_pay,esi_wages,esi_1,esi_2,total,pf_wages,eps_wages,edli_wages,ee,eps,diff,ee,total_2])
        i=i+1
        total_standard_gross += standard_gross
        total_earned_gross += ss.gross_pay
        total_esi_wages += esi_wages
        esi_a += esi_1
        esi_b += esi_2
        e_total += total
        total_pf_wages += pf_wages
        total_eps_wages += eps_wages
        total_edli_wages += edli_wages
        total_ee += ee
        total_eps += eps
        total_diff += diff
        total_er += ee
        p_total += total_2
    ws.append(['Total','','','','','','',total_standard_gross,total_earned_gross,total_esi_wages,esi_a,esi_b,e_total,total_pf_wages,total_eps_wages,total_edli_wages,total_ee,total_eps,total_diff,total_er,p_total ])
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=13) 
    ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=21) 
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.merge_cells(start_row=len(salary_slips)+3, start_column=1, end_row=len(salary_slips)+3, end_column=7) 

    for header in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=13):
         for cell in header:
             cell.fill = PatternFill(fgColor='AFE1AF', fill_type = "solid")

    for header in ws.iter_rows(min_row=1, max_row=len(salary_slips)+3, min_col=14, max_col=21):
         for cell in header:
              cell.fill = PatternFill(fgColor='ADD8E6', fill_type = "solid")
    
    for header in ws.iter_rows(min_row=3, max_row=len(salary_slips)+3, min_col=8, max_col=13):
         for cell in header:
              cell.fill = PatternFill(fgColor='AFE1AF', fill_type = "solid")

    for header in ws.iter_rows(min_row=len(salary_slips)+3, max_row=len(salary_slips)+3, min_col=1, max_col=7):
         for cell in header:
              cell.fill = PatternFill(fgColor='fefe33', fill_type = "solid")

    for header in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=27):
         for cell in header:
             cell.font = Font(bold=True)

    for header in ws.iter_rows(min_row=len(salary_slips)+3, max_row=len(salary_slips)+3, min_col=1, max_col=27):
         for cell in header:
             cell.font = Font(bold=True)


    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws["2:2"]:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws["B:B"]:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws["C:C"]:
        cell.alignment = Alignment(horizontal='center')
    for cell in ws["D:D"]:
        cell.alignment = Alignment(horizontal='center')
    for rows in ws.iter_rows(min_row=len(salary_slips)+3, max_row=len(salary_slips)+3, min_col=1, max_col=7):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
        
    

    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))
        
    # font = Font(font_style='bold', color='000000'),

    for rows in ws.iter_rows(min_row=1, max_row=len(salary_slips)+3, min_col=1, max_col=21):
        for cell in rows:
            cell.border = border

    
    # for rows in ws.iter_rows(min_row=len(salary_slips)+3, max_row=len(salary_slips)+3, min_col=1, max_col=21):
    #     for cell in rows:
    #         cell.font = font
             
   
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'