# Copyright (c) 2022, TEAMPRO and contributors
# For license information, please see license.txt
from __future__ import unicode_literals
from email import message
from frappe.utils.background_jobs import enqueue
from frappe.utils import now_datetime, formatdate
from fileinput import filename
from os import stat
from sqlite3 import Row

from redis import from_url
import frappe
from frappe.utils import cstr, add_days, date_diff, format_datetime
from frappe import _
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime, time
import pandas as pd
import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


class DeliveryPlanvsActual(Document):
    pass

@frappe.whitelist()
def get_data(from_date,to_date):
    data = """<table class='table table-bordered=1'>
    <tr><td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'></td><td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'></td><td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'></td>"""
    no_of_days = date_diff(add_days(to_date, 1), from_date)
    dates = [add_days(from_date, i) for i in range(0, no_of_days)]
    for d in dates:
        d = datetime.strptime(d,'%Y-%m-%d')
        date = d.strftime('%d-%b')
        data += "<td colspan=4 style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'><center>%s</center</td>"%(date)
    data += "<td colspan=4 style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'><center>TOTAL</center></td>"
    data += "</tr><tr><td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'><center>MAT NO</center></td><td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'><center>PART NO</center</td><td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'><center>PART NAME</center></td>"
    for d in dates:
        data += "<td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'>PLAN</td><td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'>ACTUAL</td><td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'>DIFF</td><td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'>%</td>"
    data += "<td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'>PLAN</td><td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'>ACTUAL</td><td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'>DIFF</td><td style='background-color:#ffedcc; padding:2px; border: 1px solid black; font-size:15px;'>%</td></tr>"
    pos = frappe.db.sql("select DISTINCT mat_no from `tabTSAI PO` where plan_date between '%s' and '%s' " % (from_date, to_date), as_dict=True)
    for po in pos:
        total_plan = 0
        total_actual = 0
        total_diff = 0
        total_percent = 0
        part = frappe.db.get_value('TSAI PO',{'mat_no':po.mat_no},['parts_no','parts_name'])
        # row = [po.mat_no, part[0], part[1]]
        data += "<tr><td style='padding:1px; border: 1px solid black; font-size:10px;'>%s</td><td style='padding:1px; border: 1px solid black; font-size:10px;'>%s</td><td style='padding:1px; border: 1px solid black; font-size:10px;'>%s</td>"%(po.mat_no, part[0], part[1])
        for date in dates:
            plan = frappe.db.sql('select sum(po_qty) as qty from `tabTSAI PO` where mat_no = "%s" and plan_date = "%s" and po_status = "O" '%(po.mat_no,date),as_dict=True)[0].qty or 0
            actual = frappe.db.sql(
                """select sum(`tabInvoice Items`.key_qty) as qty from `tabTSAI Invoice` 
                left join `tabInvoice Items` on `tabTSAI Invoice`.name = `tabInvoice Items`.parent where `tabTSAI Invoice`.invoice_date = '%s' and `tabTSAI Invoice`.status = 'OPEN' and `tabInvoice Items`.mat_no = '%s' """ % (date,po.mat_no), as_dict=True)[0].qty or 0
            diff = plan - actual
            percent = 0
            if actual > 0:
                if plan > 0:
                    percent = round((actual/plan)*100,2)
            total_plan += plan
            total_actual += actual
            total_diff += diff
            total_percent += percent
            if plan == 0:
                plan = '-'
            if actual == 0:
                actual = '-'
            if diff == 0:
                diff = '-'
            if percent == 0:
                percent = '-'
            # row.extend([plan, actual, diff, percent])
            data += "<td style='padding:1px; border: 1px solid black; font-size:10px;'><center>%s</center></td><td style='padding:1px; border: 1px solid black; font-size:10px;'><center>%s</center></td><td style='padding:1px; border: 1px solid black; font-size:10px;'><center>%s</center></td><td style='padding:1px; border: 1px solid black; font-size:10px;'><center>%s</center></td>"%(plan, actual, diff, percent)
        if total_plan > 0:
            if total_plan == 0:
                total_plan = '-'
            if total_actual == 0:
                total_actual = '-'
            if total_diff == 0:
                total_diff = ''
            if total_percent == 0:
                total_percent = ''
        data += "<td style='padding:1px; border: 1px solid black; font-size:10px;'><center>%s</center></td><td style='padding:1px; border: 1px solid black; font-size:10px;'><center>%s</center></td><td style='padding:1px; border: 1px solid black; font-size:10px;'><center>%s</center></td><td style='padding:1px; border: 1px solid black; font-size:10px;'><center>%s</center></td>"%(total_plan, total_actual, total_diff, total_percent)
        # row.extend([total_plan, total_actual, total_diff, total_percent])
        data += "</tr>"
    return data


@frappe.whitelist()
def enqueue_download(from_date,to_date):
    if not frappe.db.exists("Enqueue Methods",{'method':'Delivery Plan vs Actual','status':'Queued'}):
        doc = frappe.new_doc("Enqueue Methods")
        doc.method = "Delivery Plan vs Actual"
        doc.status = "Queued"
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        args = {'from_date':from_date,'to_date':to_date,}
        enqueue(download_delivery_plan_vs_actual, queue='default', timeout=6000, event='build_xlsx_response',args=args,enqueue_id=doc.name)
        frappe.msgprint("Delivery Plan vs Actual Download is successsfully Initiated. Kindly wait for sometime and refresh the page.")
    else:
        frappe.msgprint("Delivery Plan vs Actual Download is already in Progress. Please wait for sometime and refresh the page. ")

@frappe.whitelist()
def download_delivery_plan_vs_actual(args,enqueue_id):
    build_xlsx_response(args,enqueue_id)


def make_xlsx(args, sheet_name=None, wb=None, column_widths=None):
    # args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    dates = get_dates(args)
    header = ["", "", ""]
    for d in dates:
        d = datetime.strptime(d,'%Y-%m-%d')
        date = d.strftime('%d-%b')
        header.extend([date, "", "", ""])
    header.extend(["TOTAL", "", "", ""])

    ws.append(header)

    header_title = ["MAT NO", "PART NO", "PART NAME"]
    for d in dates:
        header_title.extend(["PLAN", "ACTUAL", "DIFF", "%"])
    header_title.extend(["PLAN", "ACTUAL", "DIFF", "%"])

    ws.append(header_title)
    i = 4
    for d in dates:
        ws.merge_cells(start_row=1, start_column=i, end_row=1, end_column=i+3)
        i += 4
    align_center = Alignment(horizontal='center',vertical='center')
    for cell in ws["1:1"]:
        cell.alignment = align_center
    for cell in ws["2:2"]:
        cell.alignment = align_center
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    for cell in ws["2:2"]:
        cell.font = Font(bold=True)
    
    pos = frappe.db.sql("select DISTINCT mat_no from `tabTSAI PO` where plan_date between '%s' and '%s' " % (
        args['from_date'], args['to_date']), as_dict=True)
    for po in pos:
        total_plan = 0
        total_actual = 0
        total_diff = 0
        total_percent = 0
        part = frappe.db.get_value('TSAI PO',{'mat_no':po.mat_no},['parts_no','parts_name'])
        row = [po.mat_no, part[0], part[1]]
        for date in dates:
            plan = frappe.db.sql('select sum(po_qty) as qty from `tabTSAI PO` where mat_no = "%s" and plan_date = "%s" and po_status = "O" '%(po.mat_no,date),as_dict=True)[0].qty or 0
            actual = frappe.db.sql(
                """select sum(`tabInvoice Items`.key_qty) as qty from `tabTSAI Invoice` 
                left join `tabInvoice Items` on `tabTSAI Invoice`.name = `tabInvoice Items`.parent where `tabTSAI Invoice`.invoice_date = '%s' and `tabTSAI Invoice`.status = 'OPEN' and `tabInvoice Items`.mat_no = '%s' """ % (date,po.mat_no), as_dict=True)[0].qty or 0
            diff = plan - actual
            percent = 0
            if actual > 0:
                if plan > 0:
                    percent = round((actual/plan)*100,2)
            total_plan += plan
            total_actual += actual
            total_diff += diff
            total_percent += percent
            if plan == 0:
                plan = '-'
            if actual == 0:
                actual = '-'
            if diff == 0:
                diff = '-'
            if percent == 0:
                percent = '-'
            row.extend([plan, actual, diff, percent])
        if total_plan > 0:
            if total_plan == 0:
                total_plan = '-'
            if total_actual == 0:
                total_actual = '-'
            if total_diff == 0:
                total_diff = ''
            if total_percent == 0:
                total_percent = ''
        row.extend([total_plan, total_actual, total_diff, total_percent])
        ws.append(row)

    total=["TOTAL","",""]
    total_plan = 0
    total_actual = 0
    total_diff = 0
    total_percent = 0
    for date in dates:
        plan = frappe.db.sql('select sum(po_qty) as qty from `tabTSAI PO` where plan_date = "%s" and po_status = "O" '%(date),as_dict=True)[0].qty or 0
        actual = frappe.db.sql(
            """select sum(`tabInvoice Items`.key_qty) as qty from `tabTSAI Invoice` 
            left join `tabInvoice Items` on `tabTSAI Invoice`.name = `tabInvoice Items`.parent where `tabTSAI Invoice`.invoice_date = '%s' and `tabTSAI Invoice`.status = 'OPEN' """ % (date), as_dict=True)[0].qty or 0
        diff = plan - actual
        percent = 0
        if actual > 0:
            if plan > 0:
                percent = round((actual/plan)*100,2)
        total_plan += plan
        total_actual += actual
        total_diff += diff
        total_percent += percent
        total.extend([plan,actual,diff,percent])
    if total_plan > 0:
        if total_plan == 0:
            total_plan = '-'
        if total_actual == 0:
            total_actual = '-'
        if total_diff == 0:
            total_diff = ''
        if total_percent == 0:
            total_percent = ''
        total.extend([total_plan, total_actual, total_diff, total_percent])            
        ws.append(total)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file


def build_xlsx_response(args,enqueue_id):
    xlsx_file = make_xlsx(args)
    # write out response as a xlsx type
    # frappe.response['filename'] = filename + '.xlsx'
    # frappe.response['filecontent'] = xlsx_file.getvalue()
    # frappe.response['type'] = 'binary'
    ret = frappe.get_doc({
            "doctype": "File",
            "attached_to_name": '',
            "attached_to_doctype": 'Delivery Plan vs Actual',
            "attached_to_field": 'attach',
            "file_name": 'Delivery_Plan_vs_Actual ' + str(formatdate(args['from_date'])) + ' - ' + str(formatdate(args['to_date']))   +'.xlsx',
            "is_private": 0,
            "content": xlsx_file.getvalue(),
            "decode": False
        })
    ret.save(ignore_permissions=True)
    frappe.db.commit()
    attached_file = frappe.get_doc("File", ret.name)
    frappe.db.set_value('Delivery Plan vs Actual',None,'attach',attached_file.file_url)
    frappe.db.set_value('Delivery Plan vs Actual',None,'last_download_on',now_datetime())
    frappe.db.set_value('Enqueue Methods',enqueue_id,'status','Completed')



def get_dates(args):
    no_of_days = date_diff(add_days(args['to_date'], 1), args['from_date'])
    dates = [add_days(args['from_date'], i) for i in range(0, no_of_days)]
    return dates


def test_method():
    pos = frappe.db.sql(
        "select DISTINCT mat_no from `tabTSAI PO` ", as_dict=True)
    for po in pos:
        if po.mat_no == '31130003':
            print(po.mat_no)
        # plan = frappe.get_value(
        #     "TSAI PO", {'mat_no': po.mat_no, 'po_status': 'O'}, ['sum(po_qty)'])
        # print(plan)
    # tsa_po = frappe.get_all("TSAI PO",{"po_date": ("between", ('2022-03-01','2022-03-08'))},['mat_no','parts_no'])

    # cols = ['mat_no','part_no','part_name','PO Date','po QTY']
    # data = []
    # for po in tsa_po:
    #     data.append([po.mat_no,po.parts_no,po.parts_name,po.po_date,po.po_qty])
    # df = pd.DataFrame(data=data,columns=cols)
    # df2 = df.groupby('mat_no')['po_qty'].Sum()
    # print(df2)

    # no_of_days = date_diff(add_days('2022-03-08', 1), '2022-03-01')
    # dates = [add_days('2022-03-01', i) for i in range(0, no_of_days)]
    # for d in df2.to_dict()['part_no']:
    #     for date in dates:
    #         qty = frappe.db.get_value("TSAI PO",{'mat_no':d,'po_date':date},'po_qty') or 0
    #         print(d)
    #         print(df)
    #         print(df2)
    #         print('----')
