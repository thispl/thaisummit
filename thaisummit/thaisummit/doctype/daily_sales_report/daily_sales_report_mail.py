# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date,get_first_day,get_last_day,today
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def send_mail():
    args = {'from_date':get_first_day(add_days(today(),-1)),'to_date':get_last_day(add_days(today(),-1))}
    filename = 'Daily Sales Report'
    # test = build_xlsx_response(filename)
    enqueue(build_xlsx_response, queue='default', timeout=6000, event='build_xlsx_response',filename=filename,args=args)


# return xlsx file object
def make_xlsx(data, args, sheet_name=None, wb=None, column_widths=None):
    # args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header = ["DAILY SALES REPORT"]
    ws.append(header)

    group_header,customer_header = add_header()
    ws.append(group_header)
    ws.append(customer_header)

    data = get_data(args)

    for row in data:
        ws.append(row)

    groups = ('IYM','RE','FORD')
    i = 4
    for g in groups:
        customer = frappe.db.sql("""select short_name from `tabTSAI Customer` where running is not null and customer_group = '%s' order by running """%(g),as_dict=True)
        i = i +len(customer)
        ws.merge_cells(start_row=2, start_column=i, end_row=3, end_column=i)
        i += 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=i)
    ws.merge_cells(start_row=2, start_column=i, end_row=3, end_column=i)

    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)

    align_center = Alignment(horizontal='center',vertical='center')
    for cell in ws["2:2"]:
        cell.alignment = align_center
    for cell in ws["3:3"]:
        cell.alignment = align_center
    for cell in ws["A:A"]:
        cell.alignment = align_center
    for cell in ws["B:B"]:
        cell.alignment = align_center

    ws['A1'].alignment = Alignment(vertical='center',horizontal='center')

    ws.freeze_panes = 'C4'
    ws.sheet_view.zoomScale = 70

    bold_font = Font(bold=True,size=20)
    for cell in ws["1:1"]:
        cell.font = bold_font
    for cell in ws["2:2"]:
        cell.font = Font(bold=True)
    for cell in ws["3:3"]:
        cell.font = Font(bold=True)
    

    
    ws['A1'].fill = PatternFill(fgColor="dc7633", fill_type = "solid")

    dates = get_dates(args)
    for rows in ws.iter_rows(min_row=2, max_row=len(dates)+3, min_col=1, max_col=2):
        for cell in rows:
            cell.fill = PatternFill(fgColor="d6dbdf", fill_type = "solid")
    
    ws.merge_cells(start_row=len(dates)+4, start_column=1, end_row=len(dates)+4, end_column=2)

    iym = frappe.db.sql("""select short_name from `tabTSAI Customer` where running is not null and customer_group = 'IYM' """,as_dict=True)
    re = frappe.db.sql("""select short_name from `tabTSAI Customer` where running is not null and customer_group = 'RE' """,as_dict=True)
    ford = frappe.db.sql("""select short_name from `tabTSAI Customer` where running is not null and customer_group = 'FORD' """,as_dict=True)

    for rows in ws.iter_rows(min_row=2, max_row=3, min_col=3, max_col=len(iym)+4):
        for cell in rows:
            cell.fill = PatternFill(fgColor='3498db', fill_type = "solid")
    
    for rows in ws.iter_rows(min_row=2, max_row=3, min_col=len(iym)+5, max_col=len(iym)+len(re)+5):
        for cell in rows:
            cell.fill = PatternFill(fgColor='7dcea0', fill_type = "solid")
        
    for rows in ws.iter_rows(min_row=2, max_row=3, min_col=len(iym)+len(re)+6, max_col=len(iym)+len(re)+len(ford)+6):
        for cell in rows:
            cell.fill = PatternFill(fgColor='edbb99', fill_type = "solid")

    for rows in ws.iter_rows(min_row=2, max_row=3, min_col=len(iym)+len(re)+len(ford)+7, max_col=len(iym)+len(re)+len(ford)+7):
        for cell in rows:
            cell.fill = PatternFill(fgColor='d6dbdf', fill_type = "solid")
        
    for rows in ws.iter_rows(min_row=2, max_row=len(dates)+3, min_col=len(iym)+4, max_col=len(iym)+4):
        for cell in rows:
            cell.fill = PatternFill(fgColor='3498db', fill_type = "solid")

    for rows in ws.iter_rows(min_row=2, max_row=len(dates)+3, min_col=len(iym)+len(re)+5, max_col=len(iym)+len(re)+5):
        for cell in rows:
            cell.fill = PatternFill(fgColor='7dcea0', fill_type = "solid")

    for rows in ws.iter_rows(min_row=2, max_row=len(dates)+3, min_col=len(iym)+len(re)+len(ford)+6, max_col=len(iym)+len(re)+len(ford)+6):
        for cell in rows:
            cell.fill = PatternFill(fgColor='edbb99', fill_type = "solid")

    for rows in ws.iter_rows(min_row=2, max_row=len(dates)+3, min_col=len(iym)+len(re)+len(ford)+7, max_col=len(iym)+len(re)+len(ford)+7):
        for cell in rows:
            cell.fill = PatternFill(fgColor='d6dbdf', fill_type = "solid")

    for rows in ws.iter_rows(min_row=len(dates)+4, max_row=len(dates)+4, min_col=1, max_col=len(iym)+len(re)+len(ford)+7):
        for cell in rows:
            cell.fill = PatternFill(fgColor='faf423', fill_type = "solid")
        
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    for rows in ws.iter_rows(min_row=1, max_row=len(dates)+4, min_col=1, max_col=len(iym)+len(re)+len(ford)+7):
        for cell in rows:
            cell.border = border
     

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

# def set_border(ws, cell_range):
#     border = Border(left=Side(border_style='thin', color='000000'),
#                 right=Side(border_style='thin', color='000000'),
#                 top=Side(border_style='thin', color='000000'),
#                 bottom=Side(border_style='thin', color='000000'))

    # rows = ws.iter_rows(cell_range)
    # for row in rows:
    #     for cell in row:
    #         cell.border = border 

def build_xlsx_response(filename,args):
    xlsx_file = make_xlsx(filename,args)
    ret = frappe.get_doc({
            "doctype": "File",
            "attached_to_name": '',
            "attached_to_doctype": 'Daily Sales Report',
            "attached_to_field": 'attach',
            "file_name": filename + '.xlsx',
            "is_private": 0,
            "content": xlsx_file.getvalue(),
            "decode": False
        })
    ret.save(ignore_permissions=True)
    # frappe.log_error(title='res',message=ret)
    frappe.db.commit()
    attached_file = frappe.get_doc("File", ret.name)
    frappe.db.set_value('Daily Sales Report',None,'attach',attached_file.file_url)
    attachments = [{
				'fname':  ret.file_name,
				'fcontent': xlsx_file.getvalue()
			}]
    frappe.sendmail(
            recipients= ['subash.p@groupteampro.com','Dittavarong.Gat@thaisummit.co.th','chaiyasit.sas@thaisummit.co.in','ravi.nar@thaisummit.co.in','sugu.aru@thaisummit.co.in','saravanan.sub@thaisummit.co.in','gopi.sek@thaisummit.co.in','vasanth.sol@thaisummit.co.in','rajaguru.sad@thaisummit.co.in','asha.ntr@thaisummit.co.in','abhishek.sha@thaisummit.co.th','atichat.kum@thaisummit.co.th','uthaiporn.pun@thaisummit.co.th','rajaram.ram@thaisummit.co.in',' kumaresan.sel@thaisummit.co.in','deepanch.jay@thaisummit.co.in'],
            # recipients= ['subash.p@groupteampro.com'],
            subject="Daily Sales Report",
            attachments= attachments,
            message= """Dear Team, <br><br> Kindly find the attached Daily Sales Report.""")

def get_dates(args):
    no_of_days = date_diff(add_days(args["to_date"], 1), args["from_date"])
    dates = [add_days(args["from_date"], i) for i in range(0, no_of_days)]
    return dates

@frappe.whitelist()
def add_header():
    group_header = ['Date','Day','IYM']
    customer_header = ['','','HPS']
    groups = ['IYM','RE','FORD']
    for g in groups:
        customer = frappe.db.sql("""select short_name from `tabTSAI Customer` where running is not null and customer_group = '%s' order by running """%(g),as_dict=True)
        for c in customer:
            group_header.append(g)
            customer_header.append(c.short_name)
        group_header.append('TOTAL \n' + str(g))
        customer_header.append('')
    group_header.append('TOTAL')
    return group_header, customer_header

def get_data(args):
    data = []
    groups = ['IYM','RE','FORD']
    hps_parts = ('10000289','10000290','10000291','10000335','10000336','10000337','10000338')
    dates = get_dates(args)
    hps_total = 0
    for date in dates:
        row = []
        row_total = 0
        # dt = datetime.strptime(date,'%Y-%m-%d')
        d = date.strftime('%d-%b-%Y')
        day = date.strftime('%a')
        hps_out = frappe.db.sql(""" select sum(sales_value) as sales_value from `tabSAP Outgoing Report` where ar_invoice_date = '%s' and part_no in %s and customer_ref_no not like '_________/O/%%' """%(date,hps_parts),as_dict=True)
        if hps_out[0].sales_value is not None:
            hps_total += hps_out[0].sales_value
            hps_out = hps_out[0].sales_value
            row.extend([d,day,'{:,}'.format(round(hps_out,2))])
        else:
            hps_out = 0
            row.extend([d,day,'-'])
        for g in groups:
            customer = frappe.db.sql("""select short_name from `tabTSAI Customer` where running is not null and customer_group = '%s' order by running """%(g),as_dict=True)
            sap_total = 0
            for c in customer:
                sap_out = frappe.db.sql(""" select sum(sales_value) as sales_value from `tabSAP Outgoing Report` where ar_invoice_date = '%s' and short_name = '%s' and customer_ref_no not like '_________/O/%%' """%(date,c.short_name),as_dict=True)
                if sap_out[0].sales_value is not None:
                    row.append('{:,}'.format(round(sap_out[0].sales_value,2)))
                    sap_total += sap_out[0].sales_value + hps_out
                    row_total += sap_out[0].sales_value + hps_out
                else:
                    row.append('-')
                hps_out = 0
            if sap_total == 0:
                row.append('-')
            else:
                row.append('{:,}'.format(round(sap_total,2)))
        if row_total == 0:
            row.append('-')
        else:
            row.append('{:,}'.format(round(row_total,2)))

        data.append(row)
    total = ['TOTAL','','{:,}'.format(round(hps_total,2))]
    overall_total = 0
    for g in groups:
        customer = frappe.db.sql("""select short_name from `tabTSAI Customer` where running is not null and customer_group = '%s' order by running """%(g),as_dict=True)
        sap_total = 0
        for c in customer:
            no_of_days = date_diff(add_days(args["to_date"], 1), args["from_date"])
            date_list = [str(add_days(args["from_date"], i)) for i in range(0, no_of_days)]
            d_list = ''
            for date in date_list:
                d_list += '"' + str(date) + '",' 
            # date_list = str(dates).replace('[','(')
            # date_list = str(date_list).replace(']',')')
            # frappe.log_error(title='d',message=d_list)            
            sap_out = frappe.db.sql(""" select sum(sales_value) as sales_value from `tabSAP Outgoing Report` where ar_invoice_date in (%s) and short_name = '%s' and customer_ref_no not like '_________/O/%%' """%(d_list[:-1],c.short_name),as_dict=True)
            if sap_out[0].sales_value is not None:
                total.append('{:,}'.format(round(sap_out[0].sales_value,2)))
                sap_total += sap_out[0].sales_value + hps_total
                overall_total += sap_out[0].sales_value + hps_total
            else:
                total.append('-')
            hps_total = 0
        if sap_total == 0:
            total.append('-')
        else:
            total.append('{:,}'.format(round(sap_total,2)))
    if overall_total == 0:
        total.append('-')
    else:
        total.append('{:,}'.format(round(overall_total,2)))
    data.append(total)

    return data
