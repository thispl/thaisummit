# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
from email import header
from inspect import ArgSpec
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from math import floor

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def download(f_date,t_date):
    frappe.errprint('CL')
    filename = 'CL Monthly OT Register'
    args = {'from_date':f_date,'to_date':t_date}
    enqueue(build_xlsx_response, queue='default', timeout=6000, event='build_xlsx_response',filename=filename,args=args)
    # test = build_xlsx_response(filename)

def make_xlsx(data,args ,sheet_name=None, wb=None, column_widths=None):
    # args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    dates = get_dates(args)
    ws.append(['CL Monthly OT Register ','','','','','','','','','','','','','','','','','','','',''])
    header = ['Employee ID','Employee Name','Department','DOJ']
    for date in dates:
        header.append(date)
    header.extend(['Total OT','OT Basic','OT Amount'])
    ws.append(header)
    
    employees = frappe.get_all('Employee',{'employee_type':'CL','status':'Active','vacant':0},['name','employee_name','department','date_of_joining','designation'])
    left_employees = frappe.db.sql("""select name, basic, employee_name, department, date_of_joining,designation from `tabEmployee` where status = 'Left' and employee_type = 'CL' and relieving_date >= %s and vacant = 0 """%(args['from_date']),as_dict=True)
    employees.extend(left_employees)
    for emp in employees:
            total_ot_hr = 0
            dates = get_dates(args)
            if frappe.db.exists("Overtime Request",{'ot_date':('in',dates),'employee':emp.name,'workflow_state':'Approved'}):
                row = [emp.name,emp.employee_name,emp.department,emp.date_of_joining]
                for date in dates:
                    ot = frappe.db.get_value("Overtime Request",{'ot_date':date,'employee':emp.name,'workflow_state':'Approved'},'ot_hours') or ''
                    if ot:
                        day = ot.days * 24
                        hours = day + ot.seconds // 3600
                        minutes = (ot.seconds//60)%60
                        ftr = [3600,60,1]
                        hr = (sum([a*b for a,b in zip(ftr, map(int,str(str(hours) +':'+str(minutes)+':00').split(':')))]))/3600
                        row.append(hr)
                        total_ot_hr += hr
                    else:
                        row.append('-')

                row.append(total_ot_hr)
                ot_amount = frappe.db.sql("select sum(ot_amount) as total from `tabOvertime Request` where ot_date between '%s' and '%s' and workflow_state = 'Approved' and employee = '%s' "%(args['from_date'],args['to_date'],emp.name),as_dict=True)[0].total or 0
                if emp.designation == 'Skilled':
                    row.append(frappe.db.get_single_value('HR Time Settings','skilled_amount_per_hour'))
                elif emp.designation == 'Un Skilled':
                    row.append(frappe.db.get_single_value('HR Time Settings','unskilled_amount_per_hour'))
                row.append(floor(ot_amount))
                ws.append(row)
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
      
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

# def build_xlsx_response(filename):
#     xlsx_file = make_xlsx(filename)
#     frappe.response['filename'] = filename + '.xlsx'
#     frappe.response['filecontent'] = xlsx_file.getvalue()
#     frappe.response['type'] = 'binary'



def build_xlsx_response(filename,args):
    xlsx_file = make_xlsx(filename,args)
    ret = frappe.get_doc({
            "doctype": "File",
            "attached_to_name": '',
            "attached_to_doctype": 'CL Reports',
            "attached_to_field": 'attach',
            "file_name": filename + '.xlsx',
            "is_private": 0,
            "content": xlsx_file.getvalue(),
            "decode": False
        })
    ret.save(ignore_permissions=True)
    frappe.db.commit()
    # frappe.log_error(message=ret)
    attached_file = frappe.get_doc("File", ret.name)
    frappe.db.set_value('CL Reports',None,'attach',attached_file.file_url)


def get_dates(args):
    no_of_days = date_diff(add_days(args['to_date'], 1), args['from_date'])
    dates = [add_days(args['from_date'], i) for i in range(0, no_of_days)]
    return dates