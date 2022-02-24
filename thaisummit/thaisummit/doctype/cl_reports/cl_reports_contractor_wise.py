# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def download():
    filename = 'CL Wages Contractor Wise'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    contractor = frappe.get_value('Contractor',{'name':args.contractor},['contractors_company'])
    ws.append([str(contractor) + ' Manpower Summary ' + args.from_date + ' to '+ args.to_date,'','','','','','','','','','','','','','','','','','','',''])
    ws.append(['','','','','','Income','','','','','','','','','','','','Deduction'])
    ws.append(['S No','Party','List','Head count','No of Person','Wage Amount','OT Amount','Wage + OT','Shift','Transport','Attendance','SV Wages','Insurance','PPE','Service Charge','Sourcing Fee','Total','Others','Canteen','Grand Total','% OT'])
    department = frappe.db.get_all('Department',{'is_group':0},['*'])
    i = 1
    total_head_count = 0
    total_person_count = 0
    total_wages = 0
    total_others = 0
    total_shift = 0
    total_transport = 0
    total_attendance = 0
    total_ppe = 0
    total_deduction = 0
    total_canteen = 0
    total_wages_ot = 0
    total_grand_total = 0
    sum_of_total = 0
    total_ot = 0
    for dept in department:
        person_count = frappe.db.count('Employee',{'department':dept.name,'employee_type':'CL'}) or 0
        head_count = frappe.db.sql(""" select sum(payment_days) as payment_days from `tabSalary Slip` where department = '%s' and start_date = '%s' and end_date = '%s' and employee_type = 'CL' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].payment_days or 0
        basic = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as basic from `tabSalary Slip`
        left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Basic' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' and employee_type = 'CL' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].basic or 0
        others = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as others from `tabSalary Slip`
        left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Others' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' and employee_type = 'CL' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].others or 0
        shift = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as shift from `tabSalary Slip`
        left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Shift Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' and employee_type = 'CL' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].shift or 0
        transport = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as transport from `tabSalary Slip`
        left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Transport Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' and employee_type = 'CL' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].transport or 0
        attendance = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as attendance from `tabSalary Slip`
        left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Attendance Allowance' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' and employee_type = 'CL' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].attendance or 0
        ppe = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as ppe from `tabSalary Slip`
        left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'PPE Cost' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' and employee_type = 'CL' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].ppe or 0
        deduction = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as deduction from `tabSalary Slip`
        left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Other Deduction' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' and employee_type = 'CL' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].deduction or 0
        canteen = frappe.db.sql("""select sum(`tabSalary Detail`.amount) as canteen from `tabSalary Slip`
        left join `tabSalary Detail` on `tabSalary Slip`.name = `tabSalary Detail`.parent where `tabSalary Slip`.department = '%s' and `tabSalary Detail`.salary_component = 'Canteen Charges' and `tabSalary Slip`.start_date ='%s' and `tabSalary Slip`.end_date = '%s' and employee_type = 'CL' """%(dept.name,args.from_date,args.to_date),as_dict=True)[0].canteen or 0
        total = basic + others + shift + transport + attendance + ppe
        wages_ot = basic + others or 0
        deductions = deduction + canteen
        grand_total = total - deduction
        ot = 0
        if not basic == 0: 
            ot = others/basic*100
        ws.append([i,args.contractor,dept.name,head_count,person_count,basic,others,wages_ot,shift,transport,attendance,'','',ppe,'','',total,deduction,canteen,grand_total,ot])
        i = i + 1
        total_head_count = total_head_count + head_count
        total_person_count = total_person_count + person_count
        total_wages = total_wages + basic
        total_others = total_others + others
        total_shift = total_shift + shift
        total_transport = total_transport + transport
        total_attendance =  total_attendance + attendance
        total_ppe = total_ppe + ppe
        total_deduction = total_deduction + deduction
        total_canteen = total_canteen + canteen
        total_wages_ot = total_wages_ot + wages_ot
        total_grand_total = total_grand_total + grand_total
        sum_of_total =  sum_of_total + total
        total_ot = total_ot + ot
    ws.append(['','','Total',total_head_count,total_person_count, total_wages,total_others,total_wages_ot,total_shift,total_transport,total_attendance,'','',total_ppe,'','',sum_of_total,total_deduction,total_canteen,total_grand_total,total_ot])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=21)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=17)
    ws.merge_cells(start_row=2, start_column=18, end_row=2, end_column=21)
    ws.merge_cells(start_row=2, start_column=20, end_row=2, end_column=21)

    for cell in ws["2:2"]:
        cell.alignment = Alignment(horizontal='center')

    for header in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=21):
         for cell in header:
             cell.fill = PatternFill(fgColor='FFA07A', fill_type = "solid")

    for header in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=21):
         for cell in header:
             cell.fill = PatternFill(fgColor='6B8E23', fill_type = "solid")

    for header in ws.iter_rows(min_row=len(department)+4, max_row=len(department)+4, min_col=1, max_col=21):
         for cell in header:
             cell.fill = PatternFill(fgColor='6B8E23', fill_type = "solid")

    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal='center')
    for rows in ws.iter_rows(min_row = 2, max_row = 3, min_col=1, max_col=5):
        for cell in rows:
            cell.alignment = Alignment(horizontal='center')
    
    border = Border(left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'))

    for rows in ws.iter_rows(min_row=1, max_row=len(department)+4, min_col=1, max_col=21):
        for cell in rows:
            cell.border = border

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'