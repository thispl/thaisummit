from __future__ import unicode_literals
from os import stat
import frappe
from frappe.utils import cstr, add_days, date_diff,format_datetime
from frappe import _
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime, time

import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


class MonthlyIndividualAttendancereport(Document):
    pass

@frappe.whitelist()
def download():
    filename = 'Monthly Individual Attendance report'
    test = build_xlsx_response(filename)


# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)
    header = add_header(args)   
    ws.append(header)
    header_title = add_header_title()
    heading = add_heading()
    ws.append(header_title)
    ws.append(heading)

    data = get_data(args)

    for row in data:
        ws.append(row)
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
    ws.merge_cells(start_row=2, start_column=1, end_row=2  , end_column=6)
    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=10)
    ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10)

    ws['A1'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")
    ws['D1'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")    
    ws['A3'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")    
    ws['B3'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")
    ws['D3'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")
    ws['C3'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")
    ws['E3'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")
    ws['F3'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")
    ws['H1'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")
    ws['H3'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")
    ws['I3'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")
    ws['J3'].fill = PatternFill(fgColor="FEF701", fill_type = "solid")
    ws['B1'].fill = PatternFill(fgColor="33BB22", fill_type = "solid")    
    ws['E1'].fill = PatternFill(fgColor="0F7DE5", fill_type = "solid")
    ws['I1'].fill = PatternFill(fgColor="0F7DE5", fill_type = "solid")

    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
         
    dates = get_dates(args)

    for rows in ws.iter_rows(min_row=1, max_row=len(dates)+4, min_col=1, max_col=6):
        for cell in rows:
            cell.border = border

    for rows in ws.iter_rows(min_row=1, max_row=len(dates)+4, min_col=8, max_col=10):
        for cell in rows:
            cell.border = border
    
    ws.merge_cells(start_row=len(dates)+4, start_column=8, end_row= len(dates)+4, end_column=9)

    align_center = Alignment(horizontal='center')

    for rows in ws.iter_rows(min_row=1, max_row=len(dates)+4, min_col=1, max_col=10):
        for cell in rows:
            cell.alignment = align_center

    ws.sheet_view.zoomScale = 80
   

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

def get_dates(args):
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    return dates

@frappe.whitelist()
def add_header(args):
    employee_name = frappe.db.get_value("Employee",args.employee,"employee_name")
    header = ["ID:","%s"%(args.employee),'','NAME:',"%s"%(employee_name),'','','Dept', 'HR & Admin']
    return header

@frappe.whitelist()
def add_header_title():
    header = ["Attendance","","",'','','','',"Overtime",'','','']
    return header


@frappe.whitelist()
def add_heading():
    header = ["Date","Day","Working","In time","Out time","Status",'',"Start time",'End time','Total Hrs']
    return header

@frappe.whitelist()
def get_data(args):
    data = []
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
   
    total_ot = timedelta(0,0,0)
    for date in dates:
        if frappe.db.exists('Overtime Request',{'employee':args.employee,'ot_date':date,'workflow_state':'Approved'}):
            ot = frappe.db.get_value('Overtime Request',{'employee':args.employee,'ot_date':date,'workflow_state':'Approved'},'ot_hours')
            total_ot = total_ot + ot
        in_time = frappe.db.get_value('Attendance' ,{'employee':args.employee,"attendance_date":date},'in_time') or ''
        # in_time = datetime.strptime(in_time,'%H:%M:%S')
        out_time = frappe.db.get_value('Attendance' ,{'employee':args.employee,"attendance_date":date},'out_time') or ''
        # out_time = datetime.strptime(out_time,'%H:%M:%S')
        start_time =frappe.db.get_value("Overtime Request",{'employee':args.employee,'ot_date':date,'workflow_state':'Approved'},'from_time') or  ''
        if start_time:
            start_time = datetime.strptime(str(start_time),'%H:%M:%S')
            start_time = datetime.strftime(start_time,'%H:%M')
        end_time = frappe.db.get_value("Overtime Request",{'employee':args.employee,'ot_date':date,'workflow_state':'Approved'},'to_time') or  ''   
        if end_time:
            end_time = datetime.strptime(str(end_time),'%H:%M:%S')
            end_time = datetime.strftime(end_time,'%H:%M')
        total_hours = frappe.db.get_value("Overtime Request",{'employee':args.employee,'ot_date':date,'workflow_state':'Approved'},'ot_hours') or  ''
        if total_hours:
            total_hours = datetime.strptime(str(total_hours),'%H:%M:%S')
            total_hours = datetime.strftime(total_hours,'%H:%M')
        working  = check_holiday(date)  
        status_format = get_status(args,date)
        dt = datetime.strptime(date,'%Y-%m-%d')
        day_format = datetime.date(dt).strftime('%a')
        data.append([dt.strftime('%d-%b'),day_format,working or 'W',format_datetime(in_time),format_datetime(out_time),status_format,'',start_time,end_time,total_hours])
    # if total_ot:
    #     total_ot = datetime.strptime(str(total_ot),'%H:%M:%S')
    #     total_ot = datetime.strftime(total_ot,'%H:%M')
    data.append(['','','','','','','','Total OT','',total_ot])
    return data


def check_holiday(date):
    holiday = frappe.db.sql("""select `tabHoliday`.holiday_date,`tabHoliday`.weekly_off from `tabHoliday List` 
    left join `tabHoliday` on `tabHoliday`.parent = `tabHoliday List`.name where `tabHoliday List`.name = 'Holiday List - 2021' and holiday_date = '%s' """%(date),as_dict=True)
    if holiday:
        if holiday[0].weekly_off == 1:
            return "W/O"
        else:
            return "HH"

def get_status(args,date):
    status = ''
    if frappe.db.exists('Attendance',{'employee':args.employee,'attendance_date':date,'docstatus':['!=','2']}):
        shift_status = frappe.db.get_value('Attendance',{'employee':args.employee,'attendance_date':date,'docstatus':['!=','2']},['shift_status','employee_type'])
        if shift_status[1] == "WC":
            status = wc_status_map.get(shift_status[0], "")
        else:
            status = bc_status_map.get(shift_status[0], "")
    # status = frappe.db.get_value('Attendance' ,{'employee':args.employee,"attendance_date":date},'status') or ''
    # od = frappe.db.get_value('Attendance' ,{'employee':args.employee,"attendance_date":date},'on_duty_application') or '' 

    # if od:
    #     status = 'OD'
    # elif status in ("Present","Half Day"):
    #     status = frappe.db.get_value('Attendance' ,{'employee':args.employee,"attendance_date":date},'shift') or ''
    # elif status == "Absent":
    #     holiday = check_holiday(date)
    #     if holiday:
    #         status = holiday
    #     else:
    #         status = 'A'
    # elif status == "On Leave":
    #     leave_type = frappe.db.get_value('Attendance' ,{'employee':args.employee,"attendance_date":date},'leave_type') or ''
    #     if leave_type == 'Casual Leave':
    #         status = 'CL' 
    #     if leave_type == 'Sick Leave':
    #         status = 'SL'
    #     if leave_type == 'Earned Leave':
    #         status = 'EL'
    return status

bc_status_map = {
    "Absent": "AA",
    "AA":"AA",
    "Half Day": "HD",
    "Holiday": "HH",
    "Weekly Off": "WW",
    "1H": "1H",
    "2H": "2H",
    "3H": "3H",
    "1W": "1W",
    "2W": "2W",
    "3W": "3W",
    "PP1W": "PP1W",
    "PP2W": "PP2W",
    "1LH": "1LH",
    "2LH": "2LH",
    "3LH": "3LH",
    "1LW": "1LW",
    "2LW": "2LW",
    "3LW": "3LW",
    "PP1LW": "PP1LW",
    "PP2LW": "PP2LW",
    "MW": "MW",
    "On Leave": "L",
    "Present": "P",
    "Work From Home": "WFH",
    "MM": "AA",
    "11": "11",
    "22": "22",
    "33": "33",
    "PP1PP1": "P1P1",
    "PP2PP2": "P2P2",
    "1L1": "1L1",
    "2L2": "2L2",
    "3L3": "3L3",
    "2L3": "AA",
    "1L2": "AA",
    "3L1": "AA",
    "1LM": "1LM",
    "2LM": "2LM",
    "3LM": "3LM",
    "PP1LPP1": "P1LP1",
    "PP2LPP2": "P2LP2",
    "1M": "1M",
    "2M": "2M",
    "3M": "3M",
    "M1": "M1",
    "M2": "M2",
    "M3": "M3",
    "PP1M": "P1M",
    "PP2M": "P2M",
    "MPP1": "MP1",
    "MPP2": "MP2",
    "11": "11",
    "12": "12",
    "13": "13",
    "1PP1": "1P1",
    "1PP2": "1P2",
    "21": "21",
    "22": "22",
    "23": "23",
    "2PP1": "2P1",
    "2PP2": "2P2",
    "31": "31",
    "32": "32",
    "33": "33",
    "3PP1": "3P1",
    "3PP2": "3P2",
    "PP11": "P11",
    "PP12": "P12",
    "PP13": "P13",
    "PP1PP1": "P1P1",
    "PP1PP2": "P1P2",
    "PP21": "P2P1",
    "PP22": "P2P2",
    "PP23": "P2P3",
    "PP2PP1": "P2P1",
    "PP2PP2": "P2P2",
    "Earned Leave": "EL",
    "Casual Leave": "CL",
    "Sick Leave": "SL",
    "OD": "OD",
    "Compensatory Off": "CO",
    "Leave Without Pay": "LL",
    "0.5Earned Leave": "0.5EL",
    "0.5Casual Leave": "0.5CL",
    "0.5Sick Leave": "0.5SL",
    "0.5Compensatory Off": "0.5CO",
    "0.5Leave Without Pay": "0.5LL",
    "LEarned Leave/2": "0.5SL",
	"LCasual Leave/2": "0.5LCL",
	"LSick Leave/2": "0.5LSL",
	"LSpecial Leave/2": "0.5LSPL",
	"LCompensatory Off/2": "0.5LCO",
	"LLeave Without Pay/2": "0.5LLL",
    "0.5Special Leave": "0.5SPL",
    }

wc_status_map = {
    "Absent": "AA",
    "AA": "AA",
    "Half Day": "HD",
    "Holiday": "HH",
    "1H": "1H",
    "2H": "2H",
    "3H": "3H",
    "1W": "1W",
    "2W": "2W",
    "3W": "3W",
    "PP1W": "PP1W",
    "PP2W": "PP2W",
    "PP1H": "P1H",
    "PP2H": "P2H",
    "Weekly Off": "WW",
    "On Leave": "L",
    # "Present": "11",
    "Work From Home": "WFH",
    "OD": "OD",
    "Earned Leave": "EL",
    "Casual Leave": "CL",
    "Sick Leave": "SL",
    "Compensatory Off": "CO",
    "Leave Without Pay": "LL",
    "1": "11",
    "2": "22",
    "3": "33",
    " ": "MM",
    "PP1": "P1P1",
    "PP2": "P2P2",
    "1L": "1L1",
    "2L": "2L2",
    "3L": "3L3",
    "PP1L": "P1LP1",
    "PP2L": "P2LP2",
    "1LH": "1LH",
    "2LH": "2LH",
    "3LH": "3LH",
    "1LW": "1LW",
    "2LW": "2LW",
    "3LW": "3LW",
    "PP1LW": "PP1LW",
    "PP2LW": "PP2LW",
    "1M": "1M",
    "2M": "2M",
    "3M": "3M",
    "M1": "M1",
    "M2": "M2",
    "M3": "M3",
    "PP1M": "P1M",
    "PP2M": "P2M",
    "MPP1": "MP1",
    "MPP2": "MP2",
    "0.5Earned Leave": "0.5EL",
    "0.5Casual Leave": "0.5CL",
    "0.5Sick Leave": "0.5SL",
    "0.5Compensatory Off": "0.5CO",
    "0.5Leave Without Pay": "0.5LL",
    "LEarned Leave/2": "0.5SL",
	"LCasual Leave/2": "0.5LCL",
	"LSick Leave/2": "0.5LSL",
	"LSpecial Leave/2": "0.5LSPL",
	"LCompensatory Off/2": "0.5LCO",
	"LLeave Without Pay/2": "0.5LLL",
    "0.5Special Leave": "0.5SPL",
    }