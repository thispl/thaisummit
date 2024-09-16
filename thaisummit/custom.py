from codecs import ignore_errors
from multiprocessing.spawn import old_main_modules
from os import truncate
from types import FrameType
import frappe
import json
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment,Border,Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

import datetime,math
from frappe.utils.background_jobs import enqueue
from frappe import permissions
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import read_csv_content
from frappe.utils.data import format_date, get_url_to_list
from six.moves import range
from six import string_types
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
                          nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime, format_date,get_time)
from datetime import datetime
from calendar import IllegalMonthError, month, monthrange
from frappe import _, get_value, msgprint
from frappe.utils import flt, get_url_to_list
from frappe.utils import cstr, cint, getdate, get_first_day, get_last_day, today
import requests
from datetime import date, timedelta, time
import calendar
from erpnext.hr.doctype.employee.employee import get_holiday_list_for_employee
# from tabulate import tabulate
from datetime import datetime
from erpnext.hr.utils import get_holiday_dates_for_employee
import pandas as pd
from frappe.utils import getdate, cint, add_months, date_diff, add_days, nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime
from frappe import throw,_


@frappe.whitelist()
def update_list(production_line):
    prod_line_emp = []
    user = frappe.session.user
    if user != 'Administrator':
        emp_details = frappe.db.sql("""select name from `tabEmployee Production Line Details` where user_id ='%s' """%(user),as_dict=1)[0]
        emp_name = emp_details['name']
        emp_doc = frappe.get_doc('Employee Production Line Details', emp_name)
        emp_prod_line = emp_doc.get('production_line')
        for e in emp_prod_line:
            prod_line_emp.append(e.production_line_no)
        tag_production_line = production_line
        if tag_production_line in prod_line_emp:
            return tag_production_line
        else:
            frappe.throw(_('This Tag Card doesnt belongs to your production line'))

@frappe.whitelist()
def get_opq_api():
    mat_number = '20000614'
    total_open_qty = 0
    if frappe.db.exists('TSAI Part Master',{'mat_no':mat_number}):
        url = "http://apioso.thaisummit.co.th:10401/api/OpenProductionOrder"
        payload = json.dumps({
            "ProductNo": mat_number,
            "Fromdate": "",
            "Todate": ""
        })
        headers = {
            'Content-Type': 'application/json',
            'API_KEY': '/1^i[#fhSSDnC8mHNTbg;h^uR7uZe#ninearin!g9D:pos+&terpTpdaJ$|7/QYups;==~w~!AWwb&DU'
        }
        response = requests.request(
            "POST", url, headers=headers, data=payload)
        openqty = 0
        if response:
            stocks = json.loads(response.text)
            if stocks:
                openqty = stocks[0]['OpenQty']
                completed_qty = stocks[0]['CmpltQty']
                # planned_qty = stocks[0]['PlanedQty']
                for stock in stocks:
                    total_open_qty += cint(stock['OpenQty'])
    return total_open_qty or 0


@frappe.whitelist()
def get_opq(mat_number):
    total_open_qty = 0
    if frappe.db.exists('TSAI Part Master',{'mat_no':mat_number}):
        url = "http://apioso.thaisummit.co.th:10401/api/OpenProductionOrder"
        payload = json.dumps({
            "ProductNo": mat_number,
            "Fromdate": "",
            "Todate": ""
        })
        headers = {
            'Content-Type': 'application/json',
            'API_KEY': '/1^i[#fhSSDnC8mHNTbg;h^uR7uZe#ninearin!g9D:pos+&terpTpdaJ$|7/QYups;==~w~!AWwb&DU'
        }
        response = requests.request(
            "POST", url, headers=headers, data=payload)
        openqty = 0
        if response:
            stocks = json.loads(response.text)
            if stocks:
                openqty = stocks[0]['OpenQty']
                completed_qty = stocks[0]['CmpltQty']
                planned_qty = stocks[0]['PlanedQty']
                for stock in stocks:
                    total_open_qty += cint(stock['OpenQty'])
    return total_open_qty or 0

@frappe.whitelist()
def tag_mat_name(mat_number):
    mat_name = frappe.db.sql("""select parts_name,parts_no,packing_std,production_line from `tabTSAI Part Master` where mat_no ='%s' """%(mat_number),as_dict=1)[0]
    return mat_name['parts_name'],mat_name['parts_no'],mat_name['packing_std'],mat_name['production_line']

def bulk_update_from_csv(filename):
    # below is the method to get file from Frappe File manager
    from frappe.utils.file_manager import get_file
    # Method to fetch file using get_doc and stored as _file
    _file = frappe.get_doc("File", {"file_name": filename})
    # Path in the system
    filepath = get_file(filename)
    # CSV Content stored as pps

    pps = read_csv_content(filepath[1])
    for pp in pps:
        date = "2021-02-28"
        print(pp[0])
        print(frappe.db.exists('Attendance', {
              'employee': pp[0], 'attendance_date': date}))
        for p in pp:
            if p != pp[0]:
                date = add_days(date, 1)
                if not frappe.db.exists('Attendance', {'attendance_date': date, 'employee': pp[0]}):
                    print(date, p)
                    status = ''
                    late = 0
                    if p in ['1', '2', '3', 'PP1', 'PP2', '1L', '2L', '3L', 'PP1L', 'PP2L']:
                        # print(p)
                        att = frappe.new_doc("Attendance")
                        att.employee = pp[0]
                        att.attendance_date = date
                        att.status = 'Present'
                        if p in ['1L', '2L', '3L', 'PP1L', 'PP2L']:
                            att.late_entry = 1
                            att.shift = p.strip('L')
                        else:
                            att.shift = p
                        att.save(ignore_permissions=True)
                        att.submit()
                        frappe.db.commit()
                    elif p in ['A', 'SL', 'EL', 'CL']:
                        att = frappe.new_doc("Attendance")
                        att.employee = pp[0]
                        att.attendance_date = date
                        att.status = 'Absent'
                        att.save(ignore_permissions=True)
                        att.submit()
                        frappe.db.commit()
                    elif p == 'OD':
                        od = frappe.new_doc("Attendance Request")
                        od.employee = pp[0]
                        od.from_date = date
                        od.to_date = date
                        od.reason = 'On Duty'
                        od.save(ignore_permissions=True)
                        od.submit()
                        frappe.db.commit()

def mail_wc_probation():
    employees = frappe.get_all("Employee", {"employment_type": "Probation", "employee_type": "WC"}, [
                               "date_of_joining", "name", "employee_name", "personal_email"])
    # print(employees)
    for emp in employees:
        # print(emp)
        t = relativedelta(months=5)
        fifth_month = emp.date_of_joining+t
        # print(fifth_month)
        s = relativedelta(months=6)
        sixth_month = emp.date_of_joining+s
        # print(sixth_month)
        today = datetime.today()
        if(fifth_month <= today.date() <= sixth_month):
            # print(emp.personal_email)
            frappe.sendmail(
                recipients=emp.personal_email,
                subject='Probation Period Date',
                message="""<p>Dear Admin,</p>
            <p> Probation period end date for  %s %s employee  </p>""" % (emp.name, emp.employee_name))


def mail_wc_get_trainee():
    employees = frappe.get_all("Employee", {"employment_type": "Trainee", "employee_type": "WC", "designation": "Graduate Engineer Trainee"}, [
                               "date_of_joining", "name", "employee_name", "personal_email"])
    # print(employees)
    for emp in employees:
        # print(emp)
        t = relativedelta(months=11)
        eleventh_month = emp.date_of_joining+t
        # print(eleventh_month)
        s = relativedelta(months=12)
        tewlth_month = emp.date_of_joining+s
        # print(tewlth_month)
        today = datetime.today()
        if(eleventh_month <= today.date() <= tewlth_month):
            # print("hi")
            frappe.sendmail(
                recipients=emp.personal_email,
                subject='Probation Period Date',
                message="""<p>Dear Admin,</p>
            <p> Trainee period end date for  %s %s employee  </p>""" % (emp.name, emp.employee_name))


def mail_wc_get_probation():
    employees = frappe.get_all("Employee", {"employment_type": "Probation", "employee_type": "WC", "designation": "Graduate Engineer Trainee"}, [
                               "date_of_joining", "name", "employee_name", "personal_email"])
    # print(employees)
    for emp in employees:
        # print(emp)
        t = relativedelta(months=17)
        eleventh_month = emp.date_of_joining+t
        # print(eleventh_month)
        s = relativedelta(months=18)
        tewlth_month = emp.date_of_joining+s
        # print(tewlth_month)
        today = datetime.today()
        if(eleventh_month <= today.date() <= tewlth_month):
            # print("hi")
            frappe.sendmail(
                recipients=emp.personal_email,
                subject='Probation Period Date',
                message="""<p>Dear Admin,</p>
            <p> Probation period end date for  %s %s employee  </p>""" % (emp.name, emp.employee_name))


def el_leave_policy():
    employees = frappe.get_all("Employee", {"employee_type": "WC"}, [
                               "name", "employee_name", "personal_email"])
    for emp in employees:
        print(emp.name)
        count = 0
        el = 0
        today = "2022-01-01"
        start_date = add_months(today, -12)
        end_date = add_days(today, -1)
        print(end_date)
        print(start_date)
        for att in (frappe.db.sql("""select status,attendance_date from `tabAttendance` where attendance_date between '%s' and '%s'  """ % (start_date, end_date), as_dict=True)):
            print(att)
            if (att.status == "Present"):
                count += 1
            else:
                count = 0
            # print(count)
            if count >= 20:
                el = count // 20 + el
                count = 0
                break
        print(el)
        print(datetime.today().date())
        get_la = frappe.new_doc("Leave Allocation")
        get_la.employee = emp.name
        get_la.leave_type = "Earned Leave"
        get_la.from_date = datetime.today().date()
        get_la.to_date = add_months(datetime.today().date(), 12)
        get_la.new_leaves_allocated = (int(el))
        get_la.save(ignore_permissions=True)
        frappe.db.commit()


def el_leave_encashment():
    employees = frappe.get_all("Employee", {"employee_type": "WC"}, ["name"])
    for emp in employees:
        print(emp.name)
        # if(emp.name == "TSAI0195"):
        get_le = frappe.new_doc("Leave Encashment")
        get_le.employee = emp.name
        get_le.leave_type = "Earned Leave"
        get_le.leave_period = "HR-LPR-2021-00001"
        get_le.save(ignore_permissions=True)
        frappe.db.commit()


@frappe.whitelist()
def create_user(employee_number, employee_name):
    # frappe.errprint(employee_number)
    user = frappe.new_doc("User")
    user.email = employee_number + "@gmail.com"
    user.first_name = employee_name
    user.username = employee_number
    user.send_welcome_mail = 0
    user.new_password = "thai@1234"
    user.save(ignore_permissions=True)
    frappe.db.commit()
    return user.email


@frappe.whitelist()
def api_test():
    url = "http://apioso.thaisummit.co.th:10401/api/FGQuantity"
    payload = json.dumps({
                "ItemCode": "10000089"
            })
    headers = {
        'Content-Type': 'application/json',
        'API_KEY': '/1^i[#fhSSDnC8mHNTbg;h^uR7uZe#ninearin!g9D:pos+&terpTpdaJ$|7/QYups;==w!AWwb&DU'
    }
    response = requests.request(
                "POST", url, headers=headers, data=payload)
    frappe.errprint(response.status_code)

@frappe.whitelist()
def get_sap_qty():
    parts = frappe.get_all('Part Master', ['mat_no', 'name'])
    url = "http://172.16.1.18/StockDetail/Service1.svc/GetFGQuantity"
    headers = {
        'Content-Type': 'application/json'
    }
    response = requests.request("POST", url, headers=headers)
    frappe.errprint(response.status_code)
    if response.status_code == 200:
        for part in parts:
            url = "http://172.16.1.18/StockDetail/Service1.svc/GetFGQuantity"
            payload = json.dumps({
                "ItemCode": part['mat_no']
            })
            headers = {
                'Content-Type': 'application/json'
            }
            response = requests.request(
                "POST", url, headers=headers, data=payload)
            qty = json.loads(response.text)
            if qty:
                avl_qty = qty[0]['Qty']
                frappe.set_value(
                    "Part Master", part['name'], "sap_available_quantity", avl_qty)
                frappe.set_value(
                    "Part Master", part['name'], "sap_quantity_updated_on", datetime.now())
                frappe.set_value(
                    "Part Master", part['name'], "temp_avl_qty", avl_qty)
                frappe.set_value(
                    "Part Master", part['name'], "temp_qty_updated_on", datetime.now())
        return "Successfully Updated"
    else:
        return "Server Error"


@frappe.whitelist()
def deductions():
    last_mon = add_months(nowdate(), -1)
    today = nowdate()
    deduction = frappe.db.sql("""select employee, sum(amount) as total_amount from `tabDeductions` where docstatus=1 and
        MONTH(posting_date) = MONTH(%(mon)s) group by employee""", (dict(mon=last_mon)), as_dict=True)
    for ded in deduction:
        additional_salary = frappe.db.exists("Additional Salary", {
                                             "employee": ded.employee, "payroll_date": today, "salary_component": "Personal Protective Equipment", "docstatus": 1}, ["employee"])
        if not additional_salary:
            add_salary = frappe.new_doc("Additional Salary")
            add_salary.employee = ded.employee
            add_salary.salary_component = "Personal Protective Equipment"
            add_salary.amount = int(ded.total_amount)
            add_salary.payroll_date = nowdate()
            add_salary.overwrite_salary_structure_amount = "1"
            add_salary.save(ignore_permissions=True)
            add_salary.submit()
            frappe.db.commit()


@frappe.whitelist()
def change_leave_approver(employee, leave_approver):
    leave_application = frappe.db.get_all(
        "Leave Application", {"employee": employee, "docstatus": 0}, ["name"])
    for lev_app in leave_application:
        frappe.db.set_value('Leave Application', lev_app.name,
                            'leave_approver', leave_approver)


@frappe.whitelist()
def change_permission_approver(employee, permission_approver):
    permission_request = frappe.db.get_all(
        "Permission Request", {"employee": employee, "docstatus": 0}, ["name"])
    for per_req in permission_request:
        frappe.db.set_value('Permission Request', per_req.name,
                            'permission_approver', permission_approver)


@frappe.whitelist()
def change_od_approver(employee, od_approver):
    od_application = frappe.db.get_all(
        "On Duty Application", {"employee": employee, "docstatus": 0}, ["name"])
    for od_app in od_application:
        frappe.db.set_value('On Duty Application',
                            od_app.name, 'approver', od_approver)

def is_weekoff(employee, date=None, raise_exception=True):
    '''Returns True if given Employee has a weekoff on the given date
    :param employee: Employee `name`
    :param date: Date to check. Will check for today if None'''

    holiday_list = get_holiday_list_for_employee(employee, raise_exception)
    if not date:
        date = today()
    if holiday_list:
        return frappe.get_all('Holiday List', dict(name=holiday_list, holiday_date=date, weekly_off=1)) and True or False


def is_holiday(employee, date=None, raise_exception=True):
    '''Returns True if given Employee has an holiday on the given date
    :param employee: Employee `name`
    :param date: Date to check. Will check for today if None'''

    holiday_list = get_holiday_list_for_employee(employee, raise_exception)
    if not date:
        date = today()
    if holiday_list:
        return frappe.get_all('Holiday List', dict(name=holiday_list, holiday_date=date,  weekly_off=0)) and True or False


def send_birthday_wish():
    birthdays = frappe.db.sql("""select name,date_of_birth,personal_email, employee_name
        from tabEmployee where day(date_of_birth) = day(%(date)s)
        and month(date_of_birth) = month(%(date)s)
        and status = 'Active'""", {"date": today()}, as_dict=True)
    print(birthdays)
    for emp in birthdays:
        print(emp.personal_email)
        if birthdays:
            frappe.sendmail(
                recipients=emp.personal_email,
                subject="Birthday Wishes",
                message="""<p> Dear %s <br> May your birthday be the start of a year filled with good luck, good health and much happiness, <br>
            We wish you an amazing year that ends with accomplishing all the great goals that you have set!</p>""" % (emp.employee_name))


@frappe.whitelist()
def send_confirmation_mail(employee_number, employee_name, personal_email, confirmation_message):
    frappe.sendmail(
        recipients=personal_email,
        subject="Birthday Wishes",
        message="""<p> Dear %s <br>
    %s
    </p>""" % (employee_name, confirmation_message))


@frappe.whitelist()
def get_shift(emp):
    previous_month = frappe.utils.add_months(
        datetime.today(), -1).strftime("%Y-%m-26")
    current_month = (datetime.today()).strftime("%Y-%m-25")
    date_list = get_dates(previous_month, current_month)
    data = ''
    rh1 = '<tr>'
    rh2 = ''
    rh3 = ''
    rd1 = '<tr>'
    rd2 = ''
    rd3 = ''
    r1 = ''
    r2 = ''
    r3 = ''
    i = 1
    for date in date_list:
        if frappe.db.exists('Attendance', {'employee': emp, 'attendance_date': date, 'docstatus': ['!=', '2']}):
            att = frappe.get_doc('Attendance', {
                                 'employee': emp, 'attendance_date': date, 'docstatus': ['!=', '2']})
            # if rd1 == str(A):
            #     """<td style='color:red'> A </td>  """
            # frappe.errprint(date)
            frappe.errprint(att.shift)
            if i <= 10:
                rh1 += """<th><center>%s</center></th>""" % (
                    (datetime.strptime(date, '%Y-%m-%d').date()).strftime("%d-%b"))
                rd1 += """<td><center>%s</center></td>""" % (att.shift or 'A')
                if i == 10:
                    r1 = rh1 + '</tr>' + rd1 + '</tr>'
            elif 10 < i <= 20:
                rh2 += """<th><center>%s</center></th>""" % (
                    (datetime.strptime(date, '%Y-%m-%d').date()).strftime("%d-%b"))
                rd2 += """<td><center>%s</center></td>""" % (att.shift or 'A')
                if i == 20:
                    r2 = '<tr>' + rh2 + '</tr><tr>' + rd2 + '</tr>'
            elif 20 < i:
                rh3 += """<th><center>%s</center></th>""" % (
                    (datetime.strptime(date, '%Y-%m-%d').date()).strftime("%d-%b"))
                rd3 += """<td><center>%s</center></td>""" % (att.shift or 'A')
            i += 1
        data = "<table border='1' class='table table-bordered'>" + r1 + \
            r2 + '<tr>' + rh3 + '</tr><tr>' + rd3 + '</tr>' + "</table>"
    return data


def get_dates(previous_month, current_month):
    """get list of date s in between from date and to date"""
    no_of_days = date_diff(add_days(current_month, 1), previous_month)
    dates = [add_days(previous_month, i) for i in range(0, no_of_days)]
    return dates


def send_mail_hr():
    """Continous absent employees for 5 days mail sent to hr """
    current_date = (datetime.today()).strftime("%Y-%m-%d")
    five_days = frappe.utils.add_days(
        datetime.today(), -5).strftime("%Y-%m-%d")
    print(five_days)
    print(current_date)
    get_att = frappe.db.sql(""" select employee_name,employee,count(status) as absent_count from `tabAttendance` where status ='Absent' and docstatus = '1' and attendance_date between '%s' and '%s'  group by employee""" % (
        five_days, current_date), as_dict=True)
    print(get_att)
    data = ''
    header = """<tr> <th>SL No </th> <th> Employee ID</th><th> Employee Name </tr>"""
    for idx, att in enumerate(get_att):
        if att.absent_count == 5:
            data += """ <tr> <td> %s </td> <td> %s </td> <td> %s </td> </tr>""" % (
                idx+1, att.employee, att.employee_name)
    data = "<table border='1' class='table table-bordered'>" + header + data + "</table>"
    frappe.sendmail(
        recipients=['sarumathy.d@groupteampro.com'],
        subject="Absent Employees",
        message=data)


def create_leave_allocation():
    # current_date = (datetime.today()).strftime("%Y-%m-%d")
    three_days = frappe.utils.add_days(
        datetime.today(), -3).strftime("%Y-%m-%d")
    print(three_days)
    employees = frappe.get_all("Employee", {"status": "Active"}, [
                               "name", "employee_name", "personal_email"])
    for emp in employees:
        # print(emp.name)
        attendance = frappe.db.get_all('Attendance', {
                                       'employee': emp.name, 'attendance_date': three_days, 'docstatus': '1'}, ["status", "employee"])
        for att in attendance:
            # print(att.status)
            if att.status in ('Half Day'):
                from erpnext.hr.doctype.leave_application.leave_application import get_leave_details
                leave_balance = get_leave_details(att.employee, nowdate())
                try:
                    if (leave_balance['leave_allocation']['Casual Leave']['remaining_leaves']) > 0:
                        leave_app = frappe.new_doc("Leave Application")
                        leave_app.employee = att.employee
                        leave_app.from_date = three_days
                        leave_app.to_date = three_days
                        leave_app.leave_type = 'Casual Leave'
                        leave_app.description = 'Auto-Leave Application for Late In Entry'
                        leave_app.leave_approver = "Administrator"
                        leave_app.save(ignore_permissions=True)
                        frappe.db.commit()

                    elif(leave_balance['leave_allocation']['Sick Leave']['remaining_leaves']) > 0:
                        l_app = frappe.new_doc("Leave Application")
                        l_app.employee = att.employee
                        l_app.from_date = three_days
                        l_app.to_date = three_days
                        l_app.leave_type = 'Sick Leave'
                        l_app.description = 'Auto-Leave Application for Late In Entry'
                        l_app.leave_approver = "Administrator"
                        l_app.save(ignore_permissions=True)
                        frappe.db.commit()
                    elif (leave_balance['leave_allocation']['Earned Leave']['remaining_leaves']) > 0:
                        lea_app = frappe.new_doc("Leave Application")
                        lea_app.employee = att.employee
                        lea_app.from_date = three_days
                        lea_app.to_date = three_days
                        lea_app.leave_type = 'Earned Leave'
                        lea_app.description = 'Auto-Leave Application for Late In Entry'
                        lea_app.leave_approver = "Administrator"
                        lea_app.save(ignore_permissions=True)
                        frappe.db.commit()
                except KeyError:
                    leave_balance['leave_allocation']['Casual Leave'] = 0
                    leave_balance['leave_allocation']['Sick Leave'] = 0
                    leave_balance['leave_allocation']['Earned Leave'] = 0


def add_special_leave():
    lle = frappe.new_doc("Leave Ledger Entry")
    lle.employee = "5"
    lle.from_date = '2021-01-01'
    lle.to_date = '2021-12-31'
    lle.leave_type = 'Casual Leave'
    lle.transaction_type = 'Leave Allocation'
    lle.leaves = "10"
    lle.save(ignore_permissions=True)
    lle.submit()
    frappe.db.commit()


@frappe.whitelist()
def exceed_vehicle_load(name, vehicle_name):
    emp = frappe.db.count('Employee', {'vehicle_name': vehicle_name})
    load = frappe.db.get_value("Vehicle Management", {
                               'name': vehicle_name}, ['load_capacity'])
    balance_load = int(load - emp)
    # frappe.errprint(balance_load)
    # if(emp >= load):
    #     return "Vehicle Load Exceeded"
    return emp, load


@frappe.whitelist()
def mark_qr_user(user, status):
    user = frappe.get_doc('User', user)
    if status == 'Add':
        user.role_profile_name = ''
        user.add_roles('QR User')
        frappe.errprint('add')
    else:
        user.remove_roles('QR User')
        frappe.errprint('remove')
    user.save(ignore_permissions=True)


@frappe.whitelist()
def mark_qr_user_from_csv(filename):
    # below is the method to get file from Frappe File manager
    from frappe.utils.file_manager import get_file
    # Method to fetch file using get_doc and stored as _file
    _file = frappe.get_doc("File", {"file_name": filename})
    # Path in the system
    filepath = get_file(filename)
    # CSV Content stored as pps
    pps = read_csv_content(filepath[1])
    for pp in pps:
        # print(pp[0])
        userid = frappe.get_value('Employee', pp[0], 'user_id')
        username = frappe.db.exists('User', userid)
        if username:
            user = frappe.get_doc('User', username)
            user.role_profile_name = ''
            user.add_roles('QR User')
            user.save(ignore_permissions=True)

@frappe.whitelist()
def roundoff_time(time):
    time = datetime.strptime(time, '%H:%M:%S')
    if time.minute not in (0, 30):
        if time.minute < 30:
            roundoff_time = time.replace(minute=0)
        elif time.minute >= 30:
            roundoff_time = time.replace(minute=30)
        roundoff_time = roundoff_time.time()
        return str(roundoff_time)


@frappe.whitelist()
def get_employee_code(user):
    emp_id = frappe.db.get_value('Employee', {'user_id': user}, "name")
    return emp_id


@frappe.whitelist()
def get_approver(department, employee):
    user = frappe.db.get_value('Employee', employee, 'user_id')
    roles = frappe.get_roles(user)
    if 'GM' in roles:
        return frappe.db.get_value('Department', department, "ceo")
    elif 'HOD' in roles:
        return frappe.db.get_value('Department', department, "gm")
    else:
        return frappe.db.get_value('Department', department, "hod")


@frappe.whitelist()
def update_user_permission(user, department):
    if not frappe.db.exists("User Permission", {'user': user, 'allow': "Department", "for_value": department, "is_default": 1}):
        default = frappe.db.exists(
            "User Permission", {'user': user, 'allow': "Department", "is_default": '1'})
        if default:
            frappe.delete_doc("User Permission", default)
        up = frappe.db.exists("User Permission", {
                              'user': user, 'allow': "Department", 'for_value': department})
        if up:
            doc = frappe.get_doc("User Permission", up)
            doc.is_default = 1
            doc.save(ignore_permissions=True)
        else:
            doc = frappe.new_doc("User Permission")
            doc.user = user
            doc.allow = "Department"
            doc.for_value = department
            doc.is_default = 1
            doc.save(ignore_permissions=True)


@frappe.whitelist()
def check_leave_balance(employee, leave_type):
    from erpnext.hr.doctype.leave_application.leave_application import get_leave_details
    leave_balance = get_leave_details(employee, nowdate())
    if leave_type not in ('Special Leave', 'Compensatory Off', 'Leave Without Pay'):
        balance = leave_balance['leave_allocation'][leave_type]['remaining_leaves']
        la = frappe.db.count('Leave Application', {
                             'employee': employee, 'docstatus': 0})
        if la > balance:
            frappe.throw(
                'There is not enough leave balance for Leave Type %s' % (leave_type))

def bulk_mail_alerts():
    dept = frappe.get_all('Department', {'is_group': '0'})
    header = """<p>Dear sir, <br> Please find the below list of Application pending for your Approval.</p><table class='table table-bordered'> """
    link = get_url_to_list("Approval")
    view = "<table><tr><th style = 'border: 1px solid black;background-color:#ffedcc;'><center><a href='%s'>View Approval Summary</a></center></th></tr></table><br>" % (
        link)
    regards = "Thanks & Regards,<br>hrPRO"
    for d in dept:
        hod = frappe.db.get_value('Department', d.name, "hod")
        content = ''
        ots = ''
        miss_punch = ''
        leave_application = ''
        ots = frappe.get_all("Overtime Request", {"department": d.name, "workflow_state": "Pending for HOD"}, [
                             'employee', 'employee_name', 'ot_date', 'ot_hours'])
        if ots:
            content += """<table><tr><th style = 'border: 1px solid black;background-color:#ffedcc;' colspan = "5" ><center>Overtime Request</center></th></tr><tr><th style = 'border: 1px solid black'>Department</th><th style = 'border: 1px solid black'>Employee ID</th><th style = 'border: 1px solid black'>Employee Name</th><th style = 'border: 1px solid black'>OT Date</th><th style = 'border: 1px solid black'>OT Hour</th></tr>"""
            for ot in ots:
                content += '<tr><td style = "border: 1px solid black"> %s </td><td style = "border: 1px solid black"> %s </td><td style = "border: 1px solid black"> %s </td> <td style = "border: 1px solid black"> %s </td> <td style = "border: 1px solid black"> %s </td></tr>' % (
                    d['name'], ot['employee'], ot['employee_name'], format_date(ot['ot_date']), ot['ot_hours'])
            content += '</table><br>'

        miss_punch = frappe.get_all("Miss Punch Application", {"department": d.name, "workflow_state": "Pending for HOD"}, [
                                    'employee', 'employee_name', 'in_time', 'out_time', 'attendance_date'])
        if miss_punch:
            content += """<table class='table table-bordered'><tr><th style = 'border: 1px solid black;background-color:#ffedcc;' colspan = "6"><center>Manual Attendance Correction</center></th></tr><tr><th style = 'border: 1px solid black'>Department</th><th style = 'border: 1px solid black'>Employee ID</th><th style = 'border: 1px solid black'>Employee Name</th><th style = 'border: 1px solid black'>Date</th><th style = 'border: 1px solid black'>IN Time</th><th style = 'border: 1px solid black'>OUT Time</th></tr>"""
            for mp in miss_punch:
                content += '<tr><td style = "border: 1px solid black"> %s </td><td style = "border: 1px solid black"> %s </td><td style = "border: 1px solid black"> %s </td> <td style = "border: 1px solid black"> %s </td><td style = "border: 1px solid black"> %s </td> <td style = "border: 1px solid black"> %s </td></tr>' % (
                    d['name'], mp.employee, mp.employee_name, mp.attendance_date, format_datetime(mp.in_time), format_datetime(mp.out_time))
            content += '</table><br>'

        leave_application = frappe.get_all("Leave Application", {
                                           "department": d.name, "workflow_state": "Pending for HOD"}, ['*'])
        if leave_application:
            content += """<table class='table table-bordered'><tr><th colspan = "8" style = 'border: 1px solid black;background-color:#ffedcc;'><center>Leave Application</center></th></tr><tr><th style = 'border: 1px solid black'>Department</th><th style = 'border: 1px solid black'>Employee ID</th><th style = 'border: 1px solid black'>Employee Name</th><th style = 'border: 1px solid black'>From Date</th><th style = 'border: 1px solid black'>To Date</th><th style = 'border: 1px solid black'>Leave Type</th><th style = 'border: 1px solid black'>Session</th><th style = 'border: 1px solid black'>Reason</th></tr>"""
            for leave in leave_application:
                content += '<tr><td style = "border: 1px solid black"> %s </td><td style = "border: 1px solid black"> %s </td><td style = "border: 1px solid black"> %s </td><td style = "border: 1px solid black"> %s </td> <td style = "border: 1px solid black"> %s </td><td style = "border: 1px solid black"> %s </td> <td style = "border: 1px solid black"> %s </td> <td style = "border: 1px solid black"> %s </td></tr>' % (
                    d['name'], leave.employee, leave.employee_name, format_date(leave.from_date), format_date(leave.to_date), leave.leave_type, leave.session, leave.description)
            content += '</table><br>'

        if ots or miss_punch or leave_application:
            print(d.name)
            if hod:
                frappe.sendmail(
                    recipients=[hod, 'mohan.pan@thaisummit.co.in'],
                    subject='Reg.List of pending Approvals',
                    message=header+content+view+regards)


@frappe.whitelist()
def update_approver(dept, hod, gm, ceo):
    ots = frappe.db.get_all("Overtime Request", {
                            "department": dept, "docstatus": 0}, ["name", "employee"])
    for ot in ots:
        user = frappe.db.get_value('Employee', ot.employee, 'user_id')
        roles = frappe.get_roles(user)
        if 'GM' in roles:
            approver_id = frappe.db.get_value("User", ceo, 'username')
            approver_name = frappe.db.get_value("User", ceo, 'full_name')
            frappe.db.set_value('Overtime Request', ot.name, 'approver', ceo)
            frappe.db.set_value('Overtime Request', ot.name,
                                'approver_id', approver_id)
            frappe.db.set_value('Overtime Request', ot.name,
                                'approver_name', approver_name)

        elif 'HOD' in roles:
            approver_id = frappe.db.get_value("User", gm, 'username')
            approver_name = frappe.db.get_value("User", gm, 'full_name')
            frappe.db.set_value('Overtime Request', ot.name, 'approver', gm)
            frappe.db.set_value('Overtime Request', ot.name,
                                'approver_id', approver_id)
            frappe.db.set_value('Overtime Request', ot.name,
                                'approver_name', approver_name)
        else:
            approver_id = frappe.db.get_value("User", hod, 'username')
            approver_name = frappe.db.get_value("User", hod, 'full_name')
            frappe.db.set_value('Overtime Request', ot.name, 'approver', hod)
            frappe.db.set_value('Overtime Request', ot.name,
                                'approver_id', approver_id)
            frappe.db.set_value('Overtime Request', ot.name,
                                'approver_name', approver_name)

    leave_apps = frappe.db.get_all("Leave Application", {
                                   "department": dept, "docstatus": 0}, ["name", "employee"])
    for la in leave_apps:
        user = frappe.db.get_value('Employee', la.employee, 'user_id')
        roles = frappe.get_roles(user)
        if 'GM' in roles:
            approver_name = frappe.db.get_value("User", ceo, 'full_name')
            frappe.db.set_value('Leave Application',
                                la.name, 'leave_approver', ceo)
            frappe.db.set_value('Leave Application', la.name,
                                'leave_approver_name', approver_name)

        elif 'HOD' in roles:
            approver_name = frappe.db.get_value("User", gm, 'full_name')
            frappe.db.set_value('Leave Application',
                                la.name, 'leave_approver', gm)
            frappe.db.set_value('Leave Application', la.name,
                                'leave_approver_name', approver_name)
        else:
            approver_name = frappe.db.get_value("User", hod, 'full_name')
            frappe.db.set_value('Leave Application',
                                la.name, 'leave_approver', hod)
            frappe.db.set_value('Leave Application', la.name,
                                'leave_approver_name', approver_name)

    permissions = frappe.db.get_all("Permission Request", {
                                    "department": dept, "docstatus": 0}, ["name", "employee"])
    for p in permissions:
        user = frappe.db.get_value('Employee', p.employee, 'user_id')
        roles = frappe.get_roles(user)
        if 'GM' in roles:
            approver_name = frappe.db.get_value("User", ceo, 'full_name')
            frappe.db.set_value('Permission Request', p.name,
                                'permission_approver', ceo)
            frappe.db.set_value('Permission Request', p.name,
                                'permission_approver_name', approver_name)

        elif 'HOD' in roles:
            approver_name = frappe.db.get_value("User", gm, 'full_name')
            frappe.db.set_value('Permission Request', p.name,
                                'permission_approver', gm)
            frappe.db.set_value('Permission Request', p.name,
                                'permission_approver_name', approver_name)
        else:
            approver_name = frappe.db.get_value("User", hod, 'full_name')
            frappe.db.set_value('Permission Request', p.name,
                                'permission_approver', hod)
            frappe.db.set_value('Permission Request', p.name,
                                'permission_approver_name', approver_name)

    ods = frappe.db.get_all("On Duty Application", {
                            "department": dept, "docstatus": 0}, ["name", "employee"])
    for od in ods:
        user = frappe.db.get_value('Employee', od.employee, 'user_id')
        roles = frappe.get_roles(user)
        if 'GM' in roles:
            approver_name = frappe.db.get_value("User", ceo, 'full_name')
            frappe.db.set_value('On Duty Application',
                                od.name, 'approver', ceo)
            frappe.db.set_value('On Duty Application', od.name,
                                'approver_name', approver_name)

        elif 'HOD' in roles:
            approver_name = frappe.db.get_value("User", gm, 'full_name')
            frappe.db.set_value('On Duty Application', od.name, 'approver', gm)
            frappe.db.set_value('On Duty Application', od.name,
                                'approver_name', approver_name)
        else:
            approver_name = frappe.db.get_value("User", hod, 'full_name')
            frappe.db.set_value('On Duty Application',
                                od.name, 'approver', hod)
            frappe.db.set_value('On Duty Application', od.name,
                                'approver_name', approver_name)
    return 'ok'

@frappe.whitelist()
def check_qr(from_date, to_date, employee):
    emp_type = frappe.db.get_value("Employee", employee, 'employee_type')
    if emp_type != 'WC':
        qr = frappe.db.sql("select name from `tabQR Checkin` where employee = '%s' and shift_date between '%s' and '%s' " % (
            employee, from_date, to_date), as_dict=True)
        if qr:
            return qr

def create_ss():
    emps = frappe.get_all('Employee', {'status': 'Active'}, ['*'])
    for emp in emps:
        structure = ''
        if not frappe.db.exists('Salary Structure Assignment', {'employee': emp.name}):
            if emp.employee_type == 'WC':
                structure = 'WC'
            elif emp.employee_type == 'BC':
                structure = 'BC Salary Structure'
            elif emp.employee_type == 'CL':
                structure = 'CL Structure'
            if structure:
                try:
                    doc = frappe.new_doc('Salary Structure Assignment')
                    doc.employee = emp.name
                    doc.salary_structure = structure
                    doc.from_date = '2021-07-24'
                    doc.save(ignore_permissions=True)
                    doc.submit()
                    frappe.db.commit()
                except:
                    doc = frappe.new_doc('Salary Structure Assignment')
                    doc.employee = emp.name
                    doc.salary_structure = structure
                    doc.from_date = frappe.db.get_value(
                        'Employee', emp.name, 'date_of_joining')
                    doc.save(ignore_permissions=True)
                    doc.submit()
                    frappe.db.commit()

@frappe.whitelist()
def fetch_sap_stock():
    # parts = frappe.get_all('Part Master',['mat_no','name'])
    # for part in parts:
    url = "http://apioso.thaisummit.co.th:10401/api/FGQuantity"
    date = str(add_days(today(), -1)).replace('-', '')
    # date = "20220125"
    print(date)
    payload = json.dumps({
        "Fromdate": date,
        "Todate": date,
        "ItemCode": ""
    })
    headers = {
    'Content-Type': 'application/json',
    'API_KEY': '/1^i[#fhSSDnC8mHNTbg;h^uR7uZe#ninearin!g9D:pos+&terpTpdaJ$|7/QYups;==~w~!AWwb&DU'
    }
    response = requests.request("POST", url, headers=headers, data=payload)
    # frappe.log_error(response.text)
    sap_stock = json.loads(response.text)
    if sap_stock:
        d = pd.to_datetime(date)
        frappe.db.sql(
            "delete from `tabSAP Outgoing Report` where ar_invoice_date between '%s' and '%s' " % (date, date))
        for ss in sap_stock:
            ar_invoice_date = pd.to_datetime(ss['AR_Invoice_Date'])
            ar_invoice_date = ar_invoice_date.to_pydatetime()
            ar_invoice_date = ar_invoice_date.date()
            # sapog_id = frappe.db.exists("SAP Outgoing Report",{"ar_invoice_date":ar_invoice_date,"ar_invoice_no": ss['AR_Invoice_No']})
            # if sapog_id:
            #     sapog = frappe.get_doc("SAP Outgoing Report",sapog_id)
            # else:
            sapog = frappe.new_doc("SAP Outgoing Report")
            sapog.update({
                "ar_invoice_date": ar_invoice_date,
                "ar_invoice_no": ss['AR_Invoice_No'],
                "base_price": ss['Base_Price'],
                "base_value": ss['Base_Value'],
                "customer_code": ss['Customer_Code'],
                "customer_name": ss['Customer_Name'],
                "customer_ref_no": ss['Customer_Ref_No'],
                "item_service_description": ss['Dscription'],
                "fm": ss['FM'],
                "part_no": ss['Part_No'],
                "quantity": ss['Qty'],
                "sales_price": ss['Sales_Price'],
                "sales_value": ss['Sales_Value'],
            })
            sapog.save(ignore_permissions=True)
            frappe.db.commit()


# @frappe.whitelist()
# def test_hook():
    # frappe.log_error(title='hooks', message='ok')

@frappe.whitelist()
def fetch_sap_stock_bulk():
    # parts = frappe.get_all('Part Master',['mat_no','name'])
    # for part in parts:
    url = "http://172.16.1.18/StockDetail/Service1.svc/GetFGQuantity"
    date = str(add_days(today(), -1)).replace('-', '')
    print(date)
    payload = json.dumps({
        "Fromdate": "20211109",
        "Todate": "20211109",
        "ItemCode": ""
    })
    headers = {
        'Content-Type': 'application/json'
    }
    response = requests.request("POST", url, headers=headers, data=payload)
    sap_stock = json.loads(response.text)
    if sap_stock:
        for ss in sap_stock:
            ar_invoice_date = pd.to_datetime(ss['AR_Invoice_Date'])
            ar_invoice_date = ar_invoice_date.to_pydatetime()
            ar_invoice_date = ar_invoice_date.date()
            # sapog_id = frappe.db.exists("SAP Outgoing Report",{"ar_invoice_date":ar_invoice_date,"ar_invoice_no": ss['AR_Invoice_No']})
            # if sapog_id:
            #     sapog = frappe.get_doc("SAP Outgoing Report",sapog_id)
            # else:
            sapog = frappe.new_doc("SAP Outgoing Report")
            sapog.update({
                "ar_invoice_date": ar_invoice_date,
                "ar_invoice_no": ss['AR_Invoice_No'],
                "base_price": ss['Base_Price'],
                "base_value": ss['Base_Value'],
                "customer_code": ss['Customer_Code'],
                "customer_name": ss['Customer_Name'],
                "customer_ref_no": ss['Customer_Ref_No'],
                "item_service_description": ss['Dscription'],
                "fm": ss['FM'],
                "part_no": ss['PartNo'],
                "quantity": ss['Qty'],
                "sales_price": ss['Sales_Price'],
                "sales_value": ss['Sales_Value'],
            })
            sapog.save(ignore_permissions=True)
            frappe.db.commit()


@frappe.whitelist()
def fetch_sap_production():
    url = "http://172.16.1.18/StockDetail/Service1.svc/GetProductionPlanQty"
    from_date = str(add_days(today(), -1)).replace('-', '')
    to_date = str(today()).replace('-', '')
    # from_date = '20220131'
    # to_date = '20220210'
    print(from_date)
    print(to_date)
    payload = json.dumps({
        "Fromdate": from_date,
        "Todate": to_date
    })
    headers = {
        'Content-Type': 'application/json'
    }
    response = requests.request("POST", url, headers=headers, data=payload)
    sap_production = json.loads(response.text)
    print(sap_production)
    if sap_production:
        frappe.db.sql("delete from `tabSAP Production Plan` where order_date between '%s' and '%s' " % (
            pd.to_datetime(from_date).date(), pd.to_datetime(to_date).date()))
        for sp in sap_production:
            order_date = datetime.strptime(sp['OrderDate'], '%d-%m-%Y').date()
            due_date = datetime.strptime(sp['DueDate'], '%d-%m-%Y').date()
            start_date = datetime.strptime(sp['StartDate'], '%d-%m-%Y').date()
            sappp = frappe.new_doc("SAP Production Plan")
            sappp.update({
                "completed_quantity": sp['CompletedQty'],
                "doc_no": sp['DocNum'],
                "doc_series": sp['DocSeries'],
                "due_date": due_date,
                "open_quantity": sp['OpenQty'],
                "order_date": order_date,
                "planned_quantity": sp['PlanQty'],
                "product_description": sp['ProdDesc'],
                "product_no": sp['ProdNo'],
                "rejected_quantity": sp['RejectedQty'],
                "select": sp['Select'],
                "start_date": start_date,
                "status": sp['Status'],
                "type": sp['Type']
            })
            sappp.save(ignore_permissions=True)
            frappe.db.commit()

def add_role():
    user = frappe.get_doc('User', 'yukesh.sri@thaisummit.co.in')
    user.save(ignore_permissions=True)
    frappe.db.commit()


@frappe.whitelist()
def get_dispatch_data(doc):
    i = 1
    data = """
            <thead>
            <tr>
            <th>SL.No</th>
            <th>CARD RECIEVED DATE & TIME</th>
            <th>EXPECTED DISPATCH ON</th>
            <th>DISPATCH DATE & TIME</th>
            <th>MAT NO</th>
            <th>PART NO</th>
            <th>PART NAME</th>
            <th>MODEL</th>
            <th>DISPATCH QTY</th>
            <th>SAP STOCK</th>
            <th>DISPATCH READINESS STATUS</th>
            </tr>
            </thead>
    """
    for row in doc.tag_wise_list:
        exp_dispatch_time = datetime.strptime(
            str(row.datetime), "%Y-%m-%d %H:%M:%S") + timedelta(minutes=55)
        frappe.errprint(exp_dispatch_time)
        frappe.errprint(type(exp_dispatch_time))
        data += """
        <tr>
            <td>%s</td>
            <td>%s</td>
            <td>%s</td>
            <td></td>
            <td>%s</td>
            <td>%s</td>
            <td>%s</td>
            <td>%s</td>
            <td>%s</td>
            <td>%s</td>
            <td></td>
        </tr>""" % (i, row.datetime, exp_dispatch_time, row.mat_no, row.parts_no, row.parts_name, row.model, row.required_quantity, row.sap_quantity)
        i += 1

    return cstr(data)


def run_method():
    ots = frappe.db.sql('select name,employee from `tabOvertime Request` where department is null ',as_dict=True)
    for ot in ots:
        print(ot)
        dept = frappe.db.get_value('Employee',{'name':ot.employee},'department')
        print(dept)
        frappe.db.set_value('Overtime Request',ot.name,'department',dept)
    
def delete_left_att(doc, method):
    atts = frappe.get_all('Attendance', {'employee': doc.name, 'attendance_date': (
        '>', doc.relieving_date), 'docstatus': ('!=', '2')})
    for att in atts:
        checkins = frappe.get_all('Employee Checkin', {'attendance': att.name})
        for c in checkins:
            frappe.db.set_value('Employee Checkin', c.name, 'attendance', '')

        attendance = frappe.get_doc('Attendance', att.name)
        try:
            attendance.cancel()
            frappe.delete_doc('Attendance', attendance.name)
        except:
            frappe.delete_doc('Attendance', attendance.name)

@frappe.whitelist()
def leave_application(leave_day,emp,leave_type,name,reason,approve):
    leave  = frappe.db.exists('Leave Application',{'from_date':leave_day,'employee':emp,'leave_type':leave_type})
    if leave :
        return leave
    else:
        doc = frappe.new_doc('Leave Allocation')
        doc.employee = emp
        doc.leave_type = leave_type
        doc.from_date = leave_day
        doc.to_date = leave_day
        doc.session = 'Full Day'
        doc.description = reason
        doc.leave_approver = approve
        doc.save(ignore_permissions=True)
        frappe.db.commit()

def group_by():
    # data = [['10000001',50,'sdfs00'],
    # ['31000001', 500,'sdfasdf'],
    # ['10000001', 200,'sdfs00'],
    # ['31000002', 200,'j'],
    # ['31000002', 200,'2']]

    import pandas as pd
    # df = pd.DataFrame(data, columns = ['Name', 'Age','test'])
    # df2 = df.groupby(['Name','test'])['Age'].sum().reset_index()
    # l = df2.values.tolist()
    data = [['abcd', 'asdfs'], ['sasddfs', 'assdf']]
    cols = ['1', '2']

    df = pd.DataFrame(data, columns=cols)
    print(df)
    # print(df2)
    # print(l)

def check_ot():
    ots = frappe.db.sql("""select `tabOvertime Request`.employee as employee,`tabOvertime Request`.ot_hours as ot_hours from `tabOvertime Request` left JOIN 
    `tabEmployee` on `tabOvertime Request`.employee = `tabEmployee`.name where `tabEmployee`.employee_type = 'TH' and `tabOvertime Request`.ot_date between '2021-10-26' and '2021-11-25' and `tabOvertime Request`.workflow_state in ('Pending for HOD','Approved') and `tabOvertime Request`.docstatus != '2' """, as_dict=1)
    for ot in ots:
        print(ot)

@frappe.whitelist()
def ot():
    ot_request = frappe.db.get_all('Overtime Request',{'ot_date':'2022-07-01','workflow_state':'Draft'},['employee'])
    for ot in ot_request:
        print(ot)
        att = frappe.db.get_all('Attendance',{'attendance_date':'2022-07-01','employee':ot.employee},['in_time','out_time',])
        for a in att:
            if a.in_time and a.out_time:
                print(a.employee)
            else:
                print(a.employee)    


@frappe.whitelist()
def get_dates():
    data = 0
    rate = [1,2,3,4,5]
    for i in rate:
        if i:
            data+= i
            print(data)

@frappe.whitelist()
def mark_biometric_pin(doc,method):
    employee_series = 0
    if doc.employee_type == 'FT':
        employee_series = 25
       
    elif doc.employee_type == 'NT':
        employee_series = 20
        
    elif doc.employee_type == 'BC':
        employee_series = 30
       
    elif doc.employee_type == 'WC':
        employee_series = 10
    
    elif doc.employee_type == 'TH':
        employee_series = 99
       
    elif doc.employee_type == 'CL':
            employee_series = frappe.db.get_value("Contractor",{'name': doc.contractor},['employee_series'])        
    number_seperation =re.sub('\D', '', doc.employee_number)
    bio_metric_pin = str(employee_series) + str(number_seperation)
    frappe.db.set_value("Employee",doc.name,'biometric_pin',bio_metric_pin)

@frappe.whitelist()
def get_validation_dates():
    validation_days  = frappe.db.get_single_value('HR Time Settings','on_duty_validation_dates')
    # return validation_days
    print(validation_days)
    
@frappe.whitelist()
def ot_dept_count():
    data = []
    emp = []
    # wh = 0
    payroll_start_date = '2022-09-26'
    payroll_end_date = '2022-10-25'
    dept = frappe.db.get_all('Overtime Request',{'department':'HR&ADMIN','ot_date':('between',(payroll_start_date,payroll_end_date))},['*'])
    for d in dept:
        ftr = [3600,60,1]
        try:
            hr = sum([a*b for a,b in zip(ftr, map(int,str(d.ot_hours).split(':')))])
            wh = round(hr/3600,1)
            data.append(wh)
            emp.append(d.employee)
            # data +=wh
        except:
            wh = 0  
            # data.append(wh)
    print(sum(data))


# # calculating cgst and sgst for tsai invoice
# @frappe.whitelist()
# def get_gst_percent(doc,method):
#     doc_name = frappe.get_doc("TSAI Invoice",doc.name)
#     children = doc_name.invoice_items
#     cgst = 0
#     count = 0
#     sgst = 0
#     igst = 0
#     for c in children:
#         cgst += c.cgst
#         count += 1
#         sgst += c.sgst
#         igst += c.igst
#     tot_cgst = cgst/count
#     tot_sgst = sgst/count
#     tot_igst = igst/count
#     if doc.igst > 0 :
#         igst_amount = float(tot_igst / 100) * float(doc.total_basic_amount)
#         tot_amount = igst_amount + doc.total_basic_amount
#         frappe.db.set_value("TSAI Invoice",doc.name,"total_invoice_amount",tot_amount)
#     else:
#         total_gst_amount = (tot_cgst / 100) * 2
#         grand_total_gst = ((float(doc.total_basic_amount) * (total_gst_amount)))
#         tot_amount = (grand_total_gst + doc.total_basic_amount)
#         frappe.db.set_value("TSAI Invoice",doc.name,"cgst",tot_cgst)
#         frappe.db.set_value("TSAI Invoice",doc.name,"sgst",tot_sgst)
#         frappe.db.set_value("TSAI Invoice",doc.name,"total_gst_amount",grand_total_gst)
#         frappe.db.set_value("TSAI Invoice",doc.name,"total_invoice_amount",tot_amount)

@frappe.whitelist()
def update_ots():
    ot = frappe.db.sql(""" select * from `tabQR Checkin` where shift_date between '2023-04-15' and '2023-04-15' and ot = 1 """,as_dict=True)
    for o in ot:
        if frappe.db.exists("Overtime Request", {"employee":o.employee,"ot_date":o.shift_date}):
            print("HI")
        else:
            print(o.employee)
            ot = frappe.new_doc('Overtime Request')
            ot.employee = o.employee
            ot.department = o.department
            ot.ot_date = o.shift_date
            ot.shift = o.qr_shift
            shift_start = frappe.db.get_value('Shift Type',o.qr_shift,"start_time")
            shift_start_time = datetime.strptime(str(shift_start), '%H:%M:%S')
            qr_shift = datetime.strptime(str(o.created_date), '%Y-%m-%d')
            ot.from_time = datetime.combine(qr_shift,shift_start_time.time())
            ot.to_time = ""
            ot.total_hours = ""
            ot.total_wh = ""
            ot.ot_hours = ""
            ot.save(ignore_permissions=True)
            frappe.db.commit()
            print("HII")


@frappe.whitelist()
def get_ot_hours():
    ot = frappe.db.sql("""update `tabOvertime Request` set ot_hours = 15:00 where name = "OT-114317" """,as_dict = True)
    print(ot)


@frappe.whitelist()
def get_ot_amount(from_date,to_date):
    ot_req = frappe.db.sql("""select * from `tabOvertime Request` where ot_date between '%s' and '%s' and workflow_state = 'Approved' """%(from_date,to_date),as_dict=True)
    for ot in ot_req:
        ot_hr = get_time(ot.ot_hours)
        ftr = [3600,60,1]
        hr = sum([a*b for a,b in zip(ftr, map(int,str(ot_hr).split(':')))])
        ot_hrs = round(hr/3600,1)
        # print(ot.name)
        # print(ot_hrs)
        if ot.employee_type != 'CL':
            basic = ((frappe.db.get_value('Employee',ot.employee,'basic')/26)/8)*2
            print(basic)
            frappe.db.set_value('Overtime Request',ot.name,'ot_basic',basic)
            frappe.db.set_value('Overtime Request',ot.name,'ot_amount',round(ot_hrs*basic))
            # print("Changed")
        else:
            basic = 0
            designation = frappe.db.get_value('Employee',ot.employee,'designation')
            if designation == 'Skilled':
                basic = frappe.db.get_single_value('HR Time Settings','skilled_amount_per_hour')
                # print(basic)
            elif designation == 'Un Skilled':
                basic = frappe.db.get_single_value('HR Time Settings','unskilled_amount_per_hour')
                # print(basic)
            frappe.db.set_value('Overtime Request',ot.name,'ot_basic',basic)
            frappe.db.set_value('Overtime Request',ot.name,'ot_amount',round(ot_hrs*basic))

    return "ok"

from frappe import msgprint, _
@frappe.whitelist()
def set_restrictions_for_leaves(doc,method):
    user = frappe.session.user
    user_roles = frappe.get_roles(user)
    if not ("System Manager" in user_roles):
        if doc.employee_type == "BC":
            leave_list = frappe.db.sql("""select count(*) as count from `tabLeave Application` where from_date between '%s' and '%s' and to_date between '%s' and '%s' and employee_type = 'BC' and docstatus != '2' """%(doc.from_date,doc.to_date,doc.from_date,doc.to_date), as_dict=True)
            frappe.errprint(leave_list[0].count)
            frappe.errprint(int(frappe.db.get_single_value('HR Time Settings','leave_limit_for_bc')))
            if int(leave_list[0].count + 1) > int(frappe.db.get_single_value('HR Time Settings','leave_limit_for_bc')):
                frappe.throw(_('Today Leave Limit for BC employees has been Reached. For additional details kindly contact the HR Team'))

from frappe import msgprint, _
@frappe.whitelist()
def set_restrictions_for_leave(from_date,to_date,emp_type):
    date_list = get_dates(from_date,to_date)
    user = frappe.session.user
    user_roles = frappe.get_roles(user)
    # if not ("System Manager" in user_roles):
    if emp_type == "BC":
        for d in date_list:
            leave_list = frappe.db.sql("""select count(*) as count from `tabLeave Application` where employee_type = 'BC' and docstatus != '2' and from_date between '%s' and '%s' and to_date between '%s' and '%s' """%(d,d,d,d), as_dict=True)
            frappe.errprint(leave_list[0].count)
            frappe.errprint(int(frappe.db.get_single_value('HR Time Settings','leave_limit_for_bc')))
            if int(leave_list[0].count + 1) > int(frappe.db.get_single_value('HR Time Settings','leave_limit_for_bc')):
                return "OK"
                # frappe.throw(_('Today Leave Limit for BC employees has been Reached. For additional details kindly contact the HR Team'))

def get_dates(from_date,to_date):
    no_of_days = date_diff(add_days(to_date, 1), from_date)
    # frappe.errprint(no_of_days)
    dates = [add_days(from_date, i) for i in range(0, no_of_days)]
    # frappe.errprint(dates)
    return dates

@frappe.whitelist()
def leave_att(doc,method):
    if frappe.db.exists("Attendance",{'attendance_date':('between',(doc.from_date,doc.to_date)),'employee':doc.employee,'docstatus':1}):
        frappe.throw(_('Attendance Closed between these days %s and %s'%(doc.from_date,doc.to_date)))        
                    
@frappe.whitelist()
def miss_att(doc,method):
    if frappe.db.exists("Attendance",{'attendance_date':doc.attendance_date,'employee':doc.employee,'docstatus':1}):
        frappe.throw(_('Attendance Closed for this day %s'%(doc.attendance_date)))        
                    
@frappe.whitelist()
def ot_att(doc,method):
    user = frappe.session.user
    user_roles = frappe.get_roles(user)
    if not ("System Manager" in user_roles):
        if frappe.db.exists("Attendance",{'attendance_date':doc.ot_date,'employee':doc.employee,'docstatus':1}):
            att = frappe.get_doc("Attendance",{'attendance_date':doc.ot_date,'employee':doc.employee,'docstatus':1},["*"])
            if att.shift_status not in ["OD","ODW","ODH"]:
                frappe.throw(_('Attendance Closed for this day %s. For additional details kindly contact the HR Team'%(doc.ot_date)))

@frappe.whitelist()
def make_old_iym_sheet():
    args = frappe.local.form_dict
    filename = args.name
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    doc = frappe.get_doc("RM Input",args.name)
    if doc:
        ws.append(["Customer","Grade","STD Old","STD Impact","STD New",">100 Old",">100 Impact",">100 New","<100 Old","<100 Impact","<100 New"])
        for i in doc.old_iym_settings:
            ws.append([i.customer,i.grade,i.std_old,i.std_impact,i.std_new,i.old1,i.impact1,i.new1,i.old2,i.impact2,i.new2])
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary' 	
 
 
@frappe.whitelist()
def make_old_re_sheet():
    args = frappe.local.form_dict
    filename = args.name
    test = build_xlsx_response_re(filename)

def make_xlsx_file(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    doc = frappe.get_doc("RM Input",args.name)
    if doc:
        ws.append(["Customer","Grade","Strip >100 Old","Strip >100 Impact","Strip >100 New","Strip <100 Old","Strip <100 Impact","Strip <100 New","Coil >100 Old","Coil >100 Impact","Coil >100 New","Coil <100 Old","Coil <100 Impact","Coil <100 New"])
        for i in doc.old_re_settings:
            ws.append([i.customer,i.grade,i.old1,i.impact1,i.new1,i.old2,i.impact2,i.new2,i.coil_old,i.coil_impact,i.coil_new,i.coil_old1,i.coil_impact1,i.coil_new1])
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response_re(filename):
    xlsx_file = make_xlsx_file(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary' 	

def get_live_stock():
    mat_no = '20000610'
    url = "http://apioso.thaisummit.co.th:10401/api/GetItemInventory"
    payload = json.dumps({
        "ItemCode": mat_no,
    })
    headers = {
        'Content-Type': 'application/json',
        'API_KEY': '/1^i[#fhSSDnC8mHNTbg;h^uR7uZe#ninearin!g9D:pos+&terpTpdaJ$|7/QYups;==~w~!AWwb&DU',
    }
    response = requests.request(
        "POST", url, headers=headers, data=payload)
    stock = 0
    if response:
        stocks = json.loads(response.text)
        if stocks:
            ica = frappe.db.sql(
                "select warehouse from `tabInventory Control Area` where iym = 'Y' ", as_dict=True)

            wh_list = [d['warehouse'] for d in ica if 'warehouse' in d]

            df = pd.DataFrame(stocks)
            df = df[df['Warehouse'].isin(wh_list)]
            stock = pd.to_numeric(df["Qty"]).sum()
        print(stock or 0)

@frappe.whitelist()
def enqueue_checkin_bulk_upload_csv(filename):
    frappe.enqueue(
        checkin_bulk_upload_csv, # python function or a module path as string
        queue="long", # one of short, default, long
        timeout=36000, # pass timeout manually
        is_async=True, # if this is True, method is run in worker
        now=False, # if this is True, method is run directly (not in a worker) 
        job_name='Checkin Upload', # specify a job name
        enqueue_after_commit=False, # enqueue the job after the database commit is done at the end of the request
        filename=filename, # kwargs are passed to the method as arguments
    )   

@frappe.whitelist() 
def checkin_bulk_upload_csv():
    frappe.errprint("HI")
    from frappe.utils.file_manager import get_file
    _file = frappe.get_doc("File", {"file_name": "Book8.csv"})
    filepath = get_file("Book8.csv")
    pps = read_csv_content(filepath[1])
    for pp in pps:
        # print(pp[6])
        if not frappe.db.exists('Employee Checkin',{'biometric_pin':pp[0],'time':pp[1]}):
            if frappe.db.exists('Employee',{'biometric_pin':pp[0]}):
                print('Employee Checkin')
                ec = frappe.new_doc('Employee Checkin')
                ec.biometric_pin = pp[0]
                ec.employee = frappe.db.get_value('Employee',{'biometric_pin':pp[0]},['employee_number'])
                ec.time = pp[1]

                ec.device_id = pp[3]
                ec.log_type = pp[2]
                ec.save(ignore_permissions=True)
                frappe.db.commit()
            else:
                if not frappe.db.exists('Unregistered Employee Checkin',{'biometric_pin':pp[0],'time':pp[1]}):
                    print('Unregistered Checkin')
                    ec = frappe.new_doc('Unregistered Employee Checkin')
                    ec.biometric_pin = pp[0]
                    ec.biometric_time = pp[1]
                    ec.locationdevice_id = pp[3]
                    ec.log_type = pp[2]
                    ec.save(ignore_permissions=True)
                    frappe.db.commit()    
    return 'ok'    

@frappe.whitelist() 
def cost_cen(filename):
    frappe.errprint("HI")
    from frappe.utils.file_manager import get_file
    _file = frappe.get_doc("File", {"file_name": filename})
    filepath = get_file(filename)
    pps = read_csv_content(filepath[1])
    for pp in pps:
        print(pp[0])
        if frappe.db.exists('Employee',{'name':pp[0]}):
            ec = frappe.get_doc('Employee',{'name':pp[0]})
            ec.department = pp[1]
            ec.cost_centre = pp[2]
            ec.save(ignore_permissions=True)
            frappe.db.commit()
    return 'ok'    


@frappe.whitelist()
def inactive_employee(doc,method):
    if doc.status=="Active":
        if doc.relieving_date:
            throw(_("Please remove the relieving date for the Active Employee."))
    

@frappe.whitelist()
def bulk_update_from_csv1(filename):
    from frappe.utils.file_manager import get_file
    _file = frappe.get_doc("File", {"file_name": filename})
    filepath = get_file(filename)
    pps = read_csv_content(filepath[1])
    
    for pp in pps:
        print(pp[0], pp[1], pp[7], pp[5])
        if pp[0] != "Employee ID":
            print("HI")
            if not frappe.db.exists('QR Checkin', {'employee': pp[0], 'shift_date': pp[7], 'qr_shift': pp[5]}):
                print("HI")
                emp = frappe.db.get_value('Employee', {'status': 'Active', 'employee': pp[0]}, ['name', 'basic', 'ctc', 'employee_type'])
                
                if not emp:
                    print(f"No active employee found for ID {pp[0]}")
                    continue  # Skip to the next iteration if no employee is found
                
                start_date, end_date = frappe.db.get_value('Payroll Dates', {'name': 'PAYROLL OT PERIOD DATE 0001'}, ['payroll_start_date', 'payroll_end_date'])
                holidays = len(get_holiday_dates_for_employee(emp[0], start_date, end_date))
                total_working_days = (int(date_diff(end_date, start_date)) + 1) - holidays
                
                per_day_basic = emp[1] / total_working_days if emp[1] else 0
                per_day_ctc = emp[2] / total_working_days if emp[2] else 0
                
                if emp[3] == 'CL':
                    per_day_basic = emp[1] if emp[1] else 0
                    per_day_ctc = emp[2] if emp[2] else 0
                
                qr_checkin = frappe.new_doc('QR Checkin')
                qr_checkin.employee = pp[0]
                qr_checkin.employee_name = pp[1]
                qr_checkin.department = pp[2]
                qr_checkin.employee_type = pp[3]
                qr_checkin.contractor = pp[4]
                qr_checkin.basic = emp[1] if emp[1] else 0
                qr_checkin.ctc = emp[2] if emp[2] else 0
                qr_checkin.per_day_basic = per_day_basic
                qr_checkin.per_day_ctc = per_day_ctc
                qr_checkin.created_date = pp[6]
                qr_checkin.shift_date = pp[7]
                qr_checkin.qr_scan_time = pp[9]
                qr_checkin.qr_shift = pp[5]
                qr_checkin.ot = 0
                
                try:
                    qr_checkin.save()
                    frappe.db.commit()
                    print("Created")
                except Exception as e:
                    print(f"Failed to insert QR Checkin for employee {pp[0]}: {e}")

@frappe.whitelist()
def update_salary_in_qr():
    qr_checkins = frappe.db.sql("""
        SELECT * FROM `tabQR Checkin` 
        WHERE shift_date BETWEEN '2024-05-18' AND '2024-05-25'
    """, as_dict=True)

    for qr_checkin in qr_checkins:
        print(qr_checkin.employee)
        emp = frappe.db.get_value(
            'Employee', 
            {'name': qr_checkin.employee}, 
            ['name', 'basic', 'ctc', 'employee_type']
        )
        total_working_days = 24
        per_day_basic = 0
        per_day_ctc = 0
        per_day_basic = emp[1] / total_working_days if emp[1] else 0
        per_day_ctc = emp[2] / total_working_days if emp[2] else 0

        if emp[3] == 'CL':
            per_day_basic = emp[1] if emp[1] else 0
            per_day_ctc = emp[2] if emp[2] else 0

        print(qr_checkin.name)
        frappe.db.set_value("QR Checkin",qr_checkin.name,"basic",emp[1] if emp[1] else 0)
        frappe.db.set_value("QR Checkin",qr_checkin.name,"ctc",emp[2] if emp[2] else 0)
        frappe.db.set_value("QR Checkin",qr_checkin.name,"per_day_basic",per_day_basic)
        frappe.db.set_value("QR Checkin",qr_checkin.name,"per_day_ctc",per_day_ctc)
        print(f"Successfully updated QR Checkin for employee {qr_checkin.employee}")

@frappe.whitelist()
def update_leave_approver():
    value = frappe.get_all("Department",["hod","name"])
    for i in value:
        employee = frappe.get_all("Employee",{"status":"Active","department":i.name},["name"])
        for emp in employee:
            frappe.db.set_value("Employee",emp.name,"leave_approver",i.hod)

@frappe.whitelist()
def update_attendance():
    ot = frappe.db.sql("""update `tabAttendance` set out_time = NULL where name = "HR-ATT-2024-382011" """,as_dict = True)


@frappe.whitelist()
def round_up_to_pf_esi(doc,method):	
    amo = 0
    for detail in doc.get('earnings'):
        if detail.do_not_include_in_total==0:
            amo += detail.amount
        diff = detail.amount - int(detail.amount)
        if diff >= 0.50:
            amount = math.ceil(detail.amount)
        else:
            amount = math.floor(detail.amount)
        detail.amount = amount
        value = amo-int(amo)
        if value >=0.5:
            doc.gross_pay = math.ceil(amo)
        else:
            doc.gross_pay = math.floor(amo)
      	
    for detail in doc.get('deductions'):
        if detail.salary_component == "Employee State Insurance" or "Canteen Charges":
            diff = detail.amount - int(detail.amount)
            if diff >= 0.50:
                amount = math.ceil(detail.amount)
            else:
                amount = math.floor(detail.amount)
            if detail.do_not_include_in_total==0:
                detail.amount = amount
    total_deduction_diff = doc.total_deduction - int(doc.total_deduction)
    if total_deduction_diff >= 0.50:
        doc.total_deduction = math.ceil(doc.total_deduction)
    else:
        doc.total_deduction = math.floor(doc.total_deduction)
    
    doc.net_pay = doc.gross_pay - doc.total_deduction


